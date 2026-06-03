#!/usr/bin/env python3
"""
Builds multiyear_funding_stack.xlsx -- the 5-10 year funding-stack model for the
PublicLogic federal funding architecture.

Design rules (per the task's non-negotiables):
  * Every total is a LIVE FORMULA (SUM / cell math). No hardcoded totals.
  * All assumptions live in labeled input cells on the 'Inputs' sheet and are
    referenced by name-equivalent cell refs from the example sheets.
  * Zero formula errors (verified by a load-and-scan check in build_check.py).
  * Two worked examples:
      A) a small/rural MA town running the full Discover->Continue cycle on a
         realistic multi-year braided stack, with the own-source handoff.
      B) a capital/environmental deal: NMTC + Section 45Q + SRF/Brownfields,
         with the multi-year compliance tail.
  * A PublicLogic revenue sheet maps the fixed-fee / retainer / compliance-tail
    structure to allowable cost lines (2 CFR 200.459) across the stack.

Not legal/accounting advice. Allowability is "subject to the funding source, the
approved budget, and the governing cost principles." Dollar figures are
illustrative planning assumptions, not quotes or commitments.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ---- styling helpers ----------------------------------------------------------
HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
SUB = Font(bold=True, size=11, color="1F4E5F")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow = editable input
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")   # green = live total
NOTE = Font(italic=True, size=9, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'
PCT = '0%'


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=14, color="1F4E5F")
    if sub:
        ws["A2"] = sub
        ws["A2"].font = NOTE


# ==============================================================================
wb = openpyxl.Workbook()

# ---------------------------------------------------------------- README -------
ws = wb.active
ws.title = "README"
title(ws, "PublicLogic Multi-Year Federal Funding Stack",
      "Live-formula model. Yellow cells = editable assumptions (Inputs sheet). "
      "Green cells = live totals. Date verified: 2026-06-03. Not legal/accounting advice.")
readme = [
    "",
    "PURPOSE",
    "  Show how a PublicLogic client sustains the Discover -> Honor -> Understand -> Improve ->",
    "  Build -> Steward -> Continue cycle across 5-10 year budget cycles by BRAIDING federal,",
    "  state, tax-credit, and own-source money -- and how PublicLogic's non-contingent fee rides",
    "  inside money the client already controls (grant-admin / project soft cost / compliance",
    "  infrastructure) under 2 CFR 200.459.",
    "",
    "SHEETS",
    "  Inputs            - every assumption (rates, fees, escalators, award sizes). EDIT HERE.",
    "  ExampleA_RuralTown- small/rural MA town, full cycle, braided multi-year stack + own-source handoff.",
    "  ExampleB_CapitalDeal - NMTC + Section 45Q + SRF/Brownfields capital/environmental deal w/ compliance tail.",
    "  PublicLogic_Revenue - PublicLogic's own revenue architecture across both examples; indirect-rate view.",
    "",
    "CONVENTIONS",
    "  * No total is hardcoded -- every total is =SUM(...) or cell arithmetic.",
    "  * 'Bedrock' = statutory/formula/own-source/tax-credit (durable across administrations).",
    "  * 'Volatile' = competitive discretionary (BRIC/SLCGP-class -- can vanish; see volatility register).",
    "  * Cost-treatment route per line: ADMIN / SOFT / COMPLIANCE.",
    "",
    "AUTHORIZATION vs APPROPRIATION vs ALLOCATION vs OBLIGATION (why statuses differ)",
    "  Authorization = a statute permits a program to exist.",
    "  Appropriation = Congress provides actual budget authority (annual or advance).",
    "  Allocation    = the agency distributes that authority (e.g., formula to a state).",
    "  Obligation    = a binding award/commitment to a specific recipient.",
    "  DEA, BRIC, and SLCGP each looked 'alive' (authorized) while failing at appropriation or",
    "  obligation. This model weights bedrock so the stack survives if the volatile layer is removed.",
]
r = 4
for line in readme:
    ws.cell(row=r, column=1, value=line)
    if line.strip() in ("PURPOSE", "SHEETS", "CONVENTIONS",
                        "AUTHORIZATION vs APPROPRIATION vs ALLOCATION vs OBLIGATION (why statuses differ)"):
        ws.cell(row=r, column=1).font = SUB
    r += 1
ws.column_dimensions["A"].width = 100

# ---------------------------------------------------------------- INPUTS -------
wi = wb.create_sheet("Inputs")
title(wi, "Inputs -- editable assumptions",
      "Yellow cells feed every formula in the example sheets. Figures are illustrative planning "
      "assumptions, not quotes. Verify program terms against current NOFOs before use.")
wi.append([])
headers = ["Key", "Assumption", "Value", "Unit", "Source / note"]
wi.append(headers)
style_header(wi, wi.max_row, len(headers))

# (key, label, value, unit, note) -- value cells get INPUT_FILL
inputs = [
    ("infl",      "Annual escalator (fees & O&M)",            0.03, "/yr",  "planning assumption"),
    ("map_fee",   "PublicLogic Stewardship Map fee (entry)",  35000, "$",   "fixed-fee, non-contingent on award (2 CFR 200.459)"),
    ("narr_fee",  "Grant-narrative / Understand engagement",  20000, "$",   "fixed-fee soft cost"),
    ("build_pct", "PublicLogic compliance-infra on capital",  0.04, "of capital", "soft cost / 200.459 prof service; subject to approved budget"),
    ("ret_fee",   "Annual stewardship retainer (Steward/Continue)", 24000, "$/yr", "own-source operating line once federal ends"),
    ("deminimis", "Default de minimis indirect rate",         0.15, "of MTDC", "2 CFR 200.414(f), 15% effective for awards on/after 10/1/2024"),
    ("single_aud","Single Audit threshold",                  1000000, "$",   "2 CFR 200.501, raised from $750k effective FY beginning on/after 10/1/2024"),
    ("mtdc_sub",  "MTDC per-subaward inclusion cap",          50000, "$",   "2 CFR 200.1, raised from $25k in 2024 revision"),
    # Example A program sizes
    ("ccit",      "A: MA Community Compact IT grant (Y1)",    150000, "$",   "FY2026 cap $200k, no match; verified 2026-06-03"),
    ("cpg",       "A: MA Community Planning Grant (Y2)",       50000, "$",   "Community One Stop; state, low-cost planning entry"),
    ("massworks", "A: MassWorks capital award (Y3)",          1000000, "$",   "MA One Stop infrastructure (state bedrock)"),
    ("cwsrf_a",   "A: CWSRF loan principal (Y3)",             2000000, "$",   "AL 66.458; project loan, 20yr"),
    ("cwsrf_fgv", "A: CWSRF principal forgiveness share",      0.20, "of loan","disadvantaged-community subsidy (illustrative)"),
    ("slfrf_a",   "A: ARPA SLFRF remaining to expend (Y1-2)",  80000, "$",   "must expend by 12/31/2026; closing"),
    ("townpop",   "A: town population",                        4200, "ppl", "small/rural MA town"),
    # Example B program sizes
    ("qei",       "B: NMTC Qualified Equity Investment",      10000000, "$",  "AL 21.020; NMTC made permanent under OBBBA"),
    ("nmtc_rate", "B: NMTC credit rate (7-yr total)",          0.39, "of QEI","5%/yr yrs1-3, 6%/yr yrs4-7 = 39%"),
    ("nmtc_net",  "B: net subsidy to project (after costs)",   0.20, "of QEI","typical net benefit after CDE fees/leverage"),
    ("bf_grant",  "B: EPA Brownfields cleanup grant",          500000, "$",   "AL 66.818; 20% cost share (hardship-waivable)"),
    ("bf_share",  "B: Brownfields non-federal cost share",     0.20, "of grant","CERCLA 104(k)"),
    ("cwsrf_b",   "B: CWSRF loan (water-quality remediation)", 1500000, "$",  "AL 66.458 contaminated-sites use"),
    ("co2_tons",  "B: annual CO2 captured (metric tons)",       8000, "t/yr","carbon capture component"),
    ("q45_rate",  "B: Section 45Q rate (utilization/EOR)",        85, "$/ton","OBBBA raised to $85/ton, placed-in-service after 7/4/2025"),
    ("q45_years", "B: 45Q credit period",                          12, "yrs","statutory 12-yr period"),
    ("directpay", "B: elective (direct) pay share of 45Q",       1.00, "of credit","tax-exempt sponsor monetizes via IRC 6417"),
    # Grant-writing-hours engagement (org working with PublicLogic)
    ("gw_prin",   "GW: Principal (Boudreau) hourly rate",         185, "$/hr","governance/systems; MPA/MCPPO"),
    ("gw_anal",   "GW: Analyst hourly rate",                       110, "$/hr","grant analyst / budget build"),
    ("gw_sme",    "GW: Behavioral/org SME (Rothschild) rate",      165, "$/hr","sustainability/behavioral-systems narrative"),
    ("gw_fixed",  "GW: agreed fixed fee, writing block (non-contingent)", 12500, "$","fixed-fee, owed regardless of award (2 CFR 200.459 bars contingent fees)"),
    ("gw_winprob","GW: client's expected win probability",        0.45, "prob","planning assumption only; fee is NOT contingent on it"),
    ("gw_award",  "GW: target award size (if won)",            750000, "$","illustrative federal capital award being pursued"),
]
start_row = wi.max_row + 1
keymap = {}  # key -> cell ref of the Value column
for i, (key, label, val, unit, note) in enumerate(inputs):
    rr = start_row + i
    wi.cell(row=rr, column=1, value=key).font = Font(bold=True, color="888888", size=9)
    wi.cell(row=rr, column=2, value=label)
    vc = wi.cell(row=rr, column=3, value=val)
    vc.fill = INPUT_FILL
    vc.border = BORDER
    if isinstance(val, float) and val <= 1:
        vc.number_format = PCT
    elif isinstance(val, (int, float)) and val >= 1000:
        vc.number_format = MONEY
    wi.cell(row=rr, column=4, value=unit)
    wi.cell(row=rr, column=5, value=note).font = NOTE
    keymap[key] = f"Inputs!$C${rr}"
for w, col in zip([10, 42, 14, 12, 60], "ABCDE"):
    wi.column_dimensions[col].width = w


def K(key):
    return keymap[key]


# ---------------------------------------------------- EXAMPLE A: RURAL TOWN -----
wa = wb.create_sheet("ExampleA_RuralTown")
title(wa, "Example A -- Small/Rural MA Town: full cycle on a braided multi-year stack",
      "Cycle stage -> source -> cost-treatment route -> dollars by year. Totals are live formulas. "
      "Flywheel: the cheap Map produces the diagnosis AND the grant narrative; the narrative opens "
      "the capital that funds the Build; the Build budget carries the stewardship cost forward.")
# year columns Y1..Y7
YEARS_A = 7
col0 = 4  # first year column (D)
hdr_row = 4
wa.cell(row=hdr_row, column=1, value="Cycle stage")
wa.cell(row=hdr_row, column=2, value="Funding source (layer)")
wa.cell(row=hdr_row, column=3, value="Cost route")
for y in range(YEARS_A):
    wa.cell(row=hdr_row, column=col0 + y, value=f"Y{y+1}")
wa.cell(row=hdr_row, column=col0 + YEARS_A, value="Row total")
style_header(wa, hdr_row, 3 + YEARS_A + 1)

# rows: (stage, source, layer-tag, route, {year_index(0-based): formula_or_value})
def f_infl(base_key, yr):
    # escalate a fee input by infl over yr years: base*(1+infl)^yr
    return f"={K(base_key)}*(1+{K('infl')})^{yr}"

rows_a = [
    ("Discover", "MA Community Compact IT grant [state bedrock]", "SOFT",
        {0: f"={K('ccit')}"}),
    ("Discover", "  ^ PublicLogic Stewardship Map (soft cost in the grant)", "SOFT",
        {0: f"=-{K('map_fee')}"}),
    ("Honor/Understand", "MA Community Planning Grant [state bedrock]", "SOFT",
        {1: f"={K('cpg')}"}),
    ("Honor/Understand", "  ^ PublicLogic narrative/Understand (soft cost)", "SOFT",
        {1: f"=-{f_infl('narr_fee',1)[1:]}"}),
    ("Improve/Build", "MassWorks capital [state bedrock]", "SOFT",
        {2: f"={K('massworks')}"}),
    ("Improve/Build", "CWSRF loan, net of principal forgiveness [federal bedrock]", "SOFT",
        {2: f"={K('cwsrf_a')}*(1-{K('cwsrf_fgv')})"}),
    ("Improve/Build", "  ^ PublicLogic compliance-infra (soft cost on capital)", "COMPLIANCE",
        {2: f"=-({K('massworks')}+{K('cwsrf_a')})*{K('build_pct')}"}),
    ("Steward", "ARPA SLFRF remaining (expend by 12/31/2026) [closing]", "ADMIN",
        {0: f"={K('slfrf_a')}/2", 1: f"={K('slfrf_a')}/2"}),
    ("Steward/Continue", "Own-source stewardship retainer [own-source bedrock]", "ADMIN",
        {3: f"=-{f_infl('ret_fee',3)[1:]}", 4: f"=-{f_infl('ret_fee',4)[1:]}",
         5: f"=-{f_infl('ret_fee',5)[1:]}", 6: f"=-{f_infl('ret_fee',6)[1:]}"}),
]
rr = hdr_row + 1
first_data = rr
for stage, source, route, yvals in rows_a:
    wa.cell(row=rr, column=1, value=stage)
    wa.cell(row=rr, column=2, value=source).alignment = WRAP
    wa.cell(row=rr, column=3, value=route).alignment = CENTER
    for y in range(YEARS_A):
        c = wa.cell(row=rr, column=col0 + y)
        if y in yvals:
            c.value = yvals[y]
        c.number_format = MONEY
        c.border = BORDER
    # row total = live SUM across the year cells
    tot = wa.cell(row=rr, column=col0 + YEARS_A)
    first = get_column_letter(col0)
    last = get_column_letter(col0 + YEARS_A - 1)
    tot.value = f"=SUM({first}{rr}:{last}{rr})"
    tot.number_format = MONEY
    tot.fill = TOTAL_FILL
    rr += 1
last_data = rr - 1

# Net inflows row + cumulative
wa.cell(row=rr, column=1, value="NET inflow to town (all layers)").font = SUB
for y in range(YEARS_A):
    col = get_column_letter(col0 + y)
    c = wa.cell(row=rr, column=col0 + y, value=f"=SUM({col}{first_data}:{col}{last_data})")
    c.number_format = MONEY
    c.fill = TOTAL_FILL
    c.font = Font(bold=True)
gt = wa.cell(row=rr, column=col0 + YEARS_A)
gt.value = f"=SUM({get_column_letter(col0)}{rr}:{get_column_letter(col0+YEARS_A-1)}{rr})"
gt.number_format = MONEY
gt.fill = TOTAL_FILL
gt.font = Font(bold=True)
net_row = rr
rr += 1
# cumulative
wa.cell(row=rr, column=1, value="CUMULATIVE position").font = SUB
for y in range(YEARS_A):
    col = get_column_letter(col0 + y)
    if y == 0:
        c = wa.cell(row=rr, column=col0 + y, value=f"={col}{net_row}")
    else:
        prev = get_column_letter(col0 + y - 1)
        c = wa.cell(row=rr, column=col0 + y, value=f"={prev}{rr}+{col}{net_row}")
    c.number_format = MONEY
    c.font = Font(bold=True, color="1F4E5F")
rr += 2

# PublicLogic revenue from Example A (positive = PL fee)
wa.cell(row=rr, column=1, value="PublicLogic revenue from this client (Example A)").font = SUB
rr += 1
pl_rows_a = [
    ("Map (Discover)", {0: f"={K('map_fee')}"}),
    ("Narrative (Understand)", {1: f"={f_infl('narr_fee',1)[1:]}"}),
    ("Compliance-infra on capital (Build)", {2: f"=({K('massworks')}+{K('cwsrf_a')})*{K('build_pct')}"}),
    ("Retainer (Steward/Continue)", {3: f"={f_infl('ret_fee',3)[1:]}", 4: f"={f_infl('ret_fee',4)[1:]}",
                                     5: f"={f_infl('ret_fee',5)[1:]}", 6: f"={f_infl('ret_fee',6)[1:]}"}),
]
pl_first = rr
for label, yvals in pl_rows_a:
    wa.cell(row=rr, column=2, value=label).alignment = WRAP
    for y in range(YEARS_A):
        c = wa.cell(row=rr, column=col0 + y)
        if y in yvals:
            c.value = yvals[y]
        c.number_format = MONEY
    t = wa.cell(row=rr, column=col0 + YEARS_A,
                value=f"=SUM({get_column_letter(col0)}{rr}:{get_column_letter(col0+YEARS_A-1)}{rr})")
    t.number_format = MONEY
    t.fill = TOTAL_FILL
    rr += 1
wa.cell(row=rr, column=2, value="PublicLogic TOTAL (Example A)").font = SUB
for y in range(YEARS_A):
    col = get_column_letter(col0 + y)
    c = wa.cell(row=rr, column=col0 + y, value=f"=SUM({col}{pl_first}:{col}{rr-1})")
    c.number_format = MONEY
    c.fill = TOTAL_FILL
    c.font = Font(bold=True)
grand = wa.cell(row=rr, column=col0 + YEARS_A,
                value=f"=SUM({get_column_letter(col0)}{rr}:{get_column_letter(col0+YEARS_A-1)}{rr})")
grand.number_format = MONEY
grand.fill = TOTAL_FILL
grand.font = Font(bold=True)

wa.column_dimensions["A"].width = 20
wa.column_dimensions["B"].width = 46
wa.column_dimensions["C"].width = 11
for y in range(YEARS_A + 1):
    wa.column_dimensions[get_column_letter(col0 + y)].width = 13

# ---------------------------------------------------- EXAMPLE B: CAPITAL DEAL ---
wb2 = wb.create_sheet("ExampleB_CapitalDeal")
title(wb2, "Example B -- Capital/Environmental deal: NMTC + Section 45Q + SRF/Brownfields, with compliance tail",
      "Brownfield redevelopment with a carbon-capture component, sponsored by a tax-exempt entity. "
      "Tax credits are treated as federal funding. Compliance tail = 7-yr NMTC + 12-yr 45Q -> recurring stewardship.")
YEARS_B = 8
hdr_row = 4
wb2.cell(row=hdr_row, column=1, value="Component")
wb2.cell(row=hdr_row, column=2, value="Mechanism (layer)")
wb2.cell(row=hdr_row, column=3, value="Cost route")
for y in range(YEARS_B):
    wb2.cell(row=hdr_row, column=col0 + y, value=f"Y{y+1}")
wb2.cell(row=hdr_row, column=col0 + YEARS_B, value="Row total")
style_header(wb2, hdr_row, 3 + YEARS_B + 1)

# 45Q annual credit = tons * rate, claimed yrs 2..(1+credit period) but cap display window to 8 yrs
def q45_year(yr_index):
    # credit begins Y2 (placed in service end of Y1); within 12-yr period
    return f"=IF(AND({yr_index+1}>=2,{yr_index+1}<2+{K('q45_years')}),{K('co2_tons')}*{K('q45_rate')}*{K('directpay')},0)"

rows_b = [
    ("Site cleanup", "EPA Brownfields cleanup grant [federal competitive]", "SOFT",
        {0: f"={K('bf_grant')}"}),
    ("Site cleanup", "  ^ non-federal cost share (own-source)", "SOFT",
        {0: f"=-{K('bf_grant')}*{K('bf_share')}"}),
    ("Water remediation", "CWSRF loan (contaminated-site water quality) [federal bedrock]", "SOFT",
        {1: f"={K('cwsrf_b')}"}),
    ("Capital stack", "NMTC net subsidy to project [federal bedrock - permanent]", "COMPLIANCE",
        {1: f"={K('qei')}*{K('nmtc_net')}"}),
    ("Clean energy", "Section 45Q via elective pay (annual, 12-yr) [statutory bedrock]", "COMPLIANCE",
        {i: q45_year(i) for i in range(YEARS_B)}),
    ("PublicLogic", "  ^ compliance-infra at financial close (Build)", "COMPLIANCE",
        {1: f"=-({K('qei')}+{K('cwsrf_b')})*{K('build_pct')}"}),
    ("PublicLogic", "  ^ NMTC 7-yr compliance tail (CDE reporting)", "COMPLIANCE",
        {i: (f"=-{K('ret_fee')}*(1+{K('infl')})^{i}" if 1 <= i <= 7 else None) for i in range(YEARS_B)}),
]
rr = hdr_row + 1
first_data = rr
for comp, mech, route, yvals in rows_b:
    wb2.cell(row=rr, column=1, value=comp)
    wb2.cell(row=rr, column=2, value=mech).alignment = WRAP
    wb2.cell(row=rr, column=3, value=route).alignment = CENTER
    for y in range(YEARS_B):
        c = wb2.cell(row=rr, column=col0 + y)
        v = yvals.get(y)
        if v is not None:
            c.value = v
        c.number_format = MONEY
        c.border = BORDER
    tot = wb2.cell(row=rr, column=col0 + YEARS_B)
    f = get_column_letter(col0); l = get_column_letter(col0 + YEARS_B - 1)
    tot.value = f"=SUM({f}{rr}:{l}{rr})"
    tot.number_format = MONEY
    tot.fill = TOTAL_FILL
    rr += 1
last_data = rr - 1

wb2.cell(row=rr, column=1, value="NET to deal / sponsor").font = SUB
for y in range(YEARS_B):
    col = get_column_letter(col0 + y)
    c = wb2.cell(row=rr, column=col0 + y, value=f"=SUM({col}{first_data}:{col}{last_data})")
    c.number_format = MONEY; c.fill = TOTAL_FILL; c.font = Font(bold=True)
gt = wb2.cell(row=rr, column=col0 + YEARS_B,
              value=f"=SUM({get_column_letter(col0)}{rr}:{get_column_letter(col0+YEARS_B-1)}{rr})")
gt.number_format = MONEY; gt.fill = TOTAL_FILL; gt.font = Font(bold=True)
net_row_b = rr
rr += 1
wb2.cell(row=rr, column=1, value="CUMULATIVE position").font = SUB
for y in range(YEARS_B):
    col = get_column_letter(col0 + y)
    if y == 0:
        c = wb2.cell(row=rr, column=col0 + y, value=f"={col}{net_row_b}")
    else:
        prev = get_column_letter(col0 + y - 1)
        c = wb2.cell(row=rr, column=col0 + y, value=f"={prev}{rr}+{col}{net_row_b}")
    c.number_format = MONEY; c.font = Font(bold=True, color="1F4E5F")
rr += 2

# PublicLogic revenue from Example B (compliance tail)
wb2.cell(row=rr, column=1, value="PublicLogic revenue from this deal (Example B)").font = SUB
rr += 1
plb_first = rr
plb = [
    ("Compliance-infra at close", {1: f"=({K('qei')}+{K('cwsrf_b')})*{K('build_pct')}"}),
    ("NMTC 7-yr compliance tail", {i: f"={K('ret_fee')}*(1+{K('infl')})^{i}" for i in range(1, 8)}),
]
for label, yvals in plb:
    wb2.cell(row=rr, column=2, value=label).alignment = WRAP
    for y in range(YEARS_B):
        c = wb2.cell(row=rr, column=col0 + y)
        if y in yvals:
            c.value = yvals[y]
        c.number_format = MONEY
    t = wb2.cell(row=rr, column=col0 + YEARS_B,
                 value=f"=SUM({get_column_letter(col0)}{rr}:{get_column_letter(col0+YEARS_B-1)}{rr})")
    t.number_format = MONEY; t.fill = TOTAL_FILL
    rr += 1
wb2.cell(row=rr, column=2, value="PublicLogic TOTAL (Example B)").font = SUB
for y in range(YEARS_B):
    col = get_column_letter(col0 + y)
    c = wb2.cell(row=rr, column=col0 + y, value=f"=SUM({col}{plb_first}:{col}{rr-1})")
    c.number_format = MONEY; c.fill = TOTAL_FILL; c.font = Font(bold=True)
g = wb2.cell(row=rr, column=col0 + YEARS_B,
             value=f"=SUM({get_column_letter(col0)}{rr}:{get_column_letter(col0+YEARS_B-1)}{rr})")
g.number_format = MONEY; g.fill = TOTAL_FILL; g.font = Font(bold=True)

wb2.column_dimensions["A"].width = 18
wb2.column_dimensions["B"].width = 50
wb2.column_dimensions["C"].width = 11
for y in range(YEARS_B + 1):
    wb2.column_dimensions[get_column_letter(col0 + y)].width = 12

# ---------------------------------------------------- PUBLICLOGIC REVENUE ------
wp = wb.create_sheet("PublicLogic_Revenue")
title(wp, "PublicLogic revenue architecture across the stack",
      "Maps the fixed-fee / retainer / compliance-tail structure to allowable cost lines (2 CFR 200.459) "
      "and shows the indirect-rate implication. Pulls live totals from the two example sheets.")
wp.append([]); wp.append([])
rows_p = [
    ["Revenue line", "Cost-treatment route", "Allowability basis", "Recurring?", "Amount"],
]
hr = 4
for j, h in enumerate(rows_p[0], start=1):
    c = wp.cell(row=hr, column=j, value=h);
style_header(wp, hr, 5)
# reference the PublicLogic totals computed in the example sheets (row totals column)
# Example A PL total grand cell is the last grand cell; Example B similarly. We reference by SUM of their PL blocks.
data_p = [
    ("Stewardship Map (entry)", "SOFT", "Prof. service soft cost; 2 CFR 200.459", "No (per turn)",
        f"={K('map_fee')}"),
    ("Grant narrative / Understand", "SOFT", "Soft cost in planning grant; 200.459", "No (per turn)",
        f"={K('narr_fee')}*(1+{K('infl')})"),
    ("Compliance-infra on capital (Ex A+B)", "COMPLIANCE", "Soft cost on approved capital budget; 200.459",
        "Per project",
        f"=(('ExampleA_RuralTown'!{get_column_letter(col0+YEARS_A)}{0+0}))"),  # placeholder, fixed below
    ("Stewardship retainer (Ex A own-source)", "ADMIN", "Own-source operating line; 200.459 once federal ends",
        "Yes (annual)",
        f"={K('ret_fee')}"),
    ("NMTC 7-yr compliance tail (Ex B)", "COMPLIANCE", "CDE compliance-period monitoring; 200.459",
        "Yes (7 yrs)",
        f"={K('ret_fee')}*7"),
]
rp = hr + 1
for line in data_p:
    for j, val in enumerate(line, start=1):
        c = wp.cell(row=rp, column=j, value=val)
        if j == 5:
            c.number_format = MONEY
        if j in (1, 3):
            c.alignment = WRAP
    rp += 1
# fix the compliance-infra amount to a real formula (capital from both examples * build_pct)
wp.cell(row=hr+3, column=5,
        value=f"=({K('massworks')}+{K('cwsrf_a')}+{K('qei')}+{K('cwsrf_b')})*{K('build_pct')}")
wp.cell(row=hr+3, column=5).number_format = MONEY
# total
wp.cell(row=rp, column=4, value="TOTAL recurring + one-time").font = SUB
tc = wp.cell(row=rp, column=5, value=f"=SUM(E{hr+1}:E{rp-1})")
tc.number_format = MONEY; tc.fill = TOTAL_FILL; tc.font = Font(bold=True)
rp += 2
# indirect-rate note block
notes_p = [
    "INDIRECT-RATE IMPLICATIONS",
    "  * With no negotiated rate (NICRA), PublicLogic and its subrecipients may elect the de minimis",
    f"    rate of {0.15:.0%} of Modified Total Direct Cost (2 CFR 200.414(f); 15% effective for awards on/after 10/1/2024).",
    "  * Most PublicLogic fees ride as DIRECT professional-service costs (200.459), not indirect -- so the",
    "    de minimis rate mainly affects the client's own overhead recovery, not PublicLogic's fee allowability.",
    "  * Fees are FIXED-FEE / RETAINER and NON-CONTINGENT on award; that is what makes them allowable as a",
    "    professional-service cost rather than a contingent fee (which 2 CFR 200.459(b) disallows).",
    "  * SUBRECIPIENT vs CONTRACTOR (2 CFR 200.331): PublicLogic is a CONTRACTOR -- it provides goods/services",
    "    within normal business operations to many purchasers, does not carry programmatic decision-making or",
    "    eligibility determination, and is not subject to the federal program's compliance requirements in its",
    "    own right. Therefore PublicLogic is procured under 200.317-327 and is NOT itself Single-Audited.",
    f"  * Single Audit is the CLIENT's obligation only if it expends >= ${1000000:,.0f} in federal awards in a year (200.501).",
    "  * Not legal/accounting advice -- allowability is subject to the funding source, approved budget, and cost principles.",
]
for line in notes_p:
    c = wp.cell(row=rp, column=1, value=line)
    if line.strip() == "INDIRECT-RATE IMPLICATIONS":
        c.font = SUB
    else:
        c.font = NOTE
    rp += 1
for col, w in zip("ABCDE", [42, 16, 40, 14, 16]):
    wp.column_dimensions[col].width = w

# ---------------------------------------------------- GRANT-WRITING HOURS ------
# An org working with PublicLogic buys grant-writing hours. This sheet "shows the
# work": every task line is Hours x Rate = Line cost (live formula), and -- crucially --
# each line carries the CORRECT funding route under the cost principles.
#
# THE RULE (2 CFR 200.460): grant-writing / proposal preparation costs are INDIRECT
# (F&A) costs and may NOT be charged directly to the federal award being sought
# (whether or not it is won). They ride on the org's own-source / unrestricted /
# F&A pool. PublicLogic's writing fee is FIXED and NON-CONTINGENT (2 CFR 200.459(b)
# bars contingent fees) -- it is owed whether or not the grant is awarded.
# Once the grant is WON, grant-ADMINISTRATION hours (a different activity) become an
# allowable charge to that award's admin line. This sheet shows the route flip.
wg = wb.create_sheet("GrantWriting_Hours")
title(wg, "Grant-writing hours -- an org working with PublicLogic (the work, shown line by line)",
      "Hours x rate = line cost (live formulas). Funding route per line follows the cost principles: "
      "writing = proposal cost (2 CFR 200.460, indirect/own-source, non-contingent); post-award = "
      "allowable grant-admin on the won award. Not legal/accounting advice.")
hr = 4
gw_cols = ["Phase", "Task (the work)", "Role", "Hours", "Rate $/hr", "Line cost",
           "Deliverable", "Funding route / allowability"]
for j, h in enumerate(gw_cols, start=1):
    wg.cell(row=hr, column=j, value=h)
style_header(wg, hr, len(gw_cols))

# (phase, task, role_key, hours, deliverable, route)
PROP = "Proposal cost - 2 CFR 200.460: indirect/F&A or org own-source; NON-FederalAward; non-contingent"
ADMIN_WON = "ALLOWABLE on the WON award as grant-administration (200.413 / program admin line)"
gw_rows = [
    ("1. Discover/eligibility", "Review Stewardship Map; confirm Assistance Listing eligibility", "gw_prin", 4, "Eligibility memo", PROP),
    ("1. Discover/eligibility", "SAM.gov / UEI registration check; portal setup", "gw_anal", 3, "Registration confirmed", PROP),
    ("2. Narrative", "Need statement & problem analysis", "gw_prin", 8, "Need statement", PROP),
    ("2. Narrative", "Project design & logic model", "gw_prin", 10, "Project narrative", PROP),
    ("2. Narrative", "Behavioral/organizational sustainability section", "gw_sme", 6, "Sustainability narrative", PROP),
    ("3. Budget", "Budget build (cost principles, MTDC, 15% de minimis)", "gw_anal", 8, "SF-424A budget", PROP),
    ("3. Budget", "Match/cost-share strategy & budget narrative", "gw_prin", 5, "Budget narrative", PROP),
    ("4. Attachments/compliance", "SF-424 family, certifications, reps & certs", "gw_anal", 6, "Forms package", PROP),
    ("4. Attachments/compliance", "Letters of support, MOUs, env-review coordination", "gw_anal", 6, "Attachments", PROP),
    ("5. Review/submit", "Internal review & responsiveness check", "gw_prin", 4, "Reviewed package", PROP),
    ("5. Review/submit", "Portal submission (grants.gov / state e-grants)", "gw_anal", 2, "Submission receipt", PROP),
    ("6. Post-award (if won)", "Award setup, drawdown & subrecipient-monitoring setup", "gw_prin", 8, "Grants-mgmt setup", ADMIN_WON),
    ("6. Post-award (if won)", "Compliance calendar & reporting cadence", "gw_anal", 10, "Compliance plan", ADMIN_WON),
]
rr = hr + 1
write_first = rr
for phase, task, rolekey, hours, deliv, route in gw_rows:
    wg.cell(row=rr, column=1, value=phase)
    wg.cell(row=rr, column=2, value=task).alignment = WRAP
    role_label = {"gw_prin": "Principal", "gw_anal": "Analyst", "gw_sme": "Behav/org SME"}[rolekey]
    wg.cell(row=rr, column=3, value=role_label)
    hc = wg.cell(row=rr, column=4, value=hours); hc.alignment = RIGHT
    rc = wg.cell(row=rr, column=5, value=f"={K(rolekey)}"); rc.number_format = MONEY
    lc = wg.cell(row=rr, column=6, value=f"=D{rr}*E{rr}"); lc.number_format = MONEY  # the work, shown
    wg.cell(row=rr, column=7, value=deliv).alignment = WRAP
    wg.cell(row=rr, column=8, value=route).alignment = WRAP
    for j in range(1, len(gw_cols) + 1):
        wg.cell(row=rr, column=j).border = BORDER
    rr += 1
write_last = rr - 1
# split totals: writing (phases 1-5) vs post-award (phase 6)
post_start = write_first + 11  # first phase-6 row
# Writing subtotal
wg.cell(row=rr, column=2, value="WRITING subtotal (proposal cost - own-source / F&A, non-contingent)").font = SUB
hsum = wg.cell(row=rr, column=4, value=f"=SUM(D{write_first}:D{post_start-1})"); hsum.font = Font(bold=True)
csum = wg.cell(row=rr, column=6, value=f"=SUM(F{write_first}:F{post_start-1})")
csum.number_format = MONEY; csum.fill = TOTAL_FILL; csum.font = Font(bold=True)
writing_tot_row = rr
rr += 1
# Post-award subtotal
wg.cell(row=rr, column=2, value="POST-AWARD subtotal (allowable grant-admin on the WON award)").font = SUB
hsum2 = wg.cell(row=rr, column=4, value=f"=SUM(D{post_start}:D{write_last})"); hsum2.font = Font(bold=True)
csum2 = wg.cell(row=rr, column=6, value=f"=SUM(F{post_start}:F{write_last})")
csum2.number_format = MONEY; csum2.fill = TOTAL_FILL; csum2.font = Font(bold=True)
post_tot_row = rr
rr += 1
# Grand total hours/cost (time & materials)
wg.cell(row=rr, column=2, value="ENGAGEMENT TOTAL (time & materials reference)").font = SUB
ht = wg.cell(row=rr, column=4, value=f"=D{writing_tot_row}+D{post_tot_row}"); ht.font = Font(bold=True)
ct = wg.cell(row=rr, column=6, value=f"=F{writing_tot_row}+F{post_tot_row}")
ct.number_format = MONEY; ct.fill = TOTAL_FILL; ct.font = Font(bold=True)
tm_row = rr
rr += 2

# Fixed-fee reconciliation block (the non-contingent writing block)
wg.cell(row=rr, column=1, value="FIXED-FEE RECONCILIATION (writing block)").font = SUB
rr += 1
recon = [
    ("Agreed fixed fee for the writing block (non-contingent)", f"={K('gw_fixed')}", MONEY),
    ("Time & materials value of writing hours (from above)", f"=F{writing_tot_row}", MONEY),
    ("Fixed-fee variance vs. T&M (org's certainty premium / discount)", f"={K('gw_fixed')}-F{writing_tot_row}", MONEY),
    ("Writing hours (from above)", f"=D{writing_tot_row}", '#,##0.0'),
    ("Effective blended rate on the fixed fee", f"={K('gw_fixed')}/D{writing_tot_row}", '#,##0'),
]
for label, formula, fmt in recon:
    wg.cell(row=rr, column=2, value=label).alignment = WRAP
    c = wg.cell(row=rr, column=6, value=formula); c.number_format = fmt
    rr += 1
rr += 1
# Why non-contingent matters: contingent-fee illustration (what PublicLogic does NOT do)
wg.cell(row=rr, column=1, value="WHY NON-CONTINGENT (what PublicLogic does NOT do)").font = SUB
rr += 1
contra = [
    ("Target award if won (illustrative)", f"={K('gw_award')}", MONEY),
    ("Client's expected win probability (planning only)", f"={K('gw_winprob')}", PCT),
    ("A contingent 'X% of award' fee would be barred by 2 CFR 200.459(b).", None, None),
    ("PublicLogic charges the FIXED writing fee regardless of outcome -> allowable professional service.", None, None),
    ("Expected-value of award to the org (context, not a fee basis)", f"={K('gw_award')}*{K('gw_winprob')}", MONEY),
]
for label, formula, fmt in contra:
    cl = wg.cell(row=rr, column=2, value=label); cl.alignment = WRAP
    if formula is None:
        cl.font = NOTE
    else:
        c = wg.cell(row=rr, column=6, value=formula); c.number_format = fmt
    rr += 1
rr += 1
# Allowability footnotes
notes_g = [
    "FUNDING-ROUTE NOTES (subject to the funding source, approved budget, and cost principles; not legal/accounting advice)",
    "  * 2 CFR 200.460 -- Proposal costs (preparing bids/proposals/applications, successful OR unsuccessful) are normally",
    "    treated as INDIRECT (F&A) costs and allocated to current activities; they are NOT directly chargeable to the",
    "    federal award being sought. The org pays the writing fee from own-source / unrestricted funds or its F&A pool.",
    "  * 2 CFR 200.459(b) -- professional-service costs are allowable; CONTINGENT fees are a factor against allowability.",
    "    PublicLogic's writing fee is fixed and NON-CONTINGENT, which is what keeps it an allowable professional service.",
    "  * 2 CFR 200.458 -- Pre-award costs are allowable only with prior written approval and only if necessary for the",
    "    award; a 90-day-before-award window is typical. Post-award, grant-ADMINISTRATION (a different activity from",
    "    writing) is allowable on the won award (200.413 direct admin / program admin line, e.g., CDBG's planning+admin).",
    "  * Where an org has a capacity-building or planning grant (or foundation/own-source capacity funds), some writing",
    "    support may be an allowable activity of THAT award -- never of the prospective award it is being written for.",
    "  * Hours and rates here are illustrative planning assumptions, not quotes. Edit them on the Inputs sheet.",
]
for line in notes_g:
    c = wg.cell(row=rr, column=1, value=line)
    c.font = SUB if line.startswith("FUNDING-ROUTE NOTES") else NOTE
    rr += 1
for col, w in zip("ABCDEFGH", [22, 46, 14, 8, 10, 12, 22, 50]):
    wg.column_dimensions[col].width = w
wg.freeze_panes = "A5"

# freeze header rows
for s in (wa, wb2):
    s.freeze_panes = "D5"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "multiyear_funding_stack.xlsx")
wb.save(out)
print("wrote", out)
