# Single-source build: read INPUTS from the Capital Workbook (+ Financial Model forecast),
# recompute derived figures, regenerate charts, and render the Operating Plan (DOCX+PDF).
import openpyxl, os
AN="/home/user/puddlejumper/docs/analyses"
CW=f"{AN}/swanzey-capital-forecast/Swanzey Capital Workbook.xlsx"
FM=f"{AN}/swanzey-financial-review/Models/Swanzey Financial Model.xlsx"
ROOT=f"{AN}/swanzey-operating-plan"; VIS=f"{ROOT}/Visuals"; os.makedirs(VIS,exist_ok=True)
VER="v0.2"; TITLE="Swanzey Operating Plan"

def num(c,default=0.0):
    v=c.value
    if isinstance(v,(int,float)): return float(v)
    return float(default)

# ---- read Capital Workbook inputs ----
cw=openpyxl.load_workbook(CW)
A=cw["Assumptions"]; VAL=num(A["B5"],978600000); MED=num(A["B6"],300000); EXIST=num(A["B7"],403000); BSHARE=num(A["B8"],0.80)
P=cw["Projects"]; projects=[]
r=5
while P.cell(r,1).value and str(P.cell(r,1).value).upper()!="TOTAL":
    name=P.cell(r,1).value; tot=num(P.cell(r,2)); aid=num(P.cell(r,3)); srf=num(P.cell(r,4)); gr=num(P.cell(r,5))
    projects.append(dict(name=name,total=tot,aid=aid,srf=srf,gr=gr,local=tot-aid-srf-gr,nontax=aid+srf+gr)); r+=1
L=cw["Reserve Ladder"]; ladder=[]
r=5
while L.cell(r,1).value and str(L.cell(r,1).value)!="Note":
    ladder.append(dict(fund=L.cell(r,1).value,dep=num(L.cell(r,2)),funds=L.cell(r,3).value)); r+=1
TP=cw["10-Year Plan"]; glide=[]
r=5
while isinstance(TP.cell(r,1).value,(int,float)):
    yr=int(TP.cell(r,1).value); deps=[num(TP.cell(r,c)) for c in range(2,9)]; total=sum(deps)
    net=total-EXIST; rate=net/(VAL/1000.0); home=rate*MED/1000.0
    glide.append(dict(yr=yr,total=total,net=net,rate=rate,home=home)); r+=1
tot_backlog=sum(p["total"] for p in projects); tot_local=sum(p["local"] for p in projects); tot_nontax=sum(p["nontax"] for p in projects)

# ---- read Financial Model forecast inputs ----
fm=openpyxl.load_workbook(FM); F=fm["Forecast"]
fbase=num(F["B8"],8939036); g_base=num(F["B5"],0.06); g_low=num(F["B6"],0.045); g_high=num(F["B7"],0.075)
fc=[]
for k in range(0,6):
    y=2026+k
    fc.append(dict(yr=y,low=fbase*(1+g_low)**k,base=fbase*(1+g_base)**k,high=fbase*(1+g_high)**k))

print(f"Backlog ${tot_backlog:,.0f} | non-tax {tot_nontax/tot_backlog:.0%} | local ${tot_local:,.0f}")
print(f"Forecast base 2031 ${fc[-1]['base']:,.0f} | glide peak home ${max(g['home'] for g in glide):.0f}")

# ---- regenerate charts from computed data ----
import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt; import numpy as np
NAVY="#1f3a5f"; TEAL="#0f6e56"; CORAL="#993c1d"; GREY="#5f5e5a"
plt.rcParams.update({"font.size":11,"axes.spines.top":False,"axes.spines.right":False,"axes.titlesize":13,"axes.titleweight":"bold","figure.dpi":150})
# 06 funding stack
names=[p["name"].replace(" ","\n",1) for p in projects]; loc=[p["local"]/1e6 for p in projects]; ntx=[p["nontax"]/1e6 for p in projects]
fig,ax=plt.subplots(figsize=(9.5,5)); x=np.arange(len(projects))
ax.bar(x,loc,color=CORAL,label="Local property-tax share"); ax.bar(x,ntx,bottom=loc,color=TEAL,label="State / SRF-rates / grants")
for i,p in enumerate(projects): ax.text(i,(p["total"])/1e6+0.04,f"${p['total']/1e6:.2f}M",ha="center",fontsize=8.5,fontweight="bold")
ax.set_xticks(x); ax.set_xticklabels(names,fontsize=8); ax.set_ylabel("$ millions")
ax.set_title(f"Capital Funding Stack: {tot_nontax/tot_backlog:.0%} Is Other People's Money")
ax.legend(frameon=False,fontsize=9,loc="upper right")
ax.text(0.02,0.95,f"${tot_backlog/1e6:.2f}M total · ${tot_local/1e6:.2f}M local ({tot_local/tot_backlog:.0%}) · $0 new bonds",transform=ax.transAxes,fontsize=9,color=NAVY,fontweight="bold")
fig.text(0.5,-0.03,"Generated live from the Capital Workbook. Modeled estimates.",ha="center",fontsize=8,color=GREY)
fig.tight_layout(); fig.savefig(f"{VIS}/06_capital-funding-stack.png",bbox_inches="tight"); plt.close()
# 07 tax glide
gy=[g["yr"] for g in glide]; gh=[round(g["home"]) for g in glide]
fig,ax=plt.subplots(figsize=(9.5,4.8)); cols=[TEAL if v>=0 else CORAL for v in gh]
b=ax.bar(gy,gh,color=cols,width=0.6)
for r2,v in zip(b,gh): ax.text(r2.get_x()+r2.get_width()/2,v+(2 if v>=0 else -6),f"${v:+d}",ha="center",fontsize=8.3,fontweight="bold")
ax.axhline(0,color="black",lw=0.8); ax.set_title(f"Net-New Tax Impact on a ${MED/1000:.0f}K Home (from the Workbook)")
ax.set_ylabel("$ per year vs. today"); ax.set_xlabel("Fiscal year")
fig.text(0.5,-0.03,"Generated live from the Capital Workbook 10-Year Plan + Assumptions.",ha="center",fontsize=8,color=GREY)
fig.tight_layout(); fig.savefig(f"{VIS}/07_capital-tax-glide.png",bbox_inches="tight"); plt.close()
print("charts regenerated from workbook")

import json
json.dump(dict(VAL=VAL,MED=MED,backlog=tot_backlog,local=tot_local,nontax=tot_nontax,
   projects=projects,ladder=ladder,glide=glide,forecast=fc,peak_home=max(g['home'] for g in glide)),
   open("/tmp/_model.json","w"))
print("model snapshot written")
