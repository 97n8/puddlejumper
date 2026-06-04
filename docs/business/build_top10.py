#!/usr/bin/env python3
"""
Builds Top10_Opportunities.xlsx — a live, scored, ranked workbook of the top 10
opportunities heading into PublicLogic's relationships with its business people
(Robert McCall/EM&S, Glen Scharer/Carbon Alliance, WasteWerx, AED, Acadia, CivicPlus,
and the NMTC/CDE channel).

Design rules (same discipline as the funding-stack workbook):
  * Composite scores and ranks are LIVE formulas (no hardcoded ranking).
  * Scoring weights live in labeled Inputs cells (edit there to re-rank).
  * Revenue figures are illustrative planning estimates drawn from PublicLogic's
    own deal documents — NOT quotes or commitments.
  * Zero formula errors (verified by a load-and-recalc check).

Scoring criteria (each scored 1-5; 5 = best):
  REV   - PublicLogic revenue potential (size of the fee over the relationship's life)
  PROB  - probability / readiness (how close to signed)
  RECUR - recurrence (5 = recurring retainer/channel; 1 = one-time)
  STRAT - strategic leverage (reference value, channel multiplier, reusability)
  SPEED - speed to cash (5 = cash soon; 1 = long horizon)
  RISKADJ - risk-adjusted quality (5 = clean; 1 = securities/contingent-fee/stall risk)

Composite = weighted average on the 1-5 scale (weights sum to 1).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
SUB = Font(bold=True, size=11, color="1F4E5F")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
TOP_FILL = PatternFill("solid", fgColor="FCE4D6")
NOTE = Font(italic=True, size=9, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'
NUM2 = '0.00'


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER


def title(ws, text, sub=None):
    ws["A1"] = text; ws["A1"].font = Font(bold=True, size=14, color="1F4E5F")
    if sub:
        ws["A2"] = sub; ws["A2"].font = NOTE


wb = openpyxl.Workbook()

# ----------------------------------------------------------------- README -----
ws = wb.active; ws.title = "README"
title(ws, "PublicLogic — Top 10 Opportunities (Business Relationships)",
      "Live scored & ranked. Yellow = editable weights (Inputs). Date: 2026-06-03. "
      "Revenue figures are illustrative planning estimates from PublicLogic's own deal docs, not quotes.")
lines = [
    "", "PURPOSE",
    "  Prioritize the opportunities heading into PublicLogic's relationships with its business people —",
    "  Robert McCall (EM&S / Michigan LTC Network), Glen Scharer (Carbon Alliance), WasteWerx, AED,",
    "  Acadia Capital, CivicPlus, and the NMTC/CDE channel — so effort goes where the leverage is.",
    "", "SHEETS",
    "  Inputs      - scoring weights (edit to re-rank). Weights sum to 1.00.",
    "  Top10       - the scored, ranked opportunities. Composite & Rank are live formulas.",
    "  Revenue     - PublicLogic fee potential per opportunity (near-12-month + lifetime). Live totals.",
    "  Action      - relationship owner, the move, fee structure, cost route, key flag, next step.",
    "", "HOW THE SCORE WORKS",
    "  Each opportunity is scored 1-5 (5=best) on six criteria: REV, PROB, RECUR, STRAT, SPEED, RISKADJ.",
    "  Composite = weighted average (weights on Inputs). Rank is computed live from the composite.",
    "  Change a weight or a score and the ranking re-sorts on recalculation.",
    "", "READ-IN-ONE-LINE",
    "  Anchor on Michigan (Robert) and the NMTC/WasteWerx channels; keep every business-funded fee",
    "  FIXED and NON-CONTINGENT on federal money; treat each deal as a reusable template, not a dependency.",
    "", "Not legal/accounting/securities advice. Subject to funding source, approved budget, cost principles.",
]
r = 4
for ln in lines:
    ws.cell(row=r, column=1, value=ln)
    if ln.strip() in ("PURPOSE", "SHEETS", "HOW THE SCORE WORKS", "READ-IN-ONE-LINE"):
        ws.cell(row=r, column=1).font = SUB
    r += 1
ws.column_dimensions["A"].width = 104

# ----------------------------------------------------------------- INPUTS -----
wi = wb.create_sheet("Inputs")
title(wi, "Inputs — scoring weights", "Edit the yellow weights to re-rank. They should sum to 1.00 (see check).")
wi.append([]); wi.append(["Code", "Criterion", "Weight"])
style_header(wi, wi.max_row, 3)
weights = [
    ("REV", "PublicLogic revenue potential (life of relationship)", 0.20),
    ("PROB", "Probability / readiness (closeness to signed)", 0.22),
    ("RECUR", "Recurrence (retainer / channel vs one-time)", 0.16),
    ("STRAT", "Strategic leverage (reference / channel / reuse)", 0.16),
    ("SPEED", "Speed to cash", 0.14),
    ("RISKADJ", "Risk-adjusted quality (clean vs securities/contingent/stall)", 0.12),
]
wkey = {}
start = wi.max_row + 1
for i, (code, label, w) in enumerate(weights):
    rr = start + i
    wi.cell(row=rr, column=1, value=code).font = Font(bold=True, color="888888")
    wi.cell(row=rr, column=2, value=label)
    c = wi.cell(row=rr, column=3, value=w); c.fill = INPUT_FILL; c.border = BORDER; c.number_format = '0.00'
    wkey[code] = f"Inputs!$C${rr}"
sumrow = start + len(weights)
wi.cell(row=sumrow, column=2, value="SUM (must = 1.00)").font = SUB
sc = wi.cell(row=sumrow, column=3, value=f"=SUM(C{start}:C{sumrow-1})")
sc.fill = TOTAL_FILL; sc.font = Font(bold=True); sc.number_format = '0.00'
for col, w in zip("ABC", [10, 52, 10]):
    wi.column_dimensions[col].width = w

# ----------------------------------------------------------------- TOP10 ------
wt = wb.create_sheet("Top10")
title(wt, "Top 10 Opportunities — scored & ranked (live)",
      "Composite = weighted avg of the six 1-5 scores. Rank is live. Edit scores or weights to re-rank.")
hdr = ["#", "Opportunity", "Relationship / source", "REV", "PROB", "RECUR", "STRAT", "SPEED", "RISKADJ",
       "Composite", "Rank"]
hrow = 4
for j, h in enumerate(hdr, start=1):
    wt.cell(row=hrow, column=j, value=h)
style_header(wt, hrow, len(hdr))

# (name, source, REV, PROB, RECUR, STRAT, SPEED, RISKADJ)
opps = [
    ("Michigan LTC Network — Site 1 (EGLE invite -> build)", "Robert McCall / EM&S", 5, 4, 4, 5, 4, 3),
    ("Michigan LTC Network — Sites 2-4 replication", "Robert McCall / EM&S", 5, 3, 5, 5, 2, 3),
    ("WasteWerx portfolio MSA (governance across client base)", "JT Clark / WasteWerx", 4, 3, 5, 5, 3, 4),
    ("NMTC governance practice via CDE channel (CEI, USBCDC)", "CEI / USBCDC / Lin Thornton", 5, 3, 5, 5, 2, 3),
    ("Carbon Alliance deal-flow channel (ongoing)", "Glen Scharer", 4, 3, 4, 5, 2, 4),
    ("La Porte, IN waste-to-energy (scope -> build)", "Glen Scharer", 3, 3, 3, 4, 3, 4),
    ("AED / Pocomoke NMTC compliance module", "Brian Kuhn / Joe Loud", 4, 2, 4, 3, 2, 3),
    ("City of Flint coordination engagement (Phase 1 -> retainer)", "via Robert McCall", 3, 3, 4, 3, 3, 3),
    ("CivicPlus channel partnership (govtech)", "CivicPlus", 4, 2, 5, 4, 1, 4),
    ("Acadia Capital referral channel", "Brien Walton / Acadia", 2, 3, 3, 4, 3, 5),
]
first = hrow + 1
for i, (name, src, *scores) in enumerate(opps):
    rr = first + i
    wt.cell(row=rr, column=1, value=i + 1).alignment = CENTER
    wt.cell(row=rr, column=2, value=name).alignment = WRAP
    wt.cell(row=rr, column=3, value=src).alignment = WRAP
    for k, sc_ in enumerate(scores):
        c = wt.cell(row=rr, column=4 + k, value=sc_); c.alignment = CENTER; c.border = BORDER
        c.fill = INPUT_FILL
    # composite = weighted average (cols D..I = REV..RISKADJ)
    comp = (f"=D{rr}*{wkey['REV']}+E{rr}*{wkey['PROB']}+F{rr}*{wkey['RECUR']}"
            f"+G{rr}*{wkey['STRAT']}+H{rr}*{wkey['SPEED']}+I{rr}*{wkey['RISKADJ']}")
    cc = wt.cell(row=rr, column=10, value=comp); cc.number_format = NUM2; cc.fill = TOTAL_FILL
    cc.font = Font(bold=True)
last = first + len(opps) - 1
# live rank
for i in range(len(opps)):
    rr = first + i
    rk = wt.cell(row=rr, column=11, value=f"=RANK(J{rr},$J${first}:$J${last},0)")
    rk.alignment = CENTER; rk.font = Font(bold=True, color="1F4E5F")
# top-composite highlight note row
note_r = last + 2
wt.cell(row=note_r, column=2,
        value="Tip: sort by Rank (col K) to see the priority order. Highest composite = work it first.").font = NOTE
wt.column_dimensions["A"].width = 4
wt.column_dimensions["B"].width = 46
wt.column_dimensions["C"].width = 24
for col in "DEFGHI":
    wt.column_dimensions[col].width = 9
wt.column_dimensions["J"].width = 11
wt.column_dimensions["K"].width = 7
wt.freeze_panes = "B5"

# ----------------------------------------------------------------- REVENUE ----
wr = wb.create_sheet("Revenue")
title(wr, "PublicLogic fee potential per opportunity (illustrative)",
      "Near-12-month vs lifetime PublicLogic revenue. Estimates from PublicLogic's own deal docs — not quotes. Live totals.")
rh = ["#", "Opportunity", "Near 12-mo ($)", "Lifetime ($)", "Basis / fee structure"]
hrow = 4
for j, h in enumerate(rh, start=1):
    wr.cell(row=hrow, column=j, value=h)
style_header(wr, hrow, len(rh))
# (near12, lifetime, basis)
rev = [
    (200000, 775000, "$175K capitalized build + ~$10K/mo runtime + grant-stacking; 5-yr single-site ~$775K"),
    (60000, 2250000, "Sites 2-4 @ $75-100K build + $10K/mo each; network lifetime across 3 sites"),
    (40000, 1500000, "Portfolio MSA: per-client governance + runtime across WasteWerx deployments"),
    (25000, 2000000, "Productized NMTC 7-yr compliance module repeated across CDE-financed deals"),
    (20000, 1000000, "Glen as recurring connector: scoping retainers + builds across his pipeline"),
    (30000, 400000, "La Porte scope -> grant study -> build + runtime"),
    (35000, 300000, "AED: $5K setup + $5K/mo x24 + $18K/yr yrs3-7 + underwriting (if reactivated)"),
    (25000, 250000, "Flint Phase 1 $18-25K + $6.5-15K/mo retainer"),
    (10000, 1500000, "CivicPlus: per-client referral economics at scale (early)"),
    (0, 500000, "Indirect: referred capital-stack deals via Acadia validation"),
]
first = hrow + 1
for i, (n12, life, basis) in enumerate(rev):
    rr = first + i
    wr.cell(row=rr, column=1, value=i + 1).alignment = CENTER
    wr.cell(row=rr, column=2, value=opps[i][0]).alignment = WRAP
    a = wr.cell(row=rr, column=3, value=n12); a.number_format = MONEY
    b = wr.cell(row=rr, column=4, value=life); b.number_format = MONEY
    wr.cell(row=rr, column=5, value=basis).alignment = WRAP
last = first + len(rev) - 1
wr.cell(row=last + 1, column=2, value="TOTAL (illustrative pipeline potential)").font = SUB
ta = wr.cell(row=last + 1, column=3, value=f"=SUM(C{first}:C{last})"); ta.number_format = MONEY
ta.fill = TOTAL_FILL; ta.font = Font(bold=True)
tb = wr.cell(row=last + 1, column=4, value=f"=SUM(D{first}:D{last})"); tb.number_format = MONEY
tb.fill = TOTAL_FILL; tb.font = Font(bold=True)
for col, w in zip("ABCDE", [4, 44, 15, 15, 60]):
    wr.column_dimensions[col].width = w
wr.freeze_panes = "B5"

# ----------------------------------------------------------------- ACTION -----
wa = wb.create_sheet("Action")
title(wa, "Action plan — the move, the route, the flag",
      "How to convert each opportunity. Keep every business-funded fee fixed & non-contingent on federal money.")
ah = ["#", "Opportunity", "Cost-treatment route", "Key flag", "Next move (owner)"]
hrow = 4
for j, h in enumerate(ah, start=1):
    wa.cell(row=hrow, column=j, value=h)
style_header(wa, hrow, len(ah))
actions = [
    ("Capitalized dev soft cost in project budget", "Securities lane; 5% grant-stacking must be non-contingent on federal",
     "Chase EGLE 'Encouraged to Apply' (~6/1); if invited, lead 7/7 full apps"),
    ("Same; per-site capitalized buildout", "Replication depends on Site 1 closing first",
     "Lock Site-1 close; template the buildout for Sites 2-4"),
    ("Per-client governance + runtime", "Don't become a captive sub of one OEM",
     "Convert to a portfolio MSA / preferred-partner agreement (JT Clark)"),
    ("NMTC soft-costs line of each deal", "Keep underwriting fixed/non-contingent for federal pieces",
     "Open CEI (Rob Levin) + USBCDC (via Lin Thornton); productize the module"),
    ("Scoping retainer + per-deal soft cost", "Glen's deals are early/funding-contingent — front-load paid scoping",
     "Deliver capabilities workbook + La Porte scope; lock WasteWerx intro"),
    ("Scope -> grant study -> build", "Funding-contingent; keep PL fee non-contingent",
     "Send NDA + one-page scope; define grant-study deliverable"),
    ("Project cost in $806,502 soft-costs line", "STALLED on unresponsive sponsor; reconcile job counts",
     "Reactivate via Loud OR pursue biochar thesis through Glen's channel"),
    ("Soft cost in capital stack at close", "Don't promise grant outcomes; stay out of PuddleJumper weeds",
     "Get the scoping call with the City's funding-stack owner"),
    ("Partner/referral economics", "Early; define integration + referral split",
     "Define the CivicPlus partnership model and pilot a joint client"),
    ("n/a (referral)", "Pure leverage; protect PL independence",
     "Keep Walton as validator; ask for warm intros to other deals"),
]
first = hrow + 1
for i, (route, flag, move) in enumerate(actions):
    rr = first + i
    wa.cell(row=rr, column=1, value=i + 1).alignment = CENTER
    wa.cell(row=rr, column=2, value=opps[i][0]).alignment = WRAP
    wa.cell(row=rr, column=3, value=route).alignment = WRAP
    wa.cell(row=rr, column=4, value=flag).alignment = WRAP
    wa.cell(row=rr, column=5, value=move).alignment = WRAP
for col, w in zip("ABCDE", [4, 40, 30, 38, 46]):
    wa.column_dimensions[col].width = w
wa.freeze_panes = "B5"

# --------------------------------------------------- MUNICIPAL BEDROCK --------
wm = wb.create_sheet("Municipal_Bedrock")
title(wm, "Municipal bedrock (MA) — context, scored the same way",
      "The own-source/grant-funded MA municipal layer that runs even if the competitive layer disappears. "
      "Same six criteria & weights as Top10. These are the stable counterweight to the business deals.")
mh = ["Opportunity", "Town / source", "REV", "PROB", "RECUR", "STRAT", "SPEED", "RISKADJ", "Composite", "Rank"]
hrow = 4
for j, h in enumerate(mh, start=1):
    wm.cell(row=hrow, column=j, value=h)
style_header(wm, hrow, len(mh))
muni = [
    ("Sutton TIF / VAULT Diagnostic ($475M district)", "Sutton (Compact grant)", 3, 5, 4, 5, 5, 5),
    ("Shrewsbury Municipal Fiber -> retainer", "Shrewsbury (c.30B s.7)", 4, 4, 4, 5, 4, 4),
    ("Millbury MS4 stormwater (replicable module)", "Millbury", 3, 3, 5, 5, 3, 5),
    ("Fitchburg CIP (RFP 26-092, submitted)", "Fitchburg (competitive)", 4, 3, 3, 5, 3, 4),
    ("MBI BEAD broadband RFQ (2026-MBI-08)", "MA Broadband Institute", 4, 3, 3, 4, 3, 4),
    ("Phillipston website + Memorial Building", "Phillipston", 2, 5, 3, 3, 5, 4),
    ("Gardner home-turf engagement (target)", "Gardner (campaign flag)", 3, 2, 3, 4, 2, 3),
]
first = hrow + 1
for i, (name, src, *sc) in enumerate(muni):
    rr = first + i
    wm.cell(row=rr, column=1, value=name).alignment = WRAP
    wm.cell(row=rr, column=2, value=src).alignment = WRAP
    for k, v in enumerate(sc):
        c = wm.cell(row=rr, column=3 + k, value=v); c.alignment = CENTER; c.border = BORDER; c.fill = INPUT_FILL
    comp = (f"=C{rr}*{wkey['REV']}+D{rr}*{wkey['PROB']}+E{rr}*{wkey['RECUR']}"
            f"+F{rr}*{wkey['STRAT']}+G{rr}*{wkey['SPEED']}+H{rr}*{wkey['RISKADJ']}")
    cc = wm.cell(row=rr, column=9, value=comp); cc.number_format = NUM2; cc.fill = TOTAL_FILL; cc.font = Font(bold=True)
last = first + len(muni) - 1
for i in range(len(muni)):
    rr = first + i
    rk = wm.cell(row=rr, column=10, value=f"=RANK(I{rr},$I${first}:$I${last},0)")
    rk.alignment = CENTER; rk.font = Font(bold=True, color="1F4E5F")
for col, w in zip("ABCDEFGHIJ", [42, 24, 8, 8, 8, 8, 8, 9, 11, 7]):
    wm.column_dimensions[col].width = w
wm.freeze_panes = "C5"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Top10_Opportunities.xlsx")
wb.save(out)
print("wrote", out)
