#!/usr/bin/env python3
"""
Builds Challenge4_Rubric_Score.xlsx — scores the Michigan LTC Network application two ways
against the ACTUAL EGLE Challenge #4 rubric weights: (A) as-is concept, (B) the pivoted
facility-decarbonization study. Live weighted totals out of 100; shows why the pivot wins.

Rubric (from the uploaded scoring rubric):
  S1 Emissions Impact & Depth of Decarbonization  30 pts
  S2 Facility Commitment & Likelihood of Use      20 pts
  S3 Community Benefits & Co-Benefits             15 pts
  S4 Technology & Innovation Value                15 pts
  S5 Replicability & Value to Michigan            10 pts
  S6 Team Competence & Project Approach           10 pts
Each scored 0-1 as % of its max; weighted points = pct * max. Totals are live SUM formulas.
Estimates are this analysis's judgement, not EGLE's; edit the yellow cells to test.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="7A1F1F")
SUB = Font(bold=True, size=11, color="7A1F1F")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE = Font(italic=True, size=9, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = '0%'; NUM1 = '0.0'


def style_header(ws, row, n, start=1):
    for c in range(start, start + n):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER


def title(ws, t, s=None):
    ws["A1"] = t; ws["A1"].font = Font(bold=True, size=14, color="7A1F1F")
    if s:
        ws["A2"] = s; ws["A2"].font = NOTE


wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Score"
title(ws, "EGLE Challenge #4 — Rubric Score: As-is vs Pivot (live)",
      "Edit yellow %-of-max cells. Weighted points = % x max. Totals out of 100 are live. "
      "Judgement of this analysis, not EGLE. Date 2026-06-03.")
hdr = ["§", "Criterion", "Max pts", "As-is % of max", "As-is pts", "Pivot % of max", "Pivot pts", "Why the gap"]
hrow = 4
for j, h in enumerate(hdr, start=1):
    ws.cell(row=hrow, column=j, value=h)
style_header(ws, hrow, len(hdr))

# (code, criterion, maxpts, asis_pct, pivot_pct, why)
rows = [
    ("S1", "Emissions Impact & Depth of Decarbonization", 30, 0.20, 0.80,
     "As-is has no existing-facility baseline & isn't cutting a facility's process emissions >=50%; pivot studies McBain/Lincoln to >=50% with AVERT/EPA factors"),
    ("S2", "Facility Commitment & Likelihood of Use", 20, 0.25, 0.85,
     "As-is 'facility' is the new build; pivot names an existing MI facility + decision-maker (Schopp) with data access & a commitment letter"),
    ("S3", "Community Benefits & Co-Benefits", 15, 0.55, 0.85,
     "Strong community story either way (ACS, The Warehouse) but must re-anchor to the facility & cap at 5% subaward"),
    ("S4", "Technology & Innovation Value", 15, 0.40, 0.75,
     "Pyrolysis-to-fuel steered away from; pivot leads with process electrification + circularity + enabling infrastructure"),
    ("S5", "Replicability & Value to Michigan", 10, 0.55, 0.85,
     "Two facilities + transferable methodology for MI biomass/combustion fleet + EGLE knowledge-sharing commitment"),
    ("S6", "Team Competence & Project Approach", 10, 0.45, 0.80,
     "Both need a technical engineering/emissions-modeling lead (current GAP); pivot adds it + clear PL/Rothschild/facility roles"),
]
first = hrow + 1
for i, (code, crit, mx, ap, pp, why) in enumerate(rows):
    rr = first + i
    ws.cell(row=rr, column=1, value=code).alignment = CENTER
    ws.cell(row=rr, column=2, value=crit).alignment = WRAP
    ws.cell(row=rr, column=3, value=mx).alignment = CENTER
    a = ws.cell(row=rr, column=4, value=ap); a.number_format = PCT; a.fill = INPUT_FILL; a.alignment = CENTER; a.border = BORDER
    ws.cell(row=rr, column=5, value=f"=C{rr}*D{rr}").number_format = NUM1
    p = ws.cell(row=rr, column=6, value=pp); p.number_format = PCT; p.fill = INPUT_FILL; p.alignment = CENTER; p.border = BORDER
    ws.cell(row=rr, column=7, value=f"=C{rr}*F{rr}").number_format = NUM1
    ws.cell(row=rr, column=8, value=why).alignment = WRAP
last = first + len(rows) - 1
trow = last + 1
ws.cell(row=trow, column=2, value="TOTAL (out of 100)").font = SUB
ws.cell(row=trow, column=3, value=f"=SUM(C{first}:C{last})").alignment = CENTER
for col in ("E", "G"):
    c = ws.cell(row=trow, column=5 if col == "E" else 7, value=f"=SUM({col}{first}:{col}{last})")
    c.number_format = NUM1; c.fill = TOTAL_FILL; c.font = Font(bold=True)
# delta
ws.cell(row=trow + 1, column=2, value="Pivot advantage (pts)").font = SUB
d = ws.cell(row=trow + 1, column=7, value=f"=G{trow}-E{trow}"); d.number_format = NUM1
d.fill = TOTAL_FILL; d.font = Font(bold=True, color="7A1F1F")

notes = [
    "",
    "READ: As-is lands well below a fundable score and floors the 30%-weighted Emissions criterion. The pivot",
    "(study an EXISTING MI facility's process to >=50% decarbonization) is the only version that competes.",
    "",
    "GATING ITEMS before this score is real: (1) signed facility commitment letter + named decision-maker;",
    "(2) a technical engineering/energy-&-emissions-modeling partner; (3) facility baseline (EPA FLIGHT/GHGRP);",
    "(4) budget <$400K, study-only, task-broken-out, PL fee fixed & non-contingent; (5) strip carbon-credit/45Q;",
    "(6) confirm the individualized 'Encouraged to Apply' status with Haley Neuenfeldt.",
    "",
    "Not EGLE's scoring. Not legal/accounting advice. Edit the yellow %-of-max cells to test assumptions.",
]
nr = trow + 3
for ln in notes:
    c = ws.cell(row=nr, column=2, value=ln)
    if ln.startswith(("READ", "GATING")):
        c.font = SUB
    else:
        c.font = NOTE
    nr += 1
for col, w in zip("ABCDEFGH", [5, 40, 9, 14, 11, 14, 11, 60]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Challenge4_Rubric_Score.xlsx")
wb.save(out)
print("wrote", out)
