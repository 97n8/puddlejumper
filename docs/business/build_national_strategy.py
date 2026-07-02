#!/usr/bin/env python3
"""
PublicLogic — National Strategy & Sector Outreach (v1).
Expands PL from MA-municipal + the Michigan deal to a repeatable national
practice. The scalable IP is the Funding-Landscape engine (Michigan v5.1 =
reference build): every state/sector has the same three problems and an
unmapped funding stack. Grounded, phased, 2-person-capacity-aware.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

SLATE="1A1D20"; GREEN="0F4C3A"; GOLD="A9772F"; MUTEC="6B6660"
HDR=Font(bold=True,color="FFFFFF",size=11); HFILL=PatternFill("solid",fgColor=SLATE)
GHFILL=PatternFill("solid",fgColor=GREEN); TITLE=Font(bold=True,size=15,color=SLATE)
SUB=Font(italic=True,size=9,color=MUTEC); SEC=Font(bold=True,size=12,color=GREEN)
GLD=Font(bold=True,size=11,color=GOLD); BOLD=Font(bold=True); BODY=Font(size=11,color=SLATE)
CFILL=PatternFill("solid",fgColor="ECEAE3"); GFILL=PatternFill("solid",fgColor="E3EAE5"); IFILL=PatternFill("solid",fgColor="FFF2CC")
THIN=Side(style="thin",color="CCCCCC"); B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP=Alignment(wrap_text=True,vertical="top"); MID=Alignment(wrap_text=True,vertical="center")

wb=openpyxl.Workbook()
def newtab(name,title,sub=None):
    ws=wb.create_sheet(name); ws["A1"]=title; ws["A1"].font=TITLE
    if sub: ws["A2"]=sub; ws["A2"].font=SUB
    return ws
def blocks(ws,start,rows,h=34,w=(24,104)):
    r=start
    for k,v in rows:
        a=ws.cell(row=r,column=1,value=k); a.font=HDR; a.fill=GHFILL; a.border=B; a.alignment=MID
        c=ws.cell(row=r,column=2,value=v); c.alignment=WRAP; c.border=B; c.fill=CFILL; c.font=BODY
        ws.row_dimensions[r].height=h; r+=1
    ws.column_dimensions["A"].width=w[0]; ws.column_dimensions["B"].width=w[1]
    return r
def lines(ws,start,items,col=1,width=112):
    r=start
    for it in items:
        c=ws.cell(row=r,column=col,value=it); c.alignment=WRAP; c.font=BODY; r+=1
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width=width
    return r
def htable(ws,top,headers,rows,widths,editable=()):
    for i,hh in enumerate(headers):
        c=ws.cell(row=top,column=i+1,value=hh); c.font=HDR; c.fill=HFILL; c.border=B; c.alignment=MID
    for ri,row in enumerate(rows):
        for ci,val in enumerate(row):
            c=ws.cell(row=top+1+ri,column=ci+1,value=val); c.border=B; c.alignment=WRAP; c.font=BODY
            if ci in editable: c.fill=IFILL
    for i,w in enumerate(widths): ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width=w
    ws.freeze_panes=f"A{top+1}"

# ===== START HERE =====
ws=wb.active; ws.title="Start Here"
ws["A1"]="PublicLogic — National Strategy & Sector Outreach"; ws["A1"].font=Font(bold=True,size=16,color=SLATE)
ws["A2"]="v1 · How a two-person practice goes national without going wide. One method, applied where it's warm."; ws["A2"].font=SUB
blocks(ws,4,[
 ("The thesis","Our method is state-agnostic. Every jurisdiction in the country has the same three problems we already solve — institutions that lose continuity when people leave, sponsors stuck in the public-systems maze, and a funding stack no one has mapped. The programs differ by state; the method doesn't."),
 ("Why now","Michigan LTC proved the model works outside Massachusetts and outside pure municipal work — a waste-to-fuel corridor across three sites, with a full funding landscape mapped and sequenced. That engagement is the template, not the exception."),
 ("What travels","The method (Map → Embed → Encode → Sustain), the offer ladder, the Permit & Bridge navigation, the stewardship record — and the Funding-Landscape discipline. None of it is Massachusetts-specific."),
 ("What changes","Only the specifics: which programs, which agencies, which deadlines, which local sponsor. That's exactly what a Funding Landscape captures per engagement — so the variable part is itself a paid deliverable."),
 ("The engine","The Funding-Landscape methodology (Michigan v5.1 is the reference build) is our scalable IP: a repeatable per-sector, per-state map of the whole instrument stack, status-controlled and sequenced. It's the reason clients need us, and it's a product."),
 ("The wedge","National does NOT mean blanketing 50 states. Lead where it's warm and proven: municipal, plus the sectors we have proof in — waste/recovery, clean fuel, brownfield/redevelopment, rural. Follow warm deals and channel partners, not the map."),
 ("Capacity discipline","We are two people. National scale comes from three levers — productized deliverables, channel partners who bring deals, and selective depth — never from headcount or geographic sprawl."),
 ("How to use this","Sectors → the tiered target list. Sector Outreach → the working matrix (who to call, what to sell, what funding to map, scored by fit × warmth). Then Channels, Phasing, the Engine, the 36-month plan, and the guardrails that keep us from over-reaching."),
],h=44)

# ===== WHY NATIONAL =====
ws=newtab("Why National","Why the Model Is National","Three problems that exist in every jurisdiction — and one method that answers all three.")
htable(ws,4,["The universal problem","How it shows up everywhere","What PublicLogic brings"],
 [
 ["Lost continuity","Staff leave, knowledge walks out, functions run through one person — in every town, utility, and small organization in the country.","Stewardship: the record and structure that survive turnover."],
 ["The public-systems maze","Permits, boards, funding rules, and politics stall good projects everywhere — the offices differ, the maze doesn't.","Permit & Bridge: navigation through the public system."],
 ["The unmapped funding stack","Every sector and state has a stack of grants, credits, loans, and local tools no one has mapped or sequenced.","The Funding Landscape: the whole stack, status-controlled and sequenced."],
 ],[22,52,44])
ws.cell(row=9,column=1,value="ONE METHOD, MANY PROGRAMS").font=SEC; ws.merge_cells("A9:C9")
htable(ws,10,["Layer","National / universal (travels)","Local / variable (re-mapped per job)"],
 [
 ["Method","Map → Embed → Encode → Sustain","—"],
 ["Offers","Scan · Map · Sprints · Implementation · Capacity · Funding Landscape","Price bands hold; scope flexes"],
 ["Navigation","Permit & Bridge discipline","Which boards, which offices, which sequence"],
 ["Funding","The landscape method + status vocabulary","Which programs, agencies, deadlines, sponsors"],
 ["Proof","Continuity · record · readiness · adoption · outcome","The specific case"],
 ],[16,50,52])

# ===== SECTORS =====
ws=newtab("Sectors","Target Sectors — Tiered","Lead with proof and warmth. Tier A first; Tier B as A produces; Tier C is monitor-only.")
htable(ws,4,["Sector","Tier","Why we win here","Proof / adjacency","Entry offer"],
 [
 ["Municipal & local government","A — lead","Home ground; the continuity problem in its purest form; warmest relationships.","Sutton, Shrewsbury, Phillipston, Gardiner","Stewardship Map"],
 ["Waste / recycling / circular economy","A — lead","Feedstock + market + permitting + a state materials-grant stack no one maps.","Michigan LTC corridor","Funding Landscape + Permit & Bridge"],
 ["Clean energy / biofuels / SAF","A — lead","Dense credit stack (45Z/45Q/48C), RFS pathway, university validation — high complexity, high value.","Michigan LTC · WasteWerx / UCF","Funding Landscape + STTR partnership"],
 ["Brownfield & redevelopment","A — lead","Assessment→cleanup→reuse funding + TIF + NMTC almost never aligned by one party.","Sutton TIF · NMTC work · Flint (LTC)","Permit & Bridge + Funding Landscape"],
 ["Rural & agriculture","A — lead","USDA maze; grants frozen but guarantees open; wood/biomass supply chains.","Coleman (LTC) · Wood Innovations","Funding Landscape"],
 ["Water / wastewater utilities","B — next","Capital planning + SRF/WIFIA navigation + asset continuity.","Capital-planning work","Stewardship Map + Funding Landscape"],
 ["Housing & community development","B — next","Layered capital stack (CDBG/HOME/NMTC); predevelopment + community benefit.","NMTC · community-dev adjacency","Permit & Bridge + Funding Landscape"],
 ["Higher ed / research partnerships","B — next","SBIR/STTR structure + a credible adoption path; commercialization.","WasteWerx / UCF","STTR partnership ($2,500 grant-fit)"],
 ["Economic development / small mfg","B — next","Site readiness + incentives + workforce — the MBDP-type stack.","MEDC-type work (LTC)","Funding Landscape"],
 ["Broadband / digital equity","C — monitor","BEAD / digital-equity navigation + local coordination.","Adjacent municipal","Permit & Bridge + Funding"],
 ["Long-term care / human services","C — monitor","Practice adoption + staff continuity after change — Allie's core domain.","Allie's clinical/org background","Stewardship Map + Capacity Support"],
 ],[30,12,44,30,30])

# ===== SECTOR OUTREACH (the matrix) =====
ws=newtab("Sector Outreach","Sector Outreach Configuration","The working matrix. Set Fit and Warmth (1–5); Score = Fit × Warmth; sort to rank. Yellow = you edit.")
heads=["Sector","Who we call (buyers)","The pain we lead with","Entry offer","Funding stack to map","Fit","Warmth","Score"]
for i,hh in enumerate(heads):
    c=ws.cell(row=4,column=i+1,value=hh); c.font=HDR; c.fill=HFILL; c.border=B; c.alignment=MID
cfg=[
 ["Municipal & local government","Town/city managers, finance directors, DPW, select boards","Your best work lives in one person's head — and the grant window closes before anyone's ready.","Stewardship Map","CDBG · state revolving funds · EDA · state municipal grants",5,5],
 ["Waste / recycling / circular economy","Recycling authorities, MRF operators, haulers, county solid-waste","You have feedstock and a buyer somewhere — but the permits and the state grant stack are a maze.","Funding Landscape + Permit & Bridge","State materials-mgmt grants · EPA SWIFR/recycling · scrap-tire · 45Z (if fuel)",5,5],
 ["Clean energy / biofuels / SAF","Developers, equipment/tech cos, fuel producers","The credit stack is worth millions and nobody has mapped 45Z + RFS + 48C + the university piece together.","Funding Landscape + STTR partnership","45Z · 45Q · 48C · DOE BETO · USDA · RFS RINs",5,5],
 ["Brownfield & redevelopment","Municipalities, brownfield authorities (BRAs), developers","An idle, contaminated site and four funding tools that only work if someone aligns them.","Permit & Bridge + Funding Landscape","EPA Brownfields · state brownfield · Act 381-type TIF · NMTC",5,4],
 ["Rural & agriculture","Rural project SPVs, ag operators, co-ops, small manufacturers","USDA is a maze; the grants froze but the guarantees are open — you need someone who knows which door is which.","Funding Landscape","USDA B&I guarantee · Wood Innovations · REAP loan guarantee · state ag",4,4],
 ["Water / wastewater utilities","Utility directors, DPW, water/sewer districts","A capital plan you can't fund and an SRF/WIFIA process no one has time to run.","Stewardship Map + Funding Landscape","CWSRF / DWSRF · EPA WIFIA · USDA water · state",4,3],
 ["Housing & community development","CDCs, housing authorities, municipalities","A capital stack five layers deep and a predevelopment gap that kills good projects.","Permit & Bridge + Funding Landscape","CDBG · HOME · NMTC · state housing · FHLB",4,3],
 ["Higher ed / research partnerships","University research & tech-transfer offices, startups","A real technology and a real university — but no one has built the SBIR/STTR structure or the adoption path.","STTR partnership ($2,500 grant-fit)","SBIR / STTR (thru FY2031) · state match",4,4],
 ["Economic development / small mfg","EDOs, small manufacturers, MSF-type programs","Site readiness, incentives, and workforce money that only shows up if you engage before you announce.","Funding Landscape","State business-dev grants · 48C · workforce funds · EDA",4,3],
 ["Broadband / digital equity","Municipalities, regional broadband authorities","BEAD and digital-equity money with a coordination problem attached.","Permit & Bridge + Funding","BEAD · state digital equity · CDBG",3,2],
 ["Long-term care / human services","Multi-site care orgs, human-services nonprofits","A strong start that doesn't stick — new practices and staff buy-in fade once the launch is over.","Stewardship Map + Capacity Support","State/federal HHS · foundation grants",3,3],
]
for i,row in enumerate(cfg):
    rr=5+i
    for ci,val in enumerate(row):
        c=ws.cell(row=rr,column=ci+1,value=val); c.border=B; c.alignment=WRAP; c.font=BODY
        if ci in (5,6): c.fill=IFILL; c.alignment=Alignment(horizontal="center",vertical="center")
    sc=ws.cell(row=rr,column=8,value=f'=IF(AND(ISNUMBER(F{rr}),ISNUMBER(G{rr})),F{rr}*G{rr},"")')
    sc.border=B; sc.font=BOLD; sc.fill=GFILL; sc.alignment=Alignment(horizontal="center",vertical="center")
for col,w in zip("ABCDEFGH",[30,40,52,30,44,7,9,8]): ws.column_dimensions[col].width=w
nr=5+len(cfg)+1
ws.cell(row=nr,column=1,value="Fit = how well our method solves it (1–5). Warmth = how close we are to a real buyer today (1–5). Funding instruments are illustrative and MUST be re-verified per the Engine tab — the federal layer shifts (e.g., REAP grants frozen, EJ programs terminated, 45Z/45Q expanded under OBBBA).").font=SUB
ws.cell(row=nr,column=1).alignment=WRAP; ws.merge_cells(f"A{nr}:H{nr+1}")

# ===== CHANNELS =====
ws=newtab("Channels","Channels & Partners — The Multiplier","Two people don't cold-call a nation. We grow through partners who bring the deals to us.")
htable(ws,4,["Channel","What they bring","What we give them","How to engage"],
 [
 ["Regional planning agencies / COGs (via NARC nationally)","Member towns with capacity gaps and stalled projects","A continuity + funding capability their members need; they look good","Offer a Funding Landscape or Map for a member town; become the referral name"],
 ["Engineering / A&E firms","Clients mid-project who need the funding + continuity layer","The piece next to their design — 'they build it, we make sure the town can carry it and fund it'","Partner on a shared client; take the handoff/continuity slice"],
 ["Municipal management consultancies","Overflow and the stewardship work they don't do","A specialist subcontractor for continuity & funding","White-label or co-deliver"],
 ["Universities / research offices","SBIR/STTR partner pipeline; independent-validation credibility","Applicants who need the grant built and the adoption plan (Allie)","The WasteWerx/UCF model, repeated"],
 ["CDEs (NMTC) & mission lenders","Community-project deal flow needing readiness","Ready, well-structured deals","Open pipeline conversations; be the readiness partner"],
 ["USDA-approved lenders","B&I / guarantee deals that need readiness and a funding stack","Borrowers who are actually ready","Engage at the financing stage"],
 ["Economic development orgs / IEDC network","Site + incentive projects","Grant + continuity capacity","Join the network; be the funding-map specialist"],
 ["Technology / equipment vendors (e.g., WasteWerx)","Project sponsors who need funding + permitting to buy/deploy","A partner who gets their customers funded and sited","Bundle a Funding Landscape into their sales motion"],
 ],[34,40,40,42])

# ===== THE ENGINE =====
ws=newtab("The Engine","The Funding-Landscape Engine — Our National IP","The repeatable deliverable that makes national possible. Michigan v5.1 is the reference build.")
r=lines(ws,4,[
 "What it is: a per-project map of the ENTIRE instrument stack — state grants, federal grants, tax credits, loans/guarantees, and local tools (TIF, brownfield) — each one status-controlled and sequenced against the project's stages.",
 "",
 "Why it's the unlock: every sector and every state has a stack this deep, and almost no one maps it or sequences it. That's the gap PublicLogic fills anywhere in the country, without needing local staff.",
 "",
 "Productize it: sell the Funding Landscape as a paid deliverable (scoped like a Diagnostic). It stands alone, it de-risks the client's next decision, and it becomes the reason they buy the Map, the Permit & Bridge work, and the implementation that follows.",
 "",
])
ws.cell(row=r+1,column=1,value="THE DISCIPLINE THAT MAKES IT DEFENSIBLE").font=SEC
htable(ws,r+2,["Rule","Why it matters"],
 [
 ["Verify every program against agency sources, dated","Credibility; the map is only worth what its sourcing is worth"],
 ["Controlled status vocabulary (Active → Terminated)","Everyone reads the same word the same way; nothing overclaimed"],
 ["Internal vs external citation rules","Terminated/Frozen never appears in a client or funder submission"],
 ["Investor economics never touch a public submission (FOIA)","Protects the client and us"],
 ["Re-verify on a schedule — the federal layer shifts","REAP grants frozen · EJ programs terminated · 45Z/45Q expanded (OBBBA). Re-verification is recurring revenue."],
 ],[46,64])

# ===== 36-MONTH PLAN =====
ws=newtab("36-Month Plan","The 36-Month Plan","Deepen, then expand by sector, then go national by opportunity. Never by map coverage.")
htable(ws,4,["Horizon","Focus","Concrete moves","Revenue logic"],
 [
 ["0–12 months","Deepen home + active deal","Sell Maps + Funding Landscapes across the warm MA network; carry Michigan LTC through Challenge #4 and the credit stack; productize the Funding Landscape offer.","A handful of Maps + 1–2 Funding Landscapes + the LTC engagement. Prove the product."],
 ["12–24 months","Expand by sector, not geography","Follow waste/recovery, brownfield, and clean-fuel deals into 2–3 more states where a warm deal or channel partner already exists; stand up 2–3 channel relationships (an RPC network, a university, a vendor).","Repeatable Funding-Landscape + Sprint revenue; first channel-sourced deals; one recurring re-verification retainer."],
 ["24–36 months","National by opportunity","Go where a channel partner or a fundable project pulls us — nationally, sector by sector. Consider a first hire or associate only if productized demand demands it.","Diversified pipeline across 4–5 sectors and several states; recurring landscape/monitoring revenue as the base."],
 ],[16,30,52,40])

# ===== GUARDRAILS =====
ws=newtab("Guardrails","Guardrails — How We Don't Blow It","National ambition, disciplined execution. These keep the expansion from diluting the firm.")
for t in [
 "Don't dilute the wedge. Municipal + the proof sectors lead. New sectors are adjacent, not scattershot.",
 "Capacity is the constraint. Two people. Scale through products and partners, never headcount or geographic sprawl.",
 "Fees stay non-contingent, everywhere. Never a cut of an award; never contingent on a decision.",
 "The language guardrails still apply. No 'AI transformation,' no 'platform,' no hype — nationally as locally.",
 "Never build a plan on one program. The stack + scheduled re-verification is the hedge against federal volatility.",
 "Public submissions stay clean. Investor economics never touch a public/EGLE-type filing (FOIA).",
 "Confidentiality across deals is absolute. What we learn in one engagement never leaks into another.",
 "Follow warm deals and channel partners — not the map. Coverage is a vanity metric; a fundable project is not.",
]:
    c=ws.cell(row=ws.max_row+2 if ws.max_row>2 else 4,column=1,value="—  "+t); c.font=BODY; c.alignment=WRAP
ws.column_dimensions["A"].width=118

# ===== OUTREACH TRACKER =====
ws=newtab("Outreach Tracker","Outreach Tracker (live)","One row per real target. Fill it as you go; sort by Warmth. Yellow = you edit.")
heads=["Sector","Target org","Contact / role","Warmth (1–5)","Entry offer","Next action","Status"]
for i,hh in enumerate(heads):
    c=ws.cell(row=4,column=i+1,value=hh); c.font=HDR; c.fill=HFILL; c.border=B; c.alignment=MID
seed=[
 ["Municipal","Warm MA network (Sutton / Shrewsbury / Phillipston)","Known contacts","5","Stewardship Map","Book the Map on one stalled function","Warm"],
 ["Regional planning","CMRPC","Dir. / econ-dev staff","3","Funding Landscape (member town)","Intro call","Developing"],
 ["Waste / circular","Michigan LTC corridor","Robert / Glenn","5","Challenge #4 + Funding Landscape","Clear LARA + Lincoln + letters","Active"],
 ["Clean fuel / research","WasteWerx / UCF","Vincent","4","STTR partnership","Send one-pager; 1pm call","Active"],
 ["Brownfield","(target BRA / developer)","—","2","Permit & Bridge + Funding","Identify a warm site","Prospect"],
]
for i,row in enumerate(seed):
    rr=5+i
    for ci,val in enumerate(row):
        c=ws.cell(row=rr,column=ci+1,value=val); c.border=B; c.alignment=WRAP; c.font=BODY
        if ci in (1,2,3,5,6): c.fill=IFILL
for col,w in zip("ABCDEFG",[18,40,24,12,30,40,16]): ws.column_dimensions[col].width=w

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"final_deck","PublicLogic - National Strategy & Sector Outreach (v1).xlsx")
os.makedirs(os.path.dirname(out),exist_ok=True)
wb.save(out); print("wrote",out,"| tabs:",wb.sheetnames)
