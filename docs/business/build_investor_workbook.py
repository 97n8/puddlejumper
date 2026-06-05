#!/usr/bin/env python3
"""
Michigan_LTC_Investor_Workbook_FINAL.xlsx — the leave-behind workbook that ships with the
investor deck. Same canonical numbers as the corrected deck (reconciled EBITDA range),
plus the grant stack, caveats, and documents-on-file. Live formulas; zero errors.
Private & Confidential. Not financial/legal/securities advice.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

INK=openpyxl.styles.colors.Color(rgb="FF17202B")
HDR=Font(bold=True,color="FFFFFF",size=11); HFILL=PatternFill("solid",fgColor="17202B")
TITLE=Font(bold=True,size=15,color="17202B"); SUB=Font(italic=True,size=9,color="6B6258")
GOLD=Font(bold=True,size=10,color="8A6A2F"); GFILL=PatternFill("solid",fgColor="E2EFDA")
OFILL=PatternFill("solid",fgColor="FCE4D6"); IFILL=PatternFill("solid",fgColor="FFF2CC")
BOLD=Font(bold=True); MUTE=Font(size=9,italic=True,color="7A7166")
THIN=Side(style="thin",color="CCCCCC"); B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP=Alignment(wrap_text=True,vertical="top"); RIGHT=Alignment(horizontal="right"); CTR=Alignment(horizontal="center",vertical="center")
M='#,##0'

wb=openpyxl.Workbook()

def hrow(ws,r,cols,start=1):
    for i,c in enumerate(cols):
        x=ws.cell(row=r,column=start+i,value=c); x.font=HDR; x.fill=HFILL; x.border=B; x.alignment=Alignment(wrap_text=True,vertical="center")

# ---------- COVER ----------
ws=wb.active; ws.title="Cover"
ws["A1"]="Michigan LTC Network — Resource Recovery Corridor"; ws["A1"].font=TITLE
ws["A2"]="Investor & Partner Workbook  ·  to be read with the Private Investor Deck"; ws["A2"].font=Font(size=11,color="8A6A2F",bold=True)
rows=[
 ("","",),
 ("Basis of record","Canonical Financial Model v2.0 — keyed line-for-line to the executed WasteWerx Model 3 License Pro Forma (mutual NDA 2026-05-18)."),
 ("Reading","Every figure is a signed-document value or a formula over one. Economics shown as a reconciled range (signed Pro Forma vs operator-reality)."),
 ("Status","Diligence-grade working set. Open items are commercial (carbon black, RFS/§45Z, staffing reconciliation, offtake), not modeling."),
 ("Date","2026-06-05"),
 ("","",),
 ("SECURITIES NOTICE","This workbook is for confidential discussion with accredited investors only. It is NOT an offer to sell or a solicitation to buy securities. Any offering, if made, will be made only through definitive documents reviewed by securities counsel. Contains forward-looking projections that are not guarantees. PublicLogic LLC is the governance/documentation partner — NOT a placement agent, broker-dealer, or adviser; its fee is fixed and non-contingent."),
 ("HANDLING","Private & Confidential. Keep separate from EGLE/public grant materials. Do not publish proprietary WasteWerx process detail."),
]
r=4
for k,v in rows:
    ws.cell(row=r,column=1,value=k).font=BOLD if k and not k.isupper() else (GOLD if k.isupper() else Font())
    if k.isupper(): ws.cell(row=r,column=1).fill=HFILL; ws.cell(row=r,column=1).font=HDR
    c=ws.cell(row=r,column=2,value=v); c.alignment=WRAP
    if k.isupper(): c.fill=IFILL
    r+=1
ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=104

# ---------- ASSUMPTIONS (inputs) ----------
wi=wb.create_sheet("Assumptions"); wi["A1"]="Assumptions — signed-document values (Canonical v2.0)"; wi["A1"].font=TITLE
wi["A2"]="Yellow = editable. Revenue & Pro Forma opex are executed-Pro-Forma values; stress levers are the diligence swings."; wi["A2"].font=SUB
wi.append([]); wi.append(["Key","Item","Value","Source"]); hrow(wi,wi.max_row,["Key","Item","Value","Source"])
inp=[
 ("rd_sales","RD fuel sales (2,399,040 gal x $2.00)",4798080,"Canonical Revenue"),
 ("saf_sales","SAF fuel sales (2,301,120 gal x $4.50)",10355040,"Canonical Revenue"),
 ("d4_rin","D4 RIN — RD",4798080,"Canonical Revenue — contingent (RFS)"),
 ("d7_rin","D7 RIN — SAF",9204480,"Canonical Revenue — contingent (RFS)"),
 ("z45_rd","Section 45Z — RD",2399040,"Canonical Revenue — contingent"),
 ("z45_saf","Section 45Z — SAF",4026960,"Canonical Revenue — contingent"),
 ("cb","Carbon black",0,"$0 in executed Pro Forma (upside; deck implied ~$10.125M, uncontracted)"),
 ("ww_fees","WasteWerx fees (monitoring $420k + royalty $2,820,096)",3240096,"Canonical P&L"),
 ("feedstock_pf","Feedstock — Pro Forma ($0.20/gal)",940032,"Canonical P&L"),
 ("utilities","Utilities ($0.05/gal)",235008,"Canonical P&L"),
 ("distribution","Fuel distribution & shipping ($0.08/gal)",376013,"Canonical P&L"),
 ("labor_pf","Operating labor — Pro Forma (4 FTE + manager)",425000,"Canonical P&L"),
 ("insurance","Insurance",65000,"Canonical P&L"),
 ("permits","Permit renewals & compliance",25000,"Canonical P&L"),
 ("ga","G&A overhead",75000,"Canonical P&L"),
 ("labor_real","Operating labor — STAFFING pro forma (62 FTE)",5522400,"WasteWerx Staffing Pro Forma"),
 ("mi_adj","Michigan labor adjustment",1.05,"FL->MI [CONFIRM]"),
 ("xtrain","Cross-training factor (<1.0 reduces)",1.00,"Vincent Tizio: staffing is worst-case"),
 ("feedstock_real","Feedstock — operator reality",3720000,"deck $200/ton; ERR $25/ton tipping = ~$1M [CONFIRM]"),
 ("y1_cashout","Year 1 cash out",7622828,"Canonical — CONFIRMED"),
 ("raise_site","Project raise per site",15000000,"sponsor"),
 ("company_raise","Company-level raise (separate)",2500000,"sponsor"),
 ("comm","Community profit share (of EBITDA)",0.10,"Canonical — contractual"),
 ("sites","Sites in corridor",3,"Lincoln/Flint/Coleman"),
]
k={}; r0=wi.max_row+1
for i,(key,lab,val,src) in enumerate(inp):
    rr=r0+i
    wi.cell(row=rr,column=1,value=key).font=Font(bold=True,color="888888",size=9)
    wi.cell(row=rr,column=2,value=lab)
    c=wi.cell(row=rr,column=3,value=val); c.fill=IFILL; c.border=B
    c.number_format='0.00' if isinstance(val,float) else (M if isinstance(val,(int,float)) and val>=1000 else '0')
    wi.cell(row=rr,column=4,value=src).font=MUTE
    k[key]=f"Assumptions!$C${rr}"
for col,w in zip("ABCD",[14,46,14,52]): wi.column_dimensions[col].width=w
def K(x): return k[x]

# ---------- ECONOMICS ----------
we=wb.create_sheet("Economics"); we["A1"]="Per-site economics — reconciled range"; we["A1"].font=TITLE
we["A2"]="Keyed to Canonical v2.0. 'As signed' = Pro Forma (5-FTE). 'Operator-reality' = 62-FTE staffing + real feedstock."; we["A2"].font=SUB
# revenue
we.cell(row=4,column=1,value="REVENUE (steady state)").font=GOLD
hrow(we,5,["Stream","$/yr","Category"])
rev=[("Renewable diesel sales",f"={K('rd_sales')}","Base fuel"),("SAF / kerosene sales",f"={K('saf_sales')}","Base fuel"),
 ("D4 RIN",f"={K('d4_rin')}","Contingent (RFS)"),("D7 RIN",f"={K('d7_rin')}","Contingent (RFS)"),
 ("§45Z — RD",f"={K('z45_rd')}","Contingent"),("§45Z — SAF",f"={K('z45_saf')}","Contingent"),
 ("Carbon black",f"={K('cb')}","Upside (excluded)")]
r=6
for lab,f,cat in rev:
    we.cell(row=r,column=1,value=lab); c=we.cell(row=r,column=2,value=f); c.number_format=M
    c.fill=GFILL if cat=="Base fuel" else (OFILL if "Contingent" in cat else IFILL)
    we.cell(row=r,column=3,value=cat); [we.cell(row=r,column=j).__setattr__('border',B) for j in range(1,4)]
    r+=1
we.cell(row=13,column=1,value="Base fuel (RD+SAF)").font=BOLD; we.cell(row=13,column=2,value="=B6+B7").number_format=M; we.cell(row=13,column=2).fill=GFILL; we.cell(row=13,column=2).font=BOLD
we.cell(row=14,column=1,value="Contingent credits (RIN+45Z)").font=BOLD; we.cell(row=14,column=2,value="=B8+B9+B10+B11").number_format=M; we.cell(row=14,column=2).fill=OFILL; we.cell(row=14,column=2).font=BOLD
we.cell(row=15,column=1,value="TOTAL REVENUE").font=BOLD; we.cell(row=15,column=2,value="=B13+B14+B12").number_format=M; we.cell(row=15,column=2).fill=GFILL; we.cell(row=15,column=2).font=Font(bold=True,size=12)
# opex + ebitda matrix
we.cell(row=17,column=1,value="OPERATING COST & EBITDA").font=GOLD
hrow(we,18,["","As signed (Pro Forma, 5 FTE)","Operator-reality (62 FTE)"])
opex_pf=f"={K('ww_fees')}+{K('feedstock_pf')}+{K('utilities')}+{K('distribution')}+{K('labor_pf')}+{K('insurance')}+{K('permits')}+{K('ga')}"
opex_or=f"={K('ww_fees')}+{K('feedstock_real')}+{K('utilities')}+{K('distribution')}+{K('labor_real')}*{K('mi_adj')}*{K('xtrain')}+{K('insurance')}+{K('permits')}+{K('ga')}"
we.cell(row=19,column=1,value="Total operating cost"); we.cell(row=19,column=2,value=opex_pf).number_format=M; we.cell(row=19,column=3,value=opex_or).number_format=M
we.cell(row=20,column=1,value="EBITDA — fuel + credits").font=BOLD
e1=we.cell(row=20,column=2,value="=B15-B19"); e1.number_format=M; e1.font=BOLD; e1.fill=GFILL
e2=we.cell(row=20,column=3,value="=B15-C19"); e2.number_format=M; e2.font=BOLD; e2.fill=OFILL
we.cell(row=21,column=1,value="EBITDA — base fuel only (no credits)").font=BOLD
we.cell(row=21,column=2,value="=B13-B19").number_format=M; we.cell(row=21,column=2).fill=GFILL
we.cell(row=21,column=3,value="=B13-C19").number_format=M; we.cell(row=21,column=3).fill=OFILL
we.cell(row=22,column=1,value="Year 1 cash out"); we.cell(row=22,column=2,value=f"={K('y1_cashout')}").number_format=M; we.cell(row=22,column=3,value=f"={K('y1_cashout')}").number_format=M
we.cell(row=23,column=1,value="Community share (10% of EBITDA)"); we.cell(row=23,column=2,value=f"=B20*{K('comm')}").number_format=M; we.cell(row=23,column=3,value=f"=C20*{K('comm')}").number_format=M
for j in range(1,4):
    for rr in range(19,24): we.cell(row=rr,column=j).border=B
# network
we.cell(row=25,column=1,value="THREE-SITE CORRIDOR (x sites)").font=GOLD
hrow(we,26,["Metric","As signed","Operator-reality"])
net=[("Total revenue","B15","B15"),("EBITDA (fuel+credits)","B20","C20"),("Year 1 cash out","B22","B22")]
r=27
for lab,a,b2 in net:
    we.cell(row=r,column=1,value=lab); we.cell(row=r,column=2,value=f"=({a})*{K('sites')}").number_format=M; we.cell(row=r,column=3,value=f"=({b2})*{K('sites')}").number_format=M
    for j in range(1,4): we.cell(row=r,column=j).border=B
    r+=1
we.cell(row=31,column=1,value="Note: $20.4M of the $35.58M revenue is RIN/§45Z, contingent on EPA RFS registration. Carbon black is $0 in the base. The $30.2M EBITDA assumes Pro Forma 5-FTE staffing; operator-reality integrates the 62-FTE pro forma.").font=MUTE
we.cell(row=31,column=1).alignment=WRAP; we.merge_cells("A31:C32")
for col,w in zip("ABC",[40,26,26]): we.column_dimensions[col].width=w

# ---------- CAPITAL & RAISES ----------
wc=wb.create_sheet("Capital & Raises"); wc["A1"]="Capital & the two raises"; wc["A1"].font=TITLE
wc["A2"]="Site capital and company capital are distinct asks — never blended."; wc["A2"].font=SUB
hrow(wc,4,["Item","Amount","Note"])
cap=[("Project raise per site",f"={K('raise_site')}","sponsor headline"),
 ("Project raises — 3 sites",f"={K('raise_site')}*{K('sites')}","excludes company raise"),
 ("Company-level raise (separate)",f"={K('company_raise')}","working capital, payroll, governance, team [CONFIRM split]"),
 ("Combined discussion total",f"={K('raise_site')}*{K('sites')}+{K('company_raise')}","present separately"),
 ("Year 1 cash out per site",f"={K('y1_cashout')}","CONFIRMED (WasteWerx $6,670,328 + capex $952,500)")]
r=5
for lab,f,nt in cap:
    wc.cell(row=r,column=1,value=lab); c=wc.cell(row=r,column=2,value=f); c.number_format=M; c.font=BOLD
    wc.cell(row=r,column=3,value=nt).alignment=WRAP
    for j in range(1,4): wc.cell(row=r,column=j).border=B
    r+=1
wc.cell(row=11,column=1,value="ROUGH PER-SITE USE OF FUNDS (sponsor)").font=GOLD
uof=[("Build / core plant",8000000),("Added construction",1000000),("Pre-processing equipment",2000000),("Contingency / fees / payroll / training",5000000)]
r=12
for lab,v in uof:
    wc.cell(row=r,column=1,value=lab); wc.cell(row=r,column=2,value=v).number_format=M
    for j in range(1,3): wc.cell(row=r,column=j).border=B
    r+=1
wc.cell(row=16,column=1,value="Rough subtotal").font=BOLD; wc.cell(row=16,column=2,value="=SUM(B12:B15)").number_format=M; wc.cell(row=16,column=2).font=BOLD
wc.cell(row=17,column=1,value="Gap vs $15M raise [CONFIRM]").font=BOLD; g=wc.cell(row=17,column=2,value=f"={K('raise_site')}-B16"); g.number_format=M; g.fill=OFILL
for col,w in zip("ABC",[36,18,52]): wc.column_dimensions[col].width=w

# ---------- GRANT STACK ----------
wg=wb.create_sheet("Grant Stack"); wg["A1"]="Grant stack — non-dilutive fuel around the private raise"; wg["A1"].font=TITLE
wg["A2"]="Grants are readiness/de-risking lanes, kept DISTINCT from the securities offering. They do not replace the core raise."; wg["A2"].font=SUB
hrow(wg,4,["Lane","Agency","Best fit","Current posture / status","PublicLogic role"])
gr=[
 ["MI Healthy Climate Challenge #4 — Go Big, Go Clean","EGLE","Corridor (planning study)","LOIs submitted 5/19; full app due 7/7; carbon-credit research ineligible; needs ≥50% facility-decarb framing; confirm program status","Grant-stacking + application"],
 ["NextCycle Michigan","EGLE/MI","All sites","Active support ecosystem; award status [CONFIRM]","Coordination"],
 ["EGLE Scrap Tire Market Development","EGLE","Equipment funding","FY27 target; ~50% cost share — this lane CAN fund equipment","Eligibility + application"],
 ["EPA Brownfield Cleanup & Assessment","EPA","Flint/Dort; Coldwater","$500K–$2M; environmental-readiness / diligence support","Readiness"],
 ["USDA REAP / B&I Guaranteed Loans","USDA","Coleman / rural sites","Up to $1M grant / $25M guarantee; site eligibility [CONFIRM]","Stacking"],
 ["NMTC / Opportunity Zone / EDA","Treasury/EDA","Flint / all sites","Structuring + infrastructure lanes; tract/cycle checks remain","Structuring"],
 ["§45Z / RIN (clean-fuel credits)","IRS/EPA","Operating sites","POST-qualification upside (~$20.4M/yr); RFS registration + §45Z eligibility required — NOT banked base case","Compliance / MRV"],
]
r=5
for row in gr:
    for j,v in enumerate(row):
        c=wg.cell(row=r,column=j+1,value=v); c.alignment=WRAP; c.border=B
    r+=1
wg.cell(row=13,column=1,value="Note: §45Z + RIN are the operating clean-fuel credits in the model. §45Q (carbon capture) is a different mechanism and is NOT banked here. Grant proceeds reduce net investor exposure; they are not part of the securities offering.").font=MUTE
wg.cell(row=13,column=1).alignment=WRAP; wg.merge_cells("A13:E14")
for col,w in zip("ABCDE",[34,12,20,46,18]): wg.column_dimensions[col].width=w

# ---------- CAVEATS & OPEN ITEMS ----------
wv=wb.create_sheet("Caveats & Open Items"); wv["A1"]="Open items — commercial, not modeling"; wv["A1"].font=TITLE
hrow(wv,3,["#","Item","Severity","Resolution / owner"])
cav=[
 ["1","Staffing reconciliation: Pro Forma 5 FTE ($425k) vs WasteWerx staffing pro forma 62 FTE ($5.52M)","CRITICAL","Confirm actual headcount (cross-trained middle); integrate into site model — Robert/Vinny/Glen"],
 ["2","RFS / §45Z: $20,428,560 of revenue is registration-contingent","CRITICAL","EPA RFS pathway petition + §45Z qualification (Phase 1) — Glen/counsel"],
 ["3","Carbon black = $0 in executed Pro Forma","HIGH","Confirm grade + contracted price before representing — Glen/WasteWerx"],
 ["4","Feedstock swing: $940k (PF) vs $3.72M (deck) vs ~$1M (ERR $25/ton tipping)","HIGH","Resolve with ERR supply terms — Robert"],
 ["5","Securities: $47.5M solicitation; return structure unreconciled ($14M vs $40M)","CRITICAL","Securities counsel + Reg D/PPM before any send — counsel; PL not placement agent"],
 ["6","Site list: active set Lincoln/Flint/Coleman (LOIs); Coldwater = Phase 2","HIGH","Align deck/workbook to submitted record — team"],
 ["7","Community benefit %: 10% of EBITDA vs 10% of equity (differ by ~10x)","MEDIUM","Reconcile + encode in operating agreements — Robert/counsel"],
 ["8","License not depreciable (title stays with WasteWerx); MAR Yr1 70%/Yr2+ 80%","MEDIUM","Brief Walton on tax treatment; SPV transfer terms — counsel"],
 ["9","Use-of-funds ~$16M vs $15M raise","LOW","Reconcile which line absorbs ~$1M — Robert"],
]
r=4
for row in cav:
    for j,v in enumerate(row):
        c=wv.cell(row=r,column=j+1,value=v); c.alignment=WRAP; c.border=B
        if j==2: c.fill={"CRITICAL":OFILL,"HIGH":IFILL,"MEDIUM":GFILL,"LOW":GFILL}.get(v,GFILL)
    r+=1
for col,w in zip("ABCD",[4,52,12,52]): wv.column_dimensions[col].width=w

# ---------- DOCUMENTS ON FILE ----------
wd=wb.create_sheet("Documents on File"); wd["A1"]="Documents & support on file"; wd["A1"].font=TITLE
wd["A2"]="Source-backed. Draft letters marked DRAFT; do not represent draft support as commitment."; wd["A2"].font=SUB
hrow(wd,4,["Document","Party","Date","Status"])
docs=[
 ["Executed WasteWerx License Pro Forma (Model 3)","WasteWerx","2026-05","CONFIRMED (basis of model)"],
 ["WasteWerx Mutual NDA","WasteWerx / PublicLogic","2026-05-18","EXECUTED"],
 ["WasteWerx Brochure (IP-Protected)","WasteWerx","2026","ON FILE — CONFIDENTIAL (SRC-019)"],
 ["Letter of Support","Geocycle (Paolo Carollo)","2026-01-26","ON FILE"],
 ["Letter of Support","University of Arizona (Dr. Joel Cuello)","2026-02-10","ON FILE"],
 ["Advisory validation","Acadia Capital (Brien Walton)","2026-05-08","ON FILE (validation, not committed financing)"],
 ["Feedstock supply (Dort Hwy)","ERR Inc.","2026-05","10-yr, 40,000 t/yr at $25/ton tipping"],
 ["EGLE LOIs — Lincoln / Flint / Coleman","EM&S / PublicLogic","2026-05-19","SUBMITTED"],
 ["LOS — National Salvage / Lake State / Acadia / WasteWerx","Partners","2026","DRAFT — mark draft, not commitment"],
]
r=5
for row in docs:
    for j,v in enumerate(row):
        c=wd.cell(row=r,column=j+1,value=v); c.alignment=WRAP; c.border=B
    r+=1
for col,w in zip("ABCD",[44,28,12,44]): wd.column_dimensions[col].width=w

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"final_deck","Michigan LTC - Investor Workbook FINAL (2026-06-05).xlsx")
os.makedirs(os.path.dirname(out),exist_ok=True)
wb.save(out); print("wrote",out)
