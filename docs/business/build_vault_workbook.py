#!/usr/bin/env python3
"""
Builds the Michigan LTC VAULT Workbook — the CaseSpace control surface.
16 tabs + Dashboard, pre-populated with the reconciled record from this engagement
(canonical financials, sites, partners, grants, decisions, claims, risks, reuse).
Operating principle: every claim/number/role traces to a source file + workbook row + output.
Not financial/legal/securities advice. Figures keyed to Canonical_Financial_Model v2.0.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

HDR=Font(bold=True,color="FFFFFF",size=11); HFILL=PatternFill("solid",fgColor="1F2D4D")
TITLE=Font(bold=True,size=14,color="1F2D4D"); SUB=Font(italic=True,size=9,color="555555")
GREEN=PatternFill("solid",fgColor="E2EFDA"); ORANGE=PatternFill("solid",fgColor="FCE4D6"); RED=PatternFill("solid",fgColor="F4C7C3")
WRAP=Alignment(wrap_text=True,vertical="top"); THIN=Side(style="thin",color="CCCCCC")
B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

wb=openpyxl.Workbook()

def tab(name, title, headers, rows, widths, note=None):
    ws=wb.create_sheet(name)
    ws["A1"]=title; ws["A1"].font=TITLE
    if note: ws["A2"]=note; ws["A2"].font=SUB
    hr=4
    for i,h in enumerate(headers):
        c=ws.cell(row=hr,column=i+1,value=h); c.font=HDR; c.fill=HFILL; c.border=B; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")
    for ri,row in enumerate(rows):
        for ci,val in enumerate(row):
            c=ws.cell(row=hr+1+ri,column=ci+1,value=val); c.border=B; c.alignment=WRAP
    for i,w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width=w
    ws.freeze_panes=f"A{hr+1}"
    return ws

# 00 DASHBOARD
ws=wb.active; ws.title="00 Dashboard"
ws["A1"]="Michigan LTC Network — Resource Recovery Corridor · VAULT Workbook"; ws["A1"].font=TITLE
ws["A2"]="Control surface. Original files stay preserved · this workbook controls the record · deliverables come from the record. Date 2026-06-05."; ws["A2"].font=SUB
dash=[
 ("Corridor Name","Michigan LTC Network — Resource Recovery Corridor"),
 ("Public Description","One corridor converting waste streams into recoverable fuel and industrial outputs"),
 ("Active Sites","Lincoln (Alcona) · Coleman (Midland) · Flint (Genesee)"),
 ("Lead Applicant","Energy Mann & Sunn, Inc. DBA Michigan LTC"),
 ("Project Executive","Garrett Schopp (EGLE portal contact)"),
 ("Operations Lead","Robert C. McCall Jr."),
 ("Developer / Financial","Glen D. Scharer / Carbon Alliance Group"),
 ("Technology","WasteWerx LLC (licensed; IP retained by WasteWerx)"),
 ("Governance / Compliance","PublicLogic LLC (not a placement agent; fixed/non-contingent fee)"),
 ("Current Phase","Phase 1 — Discovery & Systems Mapping"),
 ("Current Priority","Reconcile staffing & financial model; resolve carbon-black + RFS/45Z; EGLE path"),
 ("Headline economics","Signed Pro Forma EBITDA $30,200,531 (5-FTE basis); operator-reality ~$22,047,043; base-fuel-only floor ~$1,618,483"),
 ("Current Risk","Staffing (5 vs 62 FTE), feedstock swing, RFS/45Z contingency ($20.4M), securities, source consistency"),
 ("Next Deliverables","VAULT workbook · proforma reconciliation · grant map · investor/partner packet · data-room index"),
]
r=4
for f,v in dash:
    ws.cell(row=r,column=1,value=f).font=Font(bold=True); ws.cell(row=r,column=1).fill=GREEN; ws.cell(row=r,column=1).border=B
    c=ws.cell(row=r,column=2,value=v); c.alignment=WRAP; c.border=B
    r+=1
ws.column_dimensions["A"].width=24; ws.column_dimensions["B"].width=92
ws.freeze_panes="A4"

# 01 SOURCE REGISTER
tab("01 Source Register","01 — Source Register",
 ["Source ID","File Name","Folder Path","Type","Date","From","Status","Used In","Notes"],
 [
 ["SRC-001","Canonical_Financial_Model v2.0.xlsx","04_FINANCE/01_Original_Models","Financial model","2026-05","Glen / PL","ACTIVE — basis of record","Revenue, P&L, Corridor","Keyed line-for-line to executed WasteWerx Model 3 Pro Forma"],
 ["SRC-002","Financial_Reconciliation_Workbook.xlsx","04_FINANCE/02_Reconciled_Models","Reconciliation","2026-05","PL","ACTIVE","Discrepancy log, WasteWerx cost inputs","3 sources -> one truth; 8 critical discrepancies"],
 ["SRC-003","Governance_Workbook.xlsx","00_CONTROL","Governance","2026-05","PL","ACTIVE","VAULT / control","Deal dashboard + VAULT exposure"],
 ["SRC-004","Investor_Brief.docx (EM&S deck)","08_PUBLIC_LANGUAGE/02_Public_Brief","Investor deck","2026-05","EM&S / PL","SUPERSEDED (economics)","Narrative / partners","Uses old $16.875M oil+CB model; credits excluded — superseded by SRC-001"],
 ["SRC-005","WasteWerx Staffing Proforma.xlsx","04_FINANCE/04_Staffing_Proforma","Proforma","2026-05-27","Vinny Tizio / WasteWerx","NEEDS RECONCILIATION","Staffing model","62 FTE worst-case 'overkill'; cross-train reduces; MI rates + benefits pending"],
 ["SRC-006","EGLE Challenge #4 Application (EQP1168)","03_GRANTS/01_EGLE/00_Guidance","Grant guidance","2026-06-02","EGLE","ACTIVE","Full application","Due 7/7/2026; planning-only; no equipment"],
 ["SRC-007","EGLE Challenge #4 Scoring Rubric","03_GRANTS/01_EGLE/00_Guidance","Grant guidance","2026","EGLE","ACTIVE","Application strategy","Emissions 30% / Facility 20% / Community 15% / Tech 15% / Replicability 10% / Team 10%"],
 ["SRC-008","EGLE Challenge #4 LOI Feedback","03_GRANTS/01_EGLE/00_Guidance","Grant guidance","2026","EGLE","ACTIVE","Pivot analysis","Carbon-credit research INELIGIBLE; >=50% facility decarb; process change focus"],
 ["SRC-009","MI_LTC_LOI_PreFill_FINAL","03_GRANTS/01_EGLE/01_PreFill","LOI pre-fill","2026-05-19","Garrett / PL","SUBMITTED","3 LOIs","Lincoln/Flint/Coleman; $420k/site"],
 ["SRC-010","Lincoln Site Brief / EGLE Project D","02_SITES/01_Lincoln","Site brief","2026-05","PL / team","ACTIVE","Lincoln LOI","National Energy of Lincoln 18 MW biomass; PPA exp 2027"],
 ["SRC-011","Coleman Site Brief / CM Rubber","02_SITES/02_Coleman","Site brief","2026-05","PL / team","ACTIVE","Coleman LOI","Geocycle / CM Rubber; Part 169; confirm naming"],
 ["SRC-012","Flint Site Brief / Dort Highway","02_SITES/03_Flint","Site brief","2026-05","PL / team","ACTIVE","Flint LOI","ERR; EJ community; brownfield"],
 ["SRC-013","LOS_Michigan_LTC_Network_Final.docx","07_PARTNERS_AND_ROLES/01_Role_Map","Support letter","2026-05","Coalition","ON FILE","Grant / partner packet","Joint letter of support"],
 ["SRC-014","WasteWerx Mutual NDA (executed)","11_LEGAL_AND_ENTITY/05_NDAs","NDA","2026-05-18","WasteWerx / PL","EXECUTED","IP protection","Protect WasteWerx proprietary tech"],
 ["SRC-015","Geocycle Letter of Support","07_PARTNERS_AND_ROLES/06_Geocycle","Support letter","2026-01-26","Paolo Carollo","ON FILE","NextCycle / corridor","Coleman co-location support"],
 ["SRC-016","University of Arizona Support (Dr. Cuello)","07_PARTNERS_AND_ROLES","Support letter","2026-02-10","Dr. Joel Cuello","ON FILE","Community / workforce","ACS aquaponic programming"],
 ["SRC-017","Acadia Capital validation (Walton)","04_FINANCE/09_Investor_Readiness","Advisory validation","2026-05-08","Brien Walton","ON FILE","Investor readiness","Named governance the critical success factor"],
 ["SRC-018","ERR Feedstock Supply (Dort Hwy)","02_SITES/03_Flint/Source","Supply terms","2026-05","ERR Inc.","ON FILE","Feedstock / finance","10-yr, 40,000 t/yr at $25/ton tipping"],
 ["SRC-019","WasteWerx Brochure (IP-Protected)","07_PARTNERS_AND_ROLES/05_WasteWerx","Technology brochure","2026","WasteWerx","ON FILE — CONFIDENTIAL","Technology / partner packet","IP-protected version; partner-facing only; do NOT publish proprietary specs/reactor photos (NDA 5/18)"],
 ],
 [10,30,30,14,11,16,20,22,40],
 "Most important tab. If a number, claim, role, cost, site fact, or public sentence matters, it points to a Source ID.")

# 02 ENTITY + ROLE MAP
tab("02 Entity + Role Map","02 — Entity + Role Map",
 ["Role","Person / Entity","Function","Authority","Public Facing?","Notes"],
 [
 ["Applicant of Record","Energy Mann & Sunn, Inc. DBA Michigan LTC","Grant applicant / project vehicle","Legal / grant","Yes","LOI submissions"],
 ["Project Executive","Garrett Schopp","EGLE-facing lead / National Salvage","Operational / external","Yes","Portal contact"],
 ["Operations Lead","Robert C. McCall Jr.","Corridor operations / EM&S principal","Operational","Yes","EM&S signatory"],
 ["Developer / Financial","Glen D. Scharer / Carbon Alliance Group","Financial development / capital structure","Strategic","Maybe","Keep clean publicly"],
 ["Technology Provider","WasteWerx LLC (JT Clark, V. Tizio, E. Rodriguez)","Pyrolysis system (licensed) / proforma source","Technical","Yes","IP retained by WasteWerx; validate assumptions"],
 ["Governance / Compliance","PublicLogic LLC (Boudreau, Rothschild)","Governance, documentation, grant-stacking, continuity","Advisory","Yes","NOT placement agent; fixed/non-contingent fee"],
 ["Legal Agent","Michael J. Smith","Registered / legal agent","Legal","Limited","Entity records"],
 ["Facility Partner","Geocycle / CM Rubber Recycling (Paolo Carollo)","Coleman site context","Site partner","Yes","Part 169; confirm exact language"],
 ["Facility Partner","Environmental Rubber Recycling (ERR)","Flint site context + feedstock","Site partner","Yes","10-yr, 40k t/yr, $25/ton tipping"],
 ["Facility Partner","National Energy of Lincoln LLC","Lincoln site context","Site partner","Yes","18 MW biomass; PPA exp 2027"],
 ["Strategic Partner","National Salvage & Service Corp.","Lead applicant affiliate / feedstock","Strategic","Yes","Garrett affiliation; McBain/Lincoln biomass"],
 ["Capital Advisory","Acadia Capital Management (Brien Walton)","Investment-readiness validation","Advisory","Maybe","Not committed financing"],
 ],
 [22,38,38,20,14,40])

# 03 SITE REGISTER
tab("03 Site Register","03 — Site Register",
 ["Site ID","Site","County","Facility / Partner","Feedstock","Status","Grant Use","Notes"],
 [
 ["SITE-001","Lincoln","Alcona","National Energy of Lincoln LLC (18 MW biomass)","Tires / biomass / railroad ties (TBD)","ACTIVE","Feasibility / corridor","First LOI priority; PPA exp 2027; ~21 jobs"],
 ["SITE-002","Coleman","Midland","Geocycle / CM Rubber Recycling","Tires / industrial material (TBD)","ACTIVE","Feasibility / corridor","Part 169; naming must be consistent"],
 ["SITE-003","Flint","Genesee","Environmental Rubber Recycling (Dort Hwy)","Tire / rubber stream","ACTIVE","Feasibility / corridor","EJ community; ERR role needs clean language"],
 ["SITE-004","Coldwater","Genesee (Genesee Twp)","RACER #13270 (TBD)","TBD","INACTIVE / Phase 2","None","Do not reintroduce unless revived; keep for memory"],
 ],
 [10,12,16,38,30,16,20,40],
 "Keep removed sites in the register (inactive), never delete — preserves institutional memory.")

# 04 GRANT REGISTER
tab("04 Grant Register","04 — Grant Register",
 ["Grant ID","Program","Agency","Status","Ask","Match","Site(s)","Deadline","Lead","Notes"],
 [
 ["GRANT-001","MI Healthy Climate Challenge #4 — Go Big, Go Clean","EGLE","LOIs submitted; full app due 7/7","$420,000/site; $1,260,000 total","TBD","Lincoln, Coleman, Flint","Full app 7/7/2026","Garrett","Planning-only; carbon credits ineligible; needs >=50% facility decarb framing"],
 ["GRANT-002","Scrap Tire Market Development Grant","EGLE / Michigan","Future target","TBD","~50% cost share","TBD","FY27","TBD","Equipment funding path (this one CAN fund equipment)"],
 ["GRANT-003","USDA REAP / B&I","USDA","Research","Up to $1M grant / $25M guarantee","TBD","Coleman (rural)","TBD","TBD","Rural facility path"],
 ["GRANT-004","EPA Brownfield Cleanup & Assessment","EPA","Research","$500K-$2M","20% (waivable)","Flint / Dort","TBD","TBD","Brownfield reuse"],
 ["GRANT-005","EDA Public Works / NextCycle MI","EDA / MI","Research","Up to $3M / $50-250K","TBD","Corridor","NextCycle applied Mar 2026","TBD","Future stack"],
 ],
 [10,40,14,22,24,16,20,18,10,40])

# 05 FINANCIAL ASSUMPTIONS (reconciled to canonical)
tab("05 Financial Assumptions","05 — Financial Assumptions (keyed to Canonical v2.0)",
 ["ID","Item","Value","Source","Status","Notes"],
 [
 ["FIN-001","Total revenue (Pro Forma)","$35,581,680","SRC-001","CONFIRMED","Fuel + RIN + 45Z"],
 ["FIN-002","Base fuel (RD $4,798,080 + SAF $10,355,040)","$15,153,120","SRC-001","CONFIRMED","Sells on contract/market"],
 ["FIN-003","RIN + Section 45Z credits","$20,428,560","SRC-001","CONTINGENT","Flag 3: requires EPA RFS registration — not guaranteed"],
 ["FIN-004","Carbon black","$0 (base)","SRC-001","OPEN — Flag 1","Not in executed Pro Forma; deck implied ~$10.125M; uncontracted"],
 ["FIN-005","Opex — Pro Forma (as signed, 4 FTE + manager)","$5,381,149","SRC-001","AS SIGNED","Labor only $425,000 — see staffing tension"],
 ["FIN-006","Opex — operator reality (62-FTE staffing + feedstock)","~$13,534,637","SRC-001/SRC-005","STRESS","Delta ~$8.15M vs Pro Forma"],
 ["FIN-007","EBITDA — signed (PF opex)","$30,200,531","SRC-001","AS SIGNED","Rests on a 5-person crew"],
 ["FIN-008","EBITDA — operator reality (fuel+credits)","~$22,047,043","Derived","DILIGENCE REALITY","Real staffing + feedstock"],
 ["FIN-009","EBITDA — base fuel only @ operator opex","~$1,618,483","Derived","FLOOR","Without credits, barely breakeven"],
 ["FIN-010","Year 1 cash out","$7,622,828","SRC-001","CONFIRMED","WasteWerx $6,670,328 + capex $952,500"],
 ["FIN-011","WasteWerx license fee","$5,500,000","SRC-001/002","CONFIRMED","Net of $50k design credit; NOT depreciable"],
 ["FIN-012","Production royalty","$0.60/gallon","SRC-001","CONFIRMED","IoT-metered"],
 ["FIN-013","Monthly monitoring","$35,000/month","SRC-001","CONFIRMED","CPI-indexed, from commissioning"],
 ["FIN-014","MAR","Yr1 70% / Yr2+ 80%","SRC-002","CONFIRMED","Of nameplate royalty"],
 ["FIN-015","Staffing (worst-case)","62 FTE / $5,522,400","SRC-005","NEEDS RECONCILIATION","Vincent: worst-case; MI rates + cross-training + benefits pending"],
 ["FIN-016","Feedstock swing","$940k PF / $3.72M deck / ~$1M ERR tipping","SRC-001/002/018","OPEN","ERR $25/ton tipping would lower opex"],
 ["FIN-017","Simple payback","~0.25 yr","SRC-001","EXPLAIN","On signed basis; aggressive — caveat required"],
 ["FIN-018","Community profit share","10% of EBITDA","SRC-001","CONFIRMED","Contractual, locally directed"],
 ],
 [9,46,28,16,18,46],
 "Separates confirmed numbers / contingent credits / stress assumptions / open items. Keeps the model from becoming mythology.")

# 06 STAFFING + OPERATIONS
tab("06 Staffing + Operations","06 — Staffing + Operations (the diligence swing)",
 ["ID","Role","Pro Forma (signed)","Staffing Proforma (worst-case)","Status","Notes"],
 [
 ["OPS-000","TOTAL LABOR","4 FTE + manager = $425,000","62 FTE = $5,522,400 (FL)","OPEN","The single biggest swing in the model"],
 ["OPS-001","Plant Manager","incl. in $425k","1 FTE","Open","Confirm per site"],
 ["OPS-002","Operating labor / operators","4 FTE","~16-30 FTE (24/7 x 4 shifts)","Open","Staffing ratio drives everything"],
 ["OPS-003","Maintenance & reliability","—","~8 FTE","Open","Shared vs per-site"],
 ["OPS-004","Feedstock / yard / logistics","—","~11 FTE","Open","Biomass moisture mgmt = full-shift"],
 ["OPS-005","Fuel processing & carbon co-product","—","~14 FTE","Open","Distillation + carbon handling"],
 ["OPS-006","QA/lab, HSE, MRV, admin","—","~13 FTE","Open","MRV analyst from Day 1"],
 ["OPS-007","Benefits / burden","not detailed","30% burden assumed","Open","Vincent: health insurance needs actuals"],
 ],
 [9,32,26,30,12,42],
 "Two WasteWerx documents conflict on staffing. Resolve actual headcount (cross-trained middle) before investor numbers.")

# 07 REGULATORY REGISTER
tab("07 Regulatory Register","07 — Regulatory Register",
 ["Reg ID","Agency","Item","Site","Status","Owner","Notes"],
 [
 ["REG-001","EGLE","Pre-application transmittal","Corridor","Sent / pending","Garrett / PL","Sent to Nicole Sanabria"],
 ["REG-002","EGLE","LOI confirmation","Lincoln","Submitted","Garrett","Confirmation requested (no auto-email)"],
 ["REG-003","EGLE","LOI confirmation","Coleman","Submitted","Garrett","Same"],
 ["REG-004","EGLE","LOI confirmation","Flint","Submitted","Garrett","Same"],
 ["REG-005","EGLE","Air PTI / NSR / Title V applicability","Each site","Open","TBD","Map by site; WasteWerx low-filtration claim helps"],
 ["REG-006","EGLE","Part 169 scrap tire / solid waste","Each site","Open","Operator","Material-specific"],
 ["REG-007","Local","Site / zoning approvals","Each site","Open","Site partner","Unknown"],
 ["REG-008","Federal/Tax","RFS pathway petition + 45Z qualification","Corridor","Open — Phase 1","Glen / counsel","Gates $20.4M of revenue (FIN-003)"],
 ],
 [9,12,40,12,16,16,42])

# 08 DECISION LOG
tab("08 Decision Log","08 — Decision Log",
 ["Decision ID","Date","Decision","Made By","Evidence","Affected Outputs","Notes"],
 [
 ["DEC-001","2026-05","Corridor framing instead of replication","PL / team","Draft history","Website, grants, briefs","'Replication' removed"],
 ["DEC-002","2026-05","One corridor publicly; no hard site count in hero","PL / team","Website canon","Public copy","Internal count remains 3"],
 ["DEC-003","2026-05-19","Submit three separate LOIs","Garrett / PL / team","LOI package","EGLE submissions","Risk noted (RISK-006)"],
 ["DEC-004","2026-05","$420,000 per site ask","Team","LOI pre-fill","EGLE LOIs","Total $1,260,000 (note: <$400k encouraged)"],
 ["DEC-005","2026-05","Exclude Coldwater from active package","Team","Site register","Grants / website","Keep inactive (SITE-004)"],
 ["DEC-006","2026-05","Garrett = portal contact","Team","LOI pre-fill","EGLE portal","EM&S remains applicant"],
 ["DEC-007","2026-06","Canonical v2.0 = single source of truth for economics","PL","SRC-001","Model / investor","Supersedes EM&S deck economics (SRC-004)"],
 ["DEC-008","2026-06","Anchor investor headline on base + labeled contingent upside","PL","Reconciliation","Investor deck","Do not lead with $30.2M alone"],
 ],
 [11,11,42,18,16,22,40])

# 09 OPEN QUESTIONS
tab("09 Open Questions","09 — Open Questions (be brutally honest)",
 ["Q ID","Question","Category","Owner","Status","Needed Source"],
 [
 ["Q-001","Actual Michigan staffing: 5 (Pro Forma), 62 (staffing), or cross-trained middle?","Staffing","Robert / Vinny / Glen","OPEN","Reconciled staffing proforma"],
 ["Q-002","Are insurance/benefits actuals included in staffing burden?","Finance","Glen / Robert","OPEN","Insurance + benefits quote"],
 ["Q-003","Is carbon black additive to distilled fuel, and at what grade/price?","Finance","Glen / WasteWerx","OPEN","Feedstock yield model + offtake"],
 ["Q-004","RFS registration + 45Z qualification timeline?","Regulatory/Tax","Glen / counsel","OPEN — Phase 1","RFS petition + 3rd-party engineering"],
 ["Q-005","Feedstock: $0.20/gal, $200/ton, or $25/ton ERR tipping?","Finance","Robert / ERR","OPEN","ERR supply agreement"],
 ["Q-006","Equipment / depolymerization cost by site?","Finance/Grant","WasteWerx / Glen","OPEN","Vendor quote"],
 ["Q-007","Which investor return structure ($3.5M/4yr vs $5M/5yr+balloon)?","Finance/Legal","Team / counsel","OPEN — CRITICAL","Securities counsel"],
 ["Q-008","Parent holdco legal name + SPV chart + LARA registration?","Legal/Entity","Robert / counsel","OPEN","Entity formation docs"],
 ["Q-009","Public role each facility partner approves?","Public language","PL / Garrett","OPEN","Partner approval"],
 ],
 [9,46,16,22,16,28])

# 10 CLAIMS + EVIDENCE
tab("10 Claims + Evidence","10 — Claims + Evidence (public-safety tab)",
 ["Claim ID","Claim","Use Case","Public Safe?","Source","Status"],
 [
 ["CLAIM-001","Michigan LTC is a resource recovery corridor","Website / grant","Yes","Corridor brief","Approved"],
 ["CLAIM-002","Corridor includes Lincoln, Coleman, and Flint","Internal / grant","Conditional","Site briefs","Avoid hard count in hero"],
 ["CLAIM-003","Converts waste into fuel + co-products (licensed pyrolysis)","Grant / technical","Yes if sourced","WasteWerx (NDA)","Exact wording; no proprietary specs/images"],
 ["CLAIM-004","Steady-state EBITDA $30.2M","Investor / internal","NOT public without context","Canonical (SRC-001)","Caveat: 5-FTE basis; operator-reality ~$22M"],
 ["CLAIM-005","Payback ~3 months","Internal","Sensitive","Canonical","High scrutiny; requires caveat"],
 ["CLAIM-006","$35.58M revenue","Investor","Conditional","Canonical","$20.4M is RFS-contingent — label as upside"],
 ["CLAIM-007","Fully permitted","Any","NO — BANNED","None","Do not use unless true"],
 ["CLAIM-008","Zero-emission","Any","NO — BANNED","None","Do not use unless proven"],
 ["CLAIM-009","Carbon black revenue","Investor","NO without contract","None (Flag 1)","$0 in executed pro forma"],
 ["CLAIM-010","Legacy is what the next generation can use","Website / brand","Yes","Canon","Approved — emotional center"],
 ],
 [10,44,18,22,18,34],
 "How you avoid accidental overclaiming. Banned claims stay listed.")

# 11 DELIVERABLES REGISTER
tab("11 Deliverables Register","11 — Deliverables Register",
 ["Del ID","Deliverable","Audience","Status","Owner","Notes"],
 [
 ["DEL-001","EGLE LOI — Lincoln/Coleman/Flint","EGLE","Submitted 5/19","Garrett / PL","Three separate LOIs"],
 ["DEL-002","EGLE Challenge #4 Lincoln Full Application (draft)","EGLE","Draft","PL","Pivoted to facility-decarb; due 7/7"],
 ["DEL-003","Canonical Financial Model (reconciled)","Internal / investor","Current","PL / Glen","SRC-001 basis; base vs operator-reality"],
 ["DEL-004","Headline Economics Memo (McCall/Scharer)","Internal","Current","PL","Base vs contingent; what to prove"],
 ["DEL-005","VAULT Assessment + Register","Internal / partners","Current","PL","5 pillars; readiness rollup"],
 ["DEL-006","Path-to-All-A + Investment-Grade Scorecard","Internal","Current","PL","C+ -> A roadmap"],
 ["DEL-007","Investor Deck Prompt + draft package","Internal","Draft","PL","Securities + IP guardrails baked in"],
 ["DEL-008","Staffing Reconciliation Brief","Glen / Robert","Planned","PL","Michigan rates + cross-training"],
 ["DEL-009","Data-Room Index","Investor / partners","Planned","PL","Catalog financials + support letters (this workbook's companion)"],
 ["DEL-010","Corridor Partner Packet","Partners / funders","Planned","PL","Needs clean role map (tab 02)"],
 ],
 [9,42,20,14,16,44])

# 12 PUBLIC LANGUAGE
tab("12 Public Language","12 — Public Language",
 ["Item","Approved Phrase","Avoid","Notes"],
 [
 ["Corridor","Michigan LTC Network — Resource Recovery Corridor","Three-plant rollout (unless context)","Public-facing"],
 ["Sites","Lincoln, Coleman, and Flint (when needed)","Hard count in hero","Internal count is 3"],
 ["Technology","Resource recovery using licensed pyrolysis technology","Magic / revolutionary / guaranteed","IP retained by WasteWerx"],
 ["Climate","Cleaner industrial recovery pathway","Zero-emission (unless proven)","Avoid overclaim"],
 ["Economics","Model-based project economics (assumptions stated)","Guaranteed returns / unqualified $30.2M","Lead with base; credits as upside"],
 ["Legacy","Legacy is what the next generation can use","Buzzword sustainability","Emotional center"],
 ["Public proof","Documentation-first, verification-ready","Trust us","Strong posture"],
 ["Mantra","Carbon to Capacity. Capital to Community. Continuity for the next generation.","Over-explaining entity structure","Use sparingly"],
 ],
 [16,52,34,28])

# 13 RISK REGISTER
tab("13 Risk Register","13 — Risk Register",
 ["Risk ID","Risk","Severity","Category","Mitigation","Owner"],
 [
 ["RISK-001","Financial model looks too strong without support","High","Finance","Base + sensitivity + source notes (SRC-001)","Glen / PL"],
 ["RISK-002","Staffing generic, not Michigan-adjusted (5 vs 62 FTE)","High","Operations","Reconcile proforma (Q-001)","Robert / Vinny"],
 ["RISK-003","Public language overclaims readiness","High","Communications","Claims register (tab 10)","PL"],
 ["RISK-004","RFS/45Z contingency = $20.4M of revenue","High","Regulatory/Finance","RFS petition Phase 1; label as upside","Glen / counsel"],
 ["RISK-005","Securities offering with no Reg D / PPM","High","Legal","Securities counsel; PL not placement agent","Counsel / PL"],
 ["RISK-006","Three LOIs vs corridor logic","Medium","Grants","Document rationale (DEC-003)","PL"],
 ["RISK-007","Carbon black narrative vs model ($0)","Medium","Finance","Clarify inclusion (Q-003)","Glen"],
 ["RISK-008","Partner roles blur across entities","Medium","Governance","Maintain role map (tab 02)","PL / Robert"],
 ["RISK-009","WasteWerx IP exposure (licensed)","Medium","Legal/IP","NDA + generic visuals + license disclosure","PL / WasteWerx"],
 ["RISK-010","Old site (Coldwater) reappears","Low","Record control","Inactive site register (SITE-004)","PL"],
 ],
 [9,44,10,16,40,16])

# 14 CONTACT REGISTER
tab("14 Contact Register","14 — Contact Register",
 ["Contact","Entity","Role","Email","Notes"],
 [
 ["Robert C. McCall Jr.","EM&S","Operations lead / signatory","mccall@acscoop.com","216-906-2907"],
 ["Garrett Schopp","National Salvage","Project Executive / EGLE lead","Robert.Mccall@nssccorp.com (portal)","812-345-2543"],
 ["Glen D. Scharer","Carbon Alliance Group","Developer / financial","glen@carbonalliance.earth","863-255-8227"],
 ["Vincent (Vinny) Tizio II","WasteWerx","Staffing proforma / technology","vinny@wwerx.com","941-600-3018"],
 ["Nathan Boudreau","PublicLogic","Governance / documentation","nate@publiclogic.org","978-807-0829"],
 ["Dr. Allison Weiss Rothschild","PublicLogic","Org systems / continuity","allie@publiclogic.org","TBD"],
 ["Michael J. Smith","Legal","Agent / attorney","TBD","Entity"],
 ["Nicole Sanabria","EGLE","Regulatory contact","TBD","Pre-app transmittal"],
 ["Haley Neuenfeldt","EGLE","Challenge #4 program contact","NeuenfeldtH@Michigan.gov","LOI / application"],
 ["Brien Walton","Acadia Capital","Capital advisory","TBD","Validation 5/8/26"],
 ["Paolo Carollo","Geocycle","Coleman host","TBD","LOS 1/26/26"],
 ],
 [24,18,28,34,20])

# 15 CHANGE LOG
tab("15 Change Log","15 — Change Log (what changed in the record)",
 ["Change ID","Date","Item Changed","From","To","Reason"],
 [
 ["CHG-001","2026-05","Grant ask","$600K/$550K/$500K concepts","$420K/site","Fit total ask"],
 ["CHG-002","2026-05","Public site count","Hard count","One corridor","Public canon"],
 ["CHG-003","2026-05","Framing","Replication","Corridor standard","Strategic language"],
 ["CHG-004","2026-05","Site list","Included Coldwater","Removed Coldwater","Active package narrowed"],
 ["CHG-005","2026-05","Applicant framing","Partner-specific","EM&S DBA Michigan LTC","Submission consistency"],
 ["CHG-006","2026-06","Revenue model basis","EM&S deck $16.875M (oil+CB)","Canonical $35.58M (fuel+credits)","Executed Pro Forma"],
 ["CHG-007","2026-06","Opex treatment","Single $5.38M (5 FTE)","Two-case: PF vs operator-reality (62 FTE)","Staffing tension surfaced"],
 ["CHG-008","2026-06","Credit portion","~$13.4M (recon note)","$20,428,560 (canonical line items)","Canonical exact split"],
 ],
 [10,11,24,34,34,28])

# 16 REUSE LIBRARY
tab("16 Reuse Library","16 — Reuse Library (the future advantage)",
 ["Reuse ID","Asset","Can Reuse For","Status","Notes"],
 [
 ["REUSE-001","Corridor description","Grants, website, partner packet","Approved","Keep source-backed"],
 ["REUSE-002","Site brief format","Future sites","Approved","Lincoln/Coleman/Flint template"],
 ["REUSE-003","EGLE LOI answer pattern","Future grants","Approved","Adapt carefully"],
 ["REUSE-004","Role map format","Partner packets","Approved","Prevents confusion"],
 ["REUSE-005","Public-safe claims table","Website / PR","Approved","Avoids overclaiming"],
 ["REUSE-006","Two-case opex / staffing reconciliation","Other industrial projects","Draft","Needs final proforma"],
 ["REUSE-007","Canonical-vs-reality financial structure","Investor memos","Draft","Base + labeled contingent upside"],
 ["REUSE-008","VAULT Workbook (this file)","Every corridor / state SPV","Approved","The control surface template"],
 ],
 [10,40,30,12,40])

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"casespace","00_CONTROL","Michigan_LTC_VAULT_Workbook.xlsx")
os.makedirs(os.path.dirname(out),exist_ok=True)
wb.save(out); print("wrote",out,"— sheets:",len(wb.sheetnames))
