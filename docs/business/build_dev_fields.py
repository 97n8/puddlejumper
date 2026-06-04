#!/usr/bin/env python3
"""
Builds Development_Fields.xlsx — 10 development fields beyond circular-economy/pyrolysis
where PublicLogic's model works: a business-funded (or grant-funded) project where PL
builds the governance + stacks the capital + stewards the compliance tail. The thesis:
PublicLogic can help build almost ANY capital-stack development, because in most fields
you can grant-stack to assemble the money and steward to hold it together over time.

Each field is scored & ranked (live formulas) and detailed with the grant/credit stack,
the compliance/stewardship tail, and PL's fee route — all consistent with the verified
federal funding architecture in ../federal-funding/ (NMTC permanent, 45Q $85/ton, LIHTC
expanded, OZ permanent, CWSRF/DWSRF + WIFIA active, Brownfields live, 48E/45Y sunsetting).

Scoring (1-5, 5=best): MARKET, STACK (capital-stack richness), TAIL (compliance/steward
recurrence), FIT (PublicLogic capability fit), TAILWIND (policy direction), ENTRY (ease of
entry / existing proof). Composite = weighted avg; Rank is live.

Not legal/accounting advice. Figures/program statuses verified as of 2026-06-03; re-verify.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="3B5E2B")
SUB = Font(bold=True, size=11, color="3B5E2B")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE = Font(italic=True, size=9, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM2 = '0.00'


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER


def title(ws, text, sub=None):
    ws["A1"] = text; ws["A1"].font = Font(bold=True, size=14, color="3B5E2B")
    if sub:
        ws["A2"] = sub; ws["A2"].font = NOTE


wb = openpyxl.Workbook()

# README -----------------------------------------------------------------------
ws = wb.active; ws.title = "README"
title(ws, "PublicLogic — 10 Development Fields the Model Works For",
      "The model is field-agnostic: build the governance, stack the capital, steward the tail. "
      "Live scored & ranked. Date: 2026-06-03. Figures illustrative; program statuses verified.")
for r, ln in enumerate([
    "", "THE THESIS",
    "  PublicLogic's model is not about pyrolysis. It is about any capital-stack development where you can",
    "  (1) GRANT-STACK to assemble the money, and (2) STEWARD the multi-year compliance tail. In most",
    "  development fields both are true — so PublicLogic can help build almost anything, the same cyclical way:",
    "  Discover -> Honor -> Understand -> Improve -> Build -> Steward -> Continue.",
    "", "THE REPEATABLE MOVE (every field)",
    "  Cheap diagnostic -> stack the capital (tax credits + grants + loans) -> build the governance spine ->",
    "  hold the compliance tail as a retainer. PL's fee rides inside the capital as a soft/compliance cost,",
    "  fixed & non-contingent. The VAULT module library makes each new field cheaper to enter than the last.",
    "", "SHEETS",
    "  Inputs   - scoring weights (edit to re-rank).",
    "  Fields   - the 10 fields, scored & ranked (live composite + RANK).",
    "  Stacks   - per field: the grant/credit stack, the compliance/steward tail, PL's fee route, the hook.",
    "", "GUARDRAILS (carry from the funding architecture)",
    "  * Keep PL fees FIXED & NON-CONTINGENT on federal money (2 CFR 200.459(b)); grant-WRITING stays own-source (200.460).",
    "  * On capital raises, PL is governance, not placement (securities lane).",
    "  * Distinguish bedrock (NMTC permanent, 45Q, LIHTC, OZ, SRF/WIFIA, Medicaid) from sunsetting/volatile",
    "    (48E/45Y wind-solar after 2027; BEAD/SLCGP/BRIC discretionary). Anchor deals on bedrock.",
    "", "Not legal/accounting/securities advice.",
], start=4):
    ws.cell(row=r, column=1, value=ln)
    if ln.strip() in ("THE THESIS", "THE REPEATABLE MOVE (every field)", "SHEETS", "GUARDRAILS (carry from the funding architecture)"):
        ws.cell(row=r, column=1).font = SUB
ws.column_dimensions["A"].width = 106

# INPUTS -----------------------------------------------------------------------
wi = wb.create_sheet("Inputs")
title(wi, "Inputs — scoring weights", "Edit yellow weights to re-rank. Should sum to 1.00.")
wi.append([]); wi.append(["Code", "Criterion", "Weight"]); style_header(wi, wi.max_row, 3)
weights = [
    ("MARKET", "Market size / volume of projects", 0.18),
    ("STACK", "Capital-stack richness (non-dilutive layers available)", 0.20),
    ("TAIL", "Compliance / stewardship tail (PL recurring revenue)", 0.20),
    ("FIT", "Fit with PublicLogic capability (governance + Boudreau/Rothschild)", 0.18),
    ("TAILWIND", "Policy tailwind (bedrock/permanent vs sunset/volatile)", 0.12),
    ("ENTRY", "Ease of entry / existing proof & relationships", 0.12),
]
wkey = {}
start = wi.max_row + 1
for i, (code, label, w) in enumerate(weights):
    rr = start + i
    wi.cell(row=rr, column=1, value=code).font = Font(bold=True, color="888888")
    wi.cell(row=rr, column=2, value=label)
    c = wi.cell(row=rr, column=3, value=w); c.fill = INPUT_FILL; c.border = BORDER; c.number_format = '0.00'
    wkey[code] = f"Inputs!$C${rr}"
sumrow = start + len(weights)
wi.cell(row=sumrow, column=2, value="SUM (must = 1.00)").font = SUB
sc = wi.cell(row=sumrow, column=3, value=f"=SUM(C{start}:C{sumrow-1})")
sc.fill = TOTAL_FILL; sc.font = Font(bold=True); sc.number_format = '0.00'
for col, w in zip("ABC", [10, 56, 10]):
    wi.column_dimensions[col].width = w

# FIELDS (scored) --------------------------------------------------------------
wf = wb.create_sheet("Fields")
title(wf, "10 Development Fields — scored & ranked (live)",
      "Composite = weighted avg of the six 1-5 scores. Rank is live. Edit scores/weights to re-rank.")
hdr = ["#", "Development field", "Representative project", "MARKET", "STACK", "TAIL", "FIT", "TAILWIND",
       "ENTRY", "Composite", "Rank"]
hrow = 4
for j, h in enumerate(hdr, start=1):
    wf.cell(row=hrow, column=j, value=h)
style_header(wf, hrow, len(hdr))

# (field, project, MARKET, STACK, TAIL, FIT, TAILWIND, ENTRY)
fields = [
    ("Affordable & workforce housing", "LIHTC apartments / mixed-income redevelopment", 5, 5, 5, 4, 5, 3),
    ("Healthcare & behavioral-health facilities", "FQHC / CCBHC / clinic (Rothschild lane)", 5, 4, 5, 5, 4, 3),
    ("Water, wastewater, PFAS & lead-line", "Treatment upgrade / LSL replacement / PFAS", 5, 5, 5, 4, 4, 4),
    ("Broadband & municipal digital infrastructure", "Municipal fiber / middle-mile (Shrewsbury)", 4, 4, 4, 5, 3, 5),
    ("Brownfield redevelopment & site reuse", "Cleanup -> mixed-use / industrial reuse", 4, 5, 5, 5, 4, 4),
    ("Renewable energy, storage & microgrids", "Solar+storage / microgrid for critical facilities", 4, 4, 4, 3, 2, 3),
    ("Advanced & domestic manufacturing / critical minerals", "Plant build / equipment (OBBBA-new NMTC)", 4, 5, 4, 3, 5, 2),
    ("Food systems & ag processing / cold storage", "Value-added processing / cold-storage hub", 3, 4, 4, 3, 4, 3),
    ("Workforce, education & childcare facilities", "Training center / charter / childcare build", 4, 4, 4, 4, 4, 3),
    ("Carbon, environmental & ecosystem-credit markets", "45Q + carbon/RIN/nutrient/PFAS credits", 4, 4, 5, 4, 4, 5),
]
first = hrow + 1
for i, (field, proj, *sc) in enumerate(fields):
    rr = first + i
    wf.cell(row=rr, column=1, value=i + 1).alignment = CENTER
    wf.cell(row=rr, column=2, value=field).alignment = WRAP
    wf.cell(row=rr, column=3, value=proj).alignment = WRAP
    for k, v in enumerate(sc):
        c = wf.cell(row=rr, column=4 + k, value=v); c.alignment = CENTER; c.border = BORDER; c.fill = INPUT_FILL
    comp = (f"=D{rr}*{wkey['MARKET']}+E{rr}*{wkey['STACK']}+F{rr}*{wkey['TAIL']}"
            f"+G{rr}*{wkey['FIT']}+H{rr}*{wkey['TAILWIND']}+I{rr}*{wkey['ENTRY']}")
    cc = wf.cell(row=rr, column=10, value=comp); cc.number_format = NUM2; cc.fill = TOTAL_FILL; cc.font = Font(bold=True)
last = first + len(fields) - 1
for i in range(len(fields)):
    rr = first + i
    rk = wf.cell(row=rr, column=11, value=f"=RANK(J{rr},$J${first}:$J${last},0)")
    rk.alignment = CENTER; rk.font = Font(bold=True, color="3B5E2B")
wf.column_dimensions["A"].width = 4
wf.column_dimensions["B"].width = 34
wf.column_dimensions["C"].width = 36
for col in "DEFGHI":
    wf.column_dimensions[col].width = 9
wf.column_dimensions["J"].width = 11
wf.column_dimensions["K"].width = 7
wf.freeze_panes = "B5"

# STACKS (detail) --------------------------------------------------------------
wstk = wb.create_sheet("Stacks")
title(wstk, "Per field — the grant/credit stack, the steward tail, PL's fee route, the hook",
      "How PublicLogic grant-stacks and stewards each field. Bedrock vs sunsetting noted. Verified 2026-06-03.")
sh = ["#", "Field", "Grant / credit stack (layers to assemble)", "Compliance & stewardship tail", "PL fee route / the hook"]
hrow = 4
for j, h in enumerate(sh, start=1):
    wstk.cell(row=hrow, column=j, value=h)
style_header(wstk, hrow, len(sh))
stacks = [
    ("LIHTC (9%/4% — expanded by OBBBA) + tax-exempt bonds + HOME + OZ (permanent) + HTC + state DHCD/EOHLC",
     "15-yr LIHTC compliance + extended-use (to 30+ yrs) — long, recurring tail",
     "Compliance soft cost in the deal; 15-yr asset-management/steward retainer"),
    ("NMTC (permanent) + HRSA + USDA Community Facilities + Medicaid/CCBHC (entitlement) + state behavioral-health",
     "7-yr NMTC + ongoing Medicaid/clinical compliance — Rothschild's CFIR lane",
     "NMTC soft cost + behavioral-systems continuity retainer (Rothschild)"),
    ("CWSRF/DWSRF (funded FY26) + WIFIA (active, ~$7B) + EPA Brownfields + IIJA lead/PFAS set-asides",
     "Loan covenants + Davis-Bacon + lead/PFAS reporting over 20-40 yr asset life",
     "Soft cost (planning/design) + ongoing SRF/WIFIA compliance retainer"),
    ("BEAD (in-flux) + USDA ReConnect + Community Compact (no match) + state middle-mile + private ISP capital",
     "Buildout-milestone + BEAD reporting; municipal governance tail",
     "Grant admin + governance (LogicOS/CivicPulse) + retainer — PROVEN at Shrewsbury"),
    ("EPA Brownfields (assessment/cleanup/RLF) + CWSRF + OZ + state + Brownfield TIF/DIF + NMTC on the reuse",
     "Institutional controls + environmental covenants — effectively permanent",
     "Cleanup soft cost + long-horizon environmental-controls stewardship"),
    ("48E/45Y ITC/PTC (SUNSETTING — begin constr. before 7/4/2026) + 45Q ($85) + USDA REAP + direct pay (6417)",
     "Prevailing-wage, domestic-content, FEOC, interconnection — front-loaded; 45Q 12-yr",
     "Soft cost + compliance; URGENT timing window — anchor on 45Q (bedrock), treat ITC as time-boxed"),
    ("NMTC (newly eligible under OBBBA) + 48C + EDA Public Works + state incentives + USDA B&I",
     "Job-creation covenants + supply-chain + 7-yr NMTC — strong recurring tail",
     "NMTC soft cost + covenant-monitoring retainer; OBBBA tailwind is strongest here"),
    ("USDA RD (Water/CF/B&I) + NMTC supply-chain category + REAP + state ag + private offtake",
     "Food-safety + job-creation + USDA loan compliance",
     "Grant stacking + compliance retainer; rural/food-desert angle"),
    ("NMTC + WIOA + state education/childcare + OZ + philanthropic/foundation capital",
     "Outcomes reporting + 7-yr NMTC + program compliance",
     "Soft cost + outcomes-governance retainer; pairs with workforce for energy/mfg sites"),
    ("45Q ($85, strengthened) + voluntary carbon/RIN + nutrient/PFAS-credit markets + brownfield + OZ",
     "Registry verification / MRV (Puro.earth etc.) — ongoing, the core circular-economy tail",
     "The Michigan/AED core extended: governance + MRV compliance retainer (highest tail)"),
]
first = hrow + 1
for i, (stack, tail, route) in enumerate(stacks):
    rr = first + i
    wstk.cell(row=rr, column=1, value=i + 1).alignment = CENTER
    wstk.cell(row=rr, column=2, value=fields[i][0]).alignment = WRAP
    wstk.cell(row=rr, column=3, value=stack).alignment = WRAP
    wstk.cell(row=rr, column=4, value=tail).alignment = WRAP
    wstk.cell(row=rr, column=5, value=route).alignment = WRAP
    for j in range(1, 6):
        wstk.cell(row=rr, column=j).border = BORDER
for col, w in zip("ABCDE", [4, 26, 52, 40, 44]):
    wstk.column_dimensions[col].width = w
wstk.freeze_panes = "B5"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Development_Fields.xlsx")
wb.save(out)
print("wrote", out)
