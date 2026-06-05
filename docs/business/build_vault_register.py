#!/usr/bin/env python3
"""
Builds MichiganLTC_VAULT_Register.xlsx — the live VAULT governance action register for the
Michigan LTC Network across the 5 pillars (Verification, Authority, Utility, Legitimacy,
Transfer). Each finding has severity, status, owner, and remediation. A live "Governance
Readiness %" rolls up by severity-weighted completion and climbs as items close.
Update the yellow Status cells (Gap / In progress / Closed). Zero formula errors.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BOLD=Font(bold=True); HDR=Font(bold=True,color="FFFFFF")
HFILL=PatternFill("solid",fgColor="1F2D5F")
INF=PatternFill("solid",fgColor="FFF2CC"); TOT=PatternFill("solid",fgColor="E2EFDA")
CRIT=PatternFill("solid",fgColor="F4C7C3"); HIGH=PatternFill("solid",fgColor="FCE4D6"); MED=PatternFill("solid",fgColor="FFF2CC")
NOTE=Font(italic=True,size=9,color="555555")
THIN=Side(style="thin",color="BBBBBB"); B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
CTR=Alignment(horizontal="center",vertical="center"); WRAP=Alignment(wrap_text=True,vertical="top")

wb=openpyxl.Workbook(); ws=wb.active; ws.title="VAULT Register"
ws["A1"]="Michigan LTC Network — VAULT Governance Register (live)"; ws["A1"].font=Font(bold=True,size=14,color="1F2D5F")
ws["A2"]=("V/A/U/L/T findings. Update yellow Status (Gap / In progress / Closed); Governance Readiness recalcs. "
          "Date 2026-06-03. Not legal/financial advice."); ws["A2"].font=NOTE

cols=["ID","Pillar","Finding","Severity","Status","Owner","Remediation / evidence to close"]
hr=4
for i,c in enumerate(cols):
    x=ws.cell(row=hr,column=i+1,value=c); x.font=HDR; x.fill=HFILL; x.border=B; x.alignment=CTR
# hidden helper headers
ws.cell(row=hr,column=8,value="wt").font=NOTE; ws.cell(row=hr,column=9,value="cmpl").font=NOTE

F=[
 ("V1","V","Headline EBITDA depends on ~$20.4M unsecured credits + ~$4.8M understated opex","Critical","In progress","PublicLogic/Glen/WasteWerx","Anchor deck on base; register credits; confirm opex actuals"),
 ("V2","V","Single source of truth (retire 3-version proforma / EM&S numbers)","Critical","In progress","PublicLogic","FINAL model is SSOT; reconcile all docs to it"),
 ("V3","V","Offtake (fuel/SAF/carbon black) unverified","High","Gap","Robert/Glen","Execute offtake: buyers, volumes, price, term"),
 ("V4","V","Fuel build-up gap ($6.75M raw oil -> $15.15M upgraded)","High","Gap","Glen/WasteWerx","Document distillation/SAF basis + offtake"),
 ("V5","V","Labor is worst-case; MI rates/cross-training/benefits pending","Medium","In progress","Robert/WasteWerx","Set xtrain + mi_adj; provide benefits actuals"),
 ("V6","V","Carbon black excluded from signed model (real product)","Medium","Gap","Robert","Confirm price + offtake; include in base"),
 ("V7","V","Use-of-funds ~$1M over the $15M/site raise","Medium","Gap","Robert","Reconcile line items to the raise"),
 ("A1","A","Parent holdco + SPVs not papered (operating agreements/resolutions)","Critical","Gap","Counsel/Robert","Form + paper entities; board resolutions"),
 ("A2","A","LARA / foreign-entity registration unconfirmed (gates EGLE filings)","Critical","Gap","Robert","MiBusiness registry lookup; register"),
 ("A3","A","Site SPVs (E.M.&S.-FLINT) formation/renaming","High","Gap","Counsel/Robert","Form/rename under Michigan LTC"),
 ("A4","A","Decision authority undefined (signing, investor terms)","High","Gap","Robert","Authority matrix; delineate McCall multi-org roles"),
 ("A5","A","Securities offering: PL not placement agent; fee fixed/non-contingent","Critical","Gap","Counsel/PublicLogic","Engage securities counsel; structure raise; align PL engagement letter"),
 ("U1","U","Org chart / cross-trained staffing plan","High","Gap","Robert/WasteWerx","Expected headcount + roles (vs worst-case)"),
 ("U2","U","Site control (host co-location agreements) execution","High","Gap","Robert","Execute site control per active site"),
 ("U3","U","No independent energy/emissions-modeling lead","High","Gap","Robert/Glen","Retain engineer (AVERT/GHG modeling)"),
 ("U4","U","Compliance calendar / permit register / MRV per site","Medium","Gap","PublicLogic","Populate VAULT spine per site"),
 ("U5","U","Data room not stood up","Medium","In progress","PublicLogic","Index data room: SSOT + contracts + entity docs"),
 ("L1","L","Securities counsel review before any circulation","Critical","Gap","Counsel","Engagement + offering structure/opinion"),
 ("L2","L","WasteWerx IP: license disclosure; no proprietary images","High","Gap","PublicLogic/WasteWerx","License/title/termination disclosure; generic visuals"),
 ("L3","L","FOIA wall (private investor deck vs public EGLE grant)","High","In progress","PublicLogic","Separate doc sets; no proprietary in grant"),
 ("L4","L","Overclaim: 'largest in US' vs verbal 'world'","Medium","Gap","PublicLogic/Robert","Use defensible US claim unless evidenced"),
 ("L5","L","Credit/tax legitimacy (RIN registration, 45Z eligibility, counsel)","High","Gap","Glen/Counsel","Register RIN; tax-counsel review 45Z"),
 ("L6","L","Community-benefit %: 10% vs 20% conflict; encode in docs","Medium","Gap","Robert/Counsel","Reconcile + encode in operating agreements"),
 ("L7","L","Per-site permit legitimacy (air PTI/NSR, Part 169/201, NPDES)","Medium","Gap","Robert/PublicLogic","EGLE pre-app follow-through; permit register"),
 ("T1","T","Succession risk: all knowledge in McCall","Critical","Gap","PublicLogic/Rothschild","Institutional-continuity protocol; named stewards"),
 ("T2","T","Replication architecture (state SPVs) template","High","Gap","PublicLogic","Document module library / replication template"),
 ("T3","T","Institutional continuity (evidence chains, steward designations)","Medium","In progress","PublicLogic/Rothschild","CFIR continuity; VAULT records"),
 ("T4","T","Community/client ownership & durability encoded","Medium","Gap","Robert/Counsel","Encode durability in governing docs"),
]
r0=hr+1
sevfill={"Critical":CRIT,"High":HIGH,"Medium":MED}
for i,(fid,pil,find,sev,stat,own,rem) in enumerate(F):
    r=r0+i
    ws.cell(row=r,column=1,value=fid).alignment=CTR
    ws.cell(row=r,column=2,value=pil).alignment=CTR
    ws.cell(row=r,column=3,value=find).alignment=WRAP
    sc=ws.cell(row=r,column=4,value=sev); sc.alignment=CTR; sc.fill=sevfill[sev]
    st=ws.cell(row=r,column=5,value=stat); st.alignment=CTR; st.fill=INF
    ws.cell(row=r,column=6,value=own).alignment=WRAP
    ws.cell(row=r,column=7,value=rem).alignment=WRAP
    ws.cell(row=r,column=8,value=f'=IF(D{r}="Critical",3,IF(D{r}="High",2,1))').number_format='0'
    ws.cell(row=r,column=9,value=f'=IF(E{r}="Closed",1,IF(E{r}="In progress",0.5,0))').number_format='0.0'
    for cc in range(1,8): ws.cell(row=r,column=cc).border=B
last=r0+len(F)-1
# readiness rollup
gr=last+2
ws.cell(row=gr,column=2,value="GOVERNANCE READINESS (severity-weighted)").font=BOLD
rd=ws.cell(row=gr,column=4,value=f"=SUMPRODUCT(H{r0}:H{last},I{r0}:I{last})/SUM(H{r0}:H{last})"); rd.number_format='0%'; rd.fill=TOT; rd.font=Font(bold=True,size=12)
ws.cell(row=gr+1,column=2,value="Critical findings open (status <> Closed)").font=BOLD
ws.cell(row=gr+1,column=4,value=f'=COUNTIFS(D{r0}:D{last},"Critical",E{r0}:E{last},"<>Closed")').alignment=CTR
ws.cell(row=gr+2,column=2,value="Total findings / Closed").font=BOLD
ws.cell(row=gr+2,column=4,value=f'=COUNTA(A{r0}:A{last})&" / "&COUNTIF(E{r0}:E{last},"Closed")').alignment=CTR

notes=[
 "","HOW TO USE / READ",
 "  * Update yellow Status: Gap (0%) -> In progress (50%) -> Closed (100%). Readiness = severity-weighted completion.",
 "  * Severity weights: Critical=3, High=2, Medium=1. Drive Critical findings to Closed first.",
 "  * 100% readiness = every VAULT gap closed = investment-grade & replication-ready. Pair with the All-A scorecard.",
 "  * Narrative & pillar logic: Michigan_LTC_VAULT_Assessment.md.",
]
nr=gr+4
for ln in notes:
    c=ws.cell(row=nr,column=2,value=ln); c.font=BOLD if ln.strip()=="HOW TO USE / READ" else NOTE; nr+=1
for col,w in zip("ABCDEFG",[6,7,46,10,12,24,46]): ws.column_dimensions[col].width=w
ws.column_dimensions["H"].width=5; ws.column_dimensions["I"].width=6
ws.freeze_panes="A5"

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"MichiganLTC_VAULT_Register.xlsx")
wb.save(out); print("wrote",out)
