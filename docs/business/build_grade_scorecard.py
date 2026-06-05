#!/usr/bin/env python3
"""
Builds Michigan_LTC_Investment_Grade_Scorecard.xlsx — a live scorecard that grades the
Michigan LTC investor case across 10 weighted dimensions and computes an overall GPA +
letter. Update the yellow "Current points" (0-4) as gating items close; the GPA and letter
recalculate. Target = 4.0 (All A). Live formulas, zero errors.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BOLD=Font(bold=True); HDR=Font(bold=True,color="FFFFFF")
HFILL=PatternFill("solid",fgColor="1F4E5F"); IN=PatternFill("solid",fgColor="FFF2CC")
TOT=PatternFill("solid",fgColor="E2EFDA"); NOTE=Font(italic=True,size=9,color="555555")
THIN=Side(style="thin",color="BBBBBB"); B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
CTR=Alignment(horizontal="center",vertical="center"); WRAP=Alignment(wrap_text=True,vertical="top")

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Scorecard"
ws["A1"]="Michigan LTC Network — Investment-Grade Scorecard (live)"; ws["A1"].font=Font(bold=True,size=14,color="1F4E5F")
ws["A2"]=("Update yellow 'Current pts' (0-4) as gating items close; GPA + letter recalc. Target = 4.0 (All A). "
          "Date 2026-06-03. Not financial/legal advice."); ws["A2"].font=NOTE

hdr=["#","Dimension","Weight","Current pts (0-4)","Target","Weighted (cur)","Gap to A","What earns the A (evidence)"]
hr=4
for i,h in enumerate(hdr):
    c=ws.cell(row=hr,column=i+1,value=h); c.font=HDR; c.fill=HFILL; c.border=B; c.alignment=CTR

# (dim, weight, current_pts, evidence)
rows=[
 ("Thesis / positioning",0.06,3.7,"One consistent narrative + approved one-pager"),
 ("Feedstock control",0.10,3.3,"Executed supply agreements (volume/term/price) per site"),
 ("Sites",0.08,2.7,"Locked list + signed site control + use-compatibility memo"),
 ("Team / operators",0.10,2.7,"Named leads + engagement letters (incl. emissions engineer, securities counsel, CFO-level)"),
 ("Headline economics",0.18,1.3,"Reconciled model: bottom-up opex confirmed, base stands w/o credits, credits registration-backed, independent review"),
 ("Product offtake",0.12,1.7,"Executed offtake (RD/SAF/carbon black): buyers, volumes, pricing, term"),
 ("Deal / corporate structure",0.10,2.0,"Holdco+SPVs formed, operating agreements, LARA, counsel-structured offering"),
 ("Capital ask realism",0.08,2.0,"Ask ties to model; use-of-funds reconciled; company-raise split documented"),
 ("Diligence readiness / data room",0.08,1.7,"Single source of truth + indexed data room; all numbers reconciled"),
 ("Compliance posture",0.10,2.3,"Securities counsel review; IP/FOIA/disclosure language set; PL fixed/non-contingent"),
]
r0=hr+1
for i,(dim,w,cur,ev) in enumerate(rows):
    r=r0+i
    ws.cell(row=r,column=1,value=i+1).alignment=CTR
    ws.cell(row=r,column=2,value=dim).alignment=WRAP
    ws.cell(row=r,column=3,value=w).number_format='0.00'
    cp=ws.cell(row=r,column=4,value=cur); cp.number_format='0.0'; cp.fill=IN; cp.alignment=CTR; cp.border=B
    ws.cell(row=r,column=5,value=4.0).number_format='0.0'; ws.cell(row=r,column=5).alignment=CTR
    ws.cell(row=r,column=6,value=f"=C{r}*D{r}").number_format='0.000'
    ws.cell(row=r,column=7,value=f"=E{r}-D{r}").number_format='0.0'
    ws.cell(row=r,column=8,value=ev).alignment=WRAP
    for cc in range(1,9): ws.cell(row=r,column=cc).border=B
last=r0+len(rows)-1
# GPA
gr=last+1
ws.cell(row=gr,column=2,value="OVERALL (weighted GPA)").font=BOLD
ws.cell(row=gr,column=3,value=f"=SUM(C{r0}:C{last})").number_format='0.00'
gpa=ws.cell(row=gr,column=6,value=f"=SUMPRODUCT(C{r0}:C{last},D{r0}:D{last})"); gpa.number_format='0.00'; gpa.fill=TOT; gpa.font=BOLD
# letter from GPA
gcell=f"F{gr}"
letter=(f'=IF({gcell}>=3.85,"A",IF({gcell}>=3.5,"A-",IF({gcell}>=3.15,"B+",IF({gcell}>=2.85,"B",'
        f'IF({gcell}>=2.5,"B-",IF({gcell}>=2.15,"C+",IF({gcell}>=1.85,"C",IF({gcell}>=1.5,"C-","D"))))))))')
ws.cell(row=gr,column=2,value="OVERALL (weighted GPA / letter)").font=BOLD
lc=ws.cell(row=gr,column=7,value=letter); lc.fill=TOT; lc.font=Font(bold=True,size=12,color="1F4E5F"); lc.alignment=CTR
ws.cell(row=gr+1,column=2,value="Target").font=BOLD
ws.cell(row=gr+1,column=6,value=4.0).number_format='0.00'
ws.cell(row=gr+1,column=7,value="A").font=Font(bold=True,color="1F4E5F"); ws.cell(row=gr+1,column=7).alignment=CTR
ws.cell(row=gr+2,column=2,value="Gap to All A (4.0)").font=BOLD
ws.cell(row=gr+2,column=6,value=f"=4-F{gr}").number_format='0.00'

notes=[
 "","HOW TO USE",
 "  * As each dimension's evidence lands (signed offtake, confirmed opex, counsel sign-off, etc.), raise its 'Current pts'.",
 "  * 4.0 in every row = All A. The biggest movers are Headline economics (0.18) and Offtake (0.12) — they gate the grade.",
 "  * Grade points: A=4.0, A-=3.7, B+=3.3, B=3.0, B-=2.7, C+=2.3, C=2.0, C-=1.7, D+=1.3, D=1.0.",
 "  * Detail and owners: see Michigan_LTC_Path_to_All_A.md.",
]
nr=gr+4
for ln in notes:
    c=ws.cell(row=nr,column=2,value=ln); c.font=BOLD if ln.strip()=="HOW TO USE" else NOTE; nr+=1
for col,w in zip("ABCDEFGH",[4,30,8,14,8,13,9,52]): ws.column_dimensions[col].width=w
ws.freeze_panes="B5"

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"Michigan_LTC_Investment_Grade_Scorecard.xlsx")
wb.save(out); print("wrote",out)
