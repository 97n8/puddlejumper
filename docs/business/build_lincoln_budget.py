#!/usr/bin/env python3
"""
Builds Challenge4_Lincoln_Budget.xlsx in the EGLE MI Healthy Climate Challenge budget
format (EQP1167): Personnel + Fringe + Contractual, with live subtotals and a grand total.
Study-only (no equipment/construction). Total = $399,000 (under the <$400k guidance).
Community subaward kept <=5% of award. Optional scope tier marked for scalability.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
SUB_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE = Font(italic=True, size=9, color="555555")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'
RIGHT = Alignment(horizontal="right")
WRAP = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Budget Template"

ws["A1"] = ("Michigan Department of Environment, Great Lakes, and Energy — Office of Climate and Energy\n"
            "MI Healthy Climate Challenge #4 — Applicant Budget Template (EQP1167 format)")
ws["A1"].font = BOLD; ws["A1"].alignment = WRAP
ws["A3"] = "Applicant Name:"; ws["B3"] = "Energy Mann & Sunn, Inc. (DBA Michigan LTC)"
ws["A4"] = "Project Name:"; ws["B4"] = "Industrial Asset Transition Study — National Energy of Lincoln Biomass Facility"
for r in (3, 4):
    ws.cell(row=r, column=1).font = BOLD

def hdr(row, cols):
    for i, c in enumerate(cols):
        cell = ws.cell(row=row, column=i + 1, value=c)
        cell.font = HDR; cell.fill = HDR_FILL; cell.border = BORDER

# ---- Personnel ----
r = 6
hdr(r, ["Personnel Staff - Name and Title", "Hours", "Hourly Rate", "Requested Amount"])
personnel = [
    ("Project Manager (EM&S / National Salvage) — study lead & coordination", 600, 60),
    ("Data / Energy Analyst — facility data, baseline support", 360, 50),
]
pr0 = r + 1
for i, (name, hrs, rate) in enumerate(personnel):
    rr = pr0 + i
    ws.cell(row=rr, column=1, value=name).alignment = WRAP
    ws.cell(row=rr, column=2, value=hrs).alignment = RIGHT
    ws.cell(row=rr, column=3, value=rate).number_format = MONEY
    ws.cell(row=rr, column=4, value=f"=B{rr}*C{rr}").number_format = MONEY
    for c in range(1, 5):
        ws.cell(row=rr, column=c).border = BORDER
pr_last = pr0 + len(personnel) - 1
psub = pr_last + 1
ws.cell(row=psub, column=1, value="Personnel Subtotal").font = BOLD
pc = ws.cell(row=psub, column=4, value=f"=SUM(D{pr0}:D{pr_last})")
pc.number_format = MONEY; pc.font = BOLD; pc.fill = SUB_FILL

# ---- Fringe ----
r = psub + 2
ws.cell(row=r, column=1, value="Fringe Benefits").font = HDR; ws.cell(row=r, column=1).fill = HDR_FILL
ws.cell(row=r, column=3, value="Rate").font = HDR; ws.cell(row=r, column=3).fill = HDR_FILL
ws.cell(row=r, column=4, value="Requested Amount").font = HDR; ws.cell(row=r, column=4).fill = HDR_FILL
fr = r + 1
ws.cell(row=fr, column=1, value="Fringe (contractor/loaded-rate basis)")
ws.cell(row=fr, column=3, value=0).number_format = '0%'
ws.cell(row=fr, column=4, value=f"=D{psub}*C{fr}").number_format = MONEY
fsub = fr + 1
ws.cell(row=fsub, column=1, value="Personnel Costs Subtotal (Personnel + Fringe)").font = BOLD
fc = ws.cell(row=fsub, column=4, value=f"=D{psub}+D{fr}")
fc.number_format = MONEY; fc.font = BOLD; fc.fill = SUB_FILL

# ---- Contractual ----
r = fsub + 2
hdr(r, ["Contractual Services", "Hours or Units", "Cost", "Requested Amount"])
contractual = [
    ("WasteWerx LLC — process & thermal engineering, system design feasibility", "fixed scope", 95000),
    ("Independent energy & emissions modeling — GHG baseline, AVERT/EPA factors, enabling-infrastructure", "fixed scope", 70000),
    ("PublicLogic LLC — project governance, GHG accounting framework, compliance, deliverables (fixed, non-contingent)", "fixed scope", 90000),
    ("Acadia Capital Management — financial, scenario & investment-readiness analysis", "fixed scope", 40000),
    ("Community & workforce subaward (Rothschild/PublicLogic + community partner) — <=5% of award", "subaward", 19000),
    ("National Energy of Lincoln — facility engineering & data support (host staff time)", "fixed scope", 25000),
    ("Knowledge-sharing, final report/deck/public summary, EGLE convening", "fixed scope", 6000),
]
cr0 = r + 1
for i, (name, units, cost) in enumerate(contractual):
    rr = cr0 + i
    ws.cell(row=rr, column=1, value=name).alignment = WRAP
    ws.cell(row=rr, column=2, value=units)
    ws.cell(row=rr, column=4, value=cost).number_format = MONEY
    for c in range(1, 5):
        ws.cell(row=rr, column=c).border = BORDER
cr_last = cr0 + len(contractual) - 1
csub = cr_last + 1
ws.cell(row=csub, column=1, value="Contractual Subtotal").font = BOLD
cc = ws.cell(row=csub, column=4, value=f"=SUM(D{cr0}:D{cr_last})")
cc.number_format = MONEY; cc.font = BOLD; cc.fill = SUB_FILL

# ---- Grand total ----
gt = csub + 2
ws.cell(row=gt, column=1, value="TOTAL GRANT REQUEST").font = BOLD
g = ws.cell(row=gt, column=4, value=f"=D{fsub}+D{csub}")
g.number_format = MONEY; g.font = Font(bold=True, size=12); g.fill = SUB_FILL
# community % check
cr = gt + 1
ws.cell(row=cr, column=1, value="Community subaward as % of total (must be <=5%)").font = NOTE
pc2 = ws.cell(row=cr, column=4, value=f"=D{cr0+4}/D{gt}")
pc2.number_format = '0.0%'

notes = [
    "",
    "NOTES",
    "  * Study/planning costs only — no equipment, installation, or construction (per program rules).",
    "  * Total $399,000 sized under the '<$400,000 encouraged' guidance; scope is scalable if a reduced award is offered.",
    "  * OPTIONAL SCOPE TIER (if full award): expand scenario modeling + add second-facility (McBain) comparative",
    "    screen, +$0 here (kept under $400k) — describe scalability in the budget narrative.",
    "  * PublicLogic fee is FIXED and NON-CONTINGENT on award (2 CFR 200.459); no grant-stacking/contingent fee in this study.",
    "  * Confirm leveraged/match (in-kind facility staff time; partner cost-share) in the narrative — strengthens Facility Commitment.",
    "  * Not legal/accounting advice. Figures illustrative pending final partner quotes.",
]
nr = cr + 2
for ln in notes:
    c = ws.cell(row=nr, column=1, value=ln)
    c.font = BOLD if ln.strip() == "NOTES" else NOTE
    nr += 1

ws.column_dimensions["A"].width = 70
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 16

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Challenge4_Lincoln_Budget.xlsx")
wb.save(out)
print("wrote", out)
