#!/usr/bin/env python3
"""
PublicLogic — Capabilities Workbook (FINAL, internal strategy for Nathan & Allie).
Grounded, humble, honest. Reveals the company that already exists (Hubbardston, Sutton,
Phillipston, Gardiner, Michigan LTC). No hype. Answers the 10 strategy questions.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

SLATE="1A1D20"; GREEN="0F4C3A"; GOLD="A9772F"; MUTEC="6B6660"
HDR=Font(bold=True,color="FFFFFF",size=11); HFILL=PatternFill("solid",fgColor=SLATE)
GHFILL=PatternFill("solid",fgColor=GREEN); TITLE=Font(bold=True,size=15,color=SLATE)
SUB=Font(italic=True,size=9,color=MUTEC); SEC=Font(bold=True,size=12,color=GREEN)
GLD=Font(bold=True,size=11,color=GOLD); BOLD=Font(bold=True); BODY=Font(size=11,color=SLATE)
CFILL=PatternFill("solid",fgColor="ECEAE3"); GFILL=PatternFill("solid",fgColor="E3EAE5"); IFILL=PatternFill("solid",fgColor="FFF2CC")
THIN=Side(style="thin",color="CCCCCC"); B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP=Alignment(wrap_text=True,vertical="top"); MID=Alignment(wrap_text=True,vertical="center")

wb=openpyxl.Workbook()
def newtab(name,title,sub=None):
    ws=wb.create_sheet(name); ws["A1"]=title; ws["A1"].font=TITLE
    if sub: ws["A2"]=sub; ws["A2"].font=SUB
    return ws
def blocks(ws,start,rows,h=34,w=(20,104)):
    r=start
    for k,v in rows:
        a=ws.cell(row=r,column=1,value=k); a.font=HDR; a.fill=GHFILL; a.border=B; a.alignment=MID
        c=ws.cell(row=r,column=2,value=v); c.alignment=WRAP; c.border=B; c.fill=CFILL; c.font=BODY
        ws.row_dimensions[r].height=h; r+=1
    ws.column_dimensions["A"].width=w[0]; ws.column_dimensions["B"].width=w[1]
    return r
def lines(ws,start,items,col=1,width=110):
    r=start
    for it in items:
        c=ws.cell(row=r,column=col,value=it); c.alignment=WRAP; c.font=BODY; r+=1
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width=width
    return r
def htable(ws,top,headers,rows,widths,fill=HFILL,editable=()):
    for i,hh in enumerate(headers):
        c=ws.cell(row=top,column=i+1,value=hh); c.font=HDR; c.fill=fill; c.border=B; c.alignment=MID
    for ri,row in enumerate(rows):
        for ci,val in enumerate(row):
            c=ws.cell(row=top+1+ri,column=ci+1,value=val); c.border=B; c.alignment=WRAP; c.font=BODY
            if ci in editable: c.fill=IFILL
    for i,w in enumerate(widths): ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width=w
    ws.freeze_panes=f"A{top+1}"

# ===== START HERE =====
ws=wb.active; ws.title="Start Here"
ws["A1"]="PublicLogic — Internal Strategy & Business Development"; ws["A1"].font=Font(bold=True,size=16,color=SLATE)
ws["A2"]="For Nathan & Allie. What we are, what we sell, who we serve, how we grow. One line each here; the rest is in the tabs."; ws["A2"].font=SUB
blocks(ws,4,[
 ("North Star","Every system should make it easier for the next person to do the right thing."),
 ("The bridge","Implementation is the service. Institutional Stewardship is the method."),
 ("The equation","Good Work + Stewardship = Work That Lasts."),
 ("The ladder","LogicCommons helps people start. PublicLogic helps them carry it through."),
 ("The architecture","LogicCommons = public access layer · Permit & Bridge = navigation layer · PublicLogic = services · Continuity System (PuddleJumper) = delivery infrastructure."),
 ("1 · Why we exist","Good work keeps failing when knowledge, ownership, or continuity disappears. We've watched it happen, and we built the practices to stop it."),
 ("2 · The problem","Capacity stops keeping pace with complexity. The 'Chair' — the place where a function lives — ends up depending on one person."),
 ("3 · What clients buy","Projects moved forward. Funding found and won. A role or function covered."),
 ("4 · What we deliver","The same — plus the stewardship that makes it survive turnover."),
 ("5 · What makes us different","Most systems manage work. PublicLogic helps steward what has to survive the work."),
 ("6 · What we sell","Stewardship Map (the safe first step) → Sprints (Project · Funding · Permit & Bridge) → Implementation & Capacity Support."),
 ("7 · Who first","Our existing municipal network, regional planning agencies, engineering firms, municipal consulting firms."),
 ("8 · How it compounds","Every engagement leaves reusable knowledge, templates, and proof. We get stronger; clients get less dependent."),
 ("9 · Nathan + Allie","Institutional systems + human systems = the intersection. The work fails when either side is ignored."),
 ("10 · Next 12 months","A handful of warm wins, the Map as the first sale, Permit & Bridge as the wedge. Modest and real."),
],h=30)

# ===== WHY WE EXIST =====
ws=newtab("Why We Exist","Why PublicLogic Exists","We didn't start from a theory. We started from the work.")
r=lines(ws,4,[
 "PublicLogic came out of real work — Hubbardston, Sutton, Phillipston, Gardiner, Michigan LTC — and out of years inside municipal administration, capital planning, and grant development.",
 "",
 "Across all of it we kept seeing the same thing: good, important work would stall or quietly disappear. Not because the idea was wrong. Because the person holding it left, the knowledge was never written down, or no one owned what came next.",
 "",
 "PublicLogic is our attempt to fix that — with the practices we wish we'd had when we were inside it.",
 "",
])
ws.cell(row=r+1,column=1,value="North Star").font=SEC
ws.cell(row=r+2,column=1,value="“Every system should make it easier for the next person to do the right thing.”").font=Font(size=14,italic=True,color=GREEN)
ws.cell(row=r+2,column=1).alignment=WRAP
ws.cell(row=r+4,column=1,value="This isn't a new company we're inventing. It's the company that already exists — named, and made repeatable.").font=Font(size=11,color=SLATE,italic=True)

# ===== WHAT BREAKS WITHOUT STEWARDSHIP =====
ws=newtab("What Breaks","What Breaks Without Stewardship","Plain language. None of this is dramatic — which is exactly why it's dangerous.")
htable(ws,4,["When stewardship is missing…","…this is what we've watched happen"],
 [
 ["Turnover","The person leaves and the knowledge leaves with them."],
 ["Lost knowledge","How and why things were done stops being written down."],
 ["Delayed projects","Work stalls because no one owns the next step."],
 ["Missed funding","A funding window closes before anyone is ready."],
 ["Dependency on individuals","The whole function runs through one person's head."],
 ["Repeated mistakes","The same problems come back because nothing was captured."],
 ["Plans that never become action","A good plan gets made, then sits on a shelf."],
 ],[30,80])
ws.cell(row=13,column=1,value="It's quiet failure. By the time anyone notices, the momentum, the money, or the knowledge is already gone. Stewardship is how we keep that from happening.").font=Font(italic=True,size=11,color=GREEN)
ws.cell(row=13,column=1).alignment=WRAP; ws.merge_cells("A13:B14")

# ===== WHAT WE DO =====
ws=newtab("What We Do","What PublicLogic Actually Does",None)
r=blocks(ws,4,[
 ("In one sentence","PublicLogic helps projects move through public systems, and helps organizations keep the capacity required to sustain what they build."),
 ("The bridge","Implementation is the service. Institutional Stewardship is the method."),
 ("What makes us different","Most systems manage work. PublicLogic helps steward what has to survive the work."),
 ("What we are NOT","Not a grant-writing firm. Not an AI consulting firm. Not a software company. Not a planning firm. Not a generic implementation consultant. The work touches all of those, but none of them is the point."),
 ("Clients buy / we deliver","Clients buy funding, progress, and capacity. We deliver those — plus the stewardship that makes them last. That's not a contradiction; it's the business model."),
],h=40)
# equation
ws.cell(row=r+1,column=1,value="THE EQUATION").font=SEC
ws.cell(row=r+1,column=2,value="Recommended (use this one):  Good Work + Stewardship = Work That Lasts").font=GLD
ws.cell(row=r+2,column=2,value="Fuller / operational:  Project + Funding + Capacity + Continuity = Implementation").font=Font(size=11,color=SLATE)
ws.cell(row=r+3,column=2,value="Fuller / outcomes:  Vision + Funding + Capacity + Stewardship = Durable Results").font=Font(size=11,color=SLATE)
# method
ws.cell(row=r+5,column=1,value="HOW WE WORK").font=SEC
ws.cell(row=r+5,column=2,value="Map  →  Embed  →  Encode  →  Sustain").font=GLD
method=[("Map","Surface how work, knowledge, authority, records, and risk actually move."),
 ("Embed","Put the right structure into the live workflow, with the people who do the work."),
 ("Encode","Capture the knowledge and rules so they outlast any individual."),
 ("Sustain","Monitor adoption and transfer ownership so it runs without us.")]
for i,(st,d) in enumerate(method):
    rr=r+6+i; a=ws.cell(row=rr,column=1,value=st); a.font=Font(bold=True,color=GOLD); a.fill=GFILL; a.border=B
    c=ws.cell(row=rr,column=2,value=d); c.font=BODY; c.alignment=WRAP; c.border=B
ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=100

# ===== OFFER STACK =====
ws=newtab("Offer Stack","The Offer Stack","Simple, honestly priced. The Map comes first; everything else follows from what it finds. Fixed-fee or retainer, non-contingent.")
htable(ws,4,["Product","Why it exists","Output","Price","Transfer asset"],
 [
 ["Stewardship Map  (entry)","Understand how work, knowledge, authority, records, and risk actually move — before anyone commits money. The safest first engagement.","Stewardship Map · Continuity Risks · Readiness Findings · Priority Roadmap","$2,500 – $7,500","CaseSpace + the Map"],
 ["Project Development Sprint","Turn a priority or a challenge into an actionable project.","Project Development Roadmap","$5,000 – $15,000","Charter / Scope template"],
 ["Funding Strategy Sprint","Find realistic funding and the readiness it requires.","Funding Roadmap","$5,000 – $12,500","Funding scan + roadmap workbook"],
 ["Permit & Bridge Sprint","Move a project through the public systems around it (permits, boards, funding, stakeholders).","Approvals path + coordination plan","$7,500 – $15,000","Bridge map / permit register"],
 ["Implementation Support","Coordinate and hold accountability while a plan gets executed by the client's team.","Implementation Framework + Accountability Structure","$3,500 – $8,500 / month","VAULT continuity record"],
 ["Capacity Support","Step in and hold a role or function directly — interim / fractional — until it transfers back.","Embedded Capacity Plan + Continuity Handoff","$4,000 – $10,000 / month","Role / Coverage Map + handoff"],
 ],[28,42,32,18,26])
ws.cell(row=11,column=1,value="Implementation vs Capacity: Implementation helps YOUR team execute — the work stays theirs (e.g., Shrewsbury). Capacity means we hold the role ourselves until it transfers back (e.g., Swanzey). Same discipline; different level of who holds the Chair.").font=Font(italic=True,size=9,color=MUTEC)
ws.cell(row=11,column=1).alignment=WRAP; ws.merge_cells("A11:E12")
ws.cell(row=13,column=1,value="Entry below the Map: the Permit & Bridge tab adds a public-facing ladder — Tier 0 Public Permit Helper (free / very low cost) and Tier 1 Permit Path Scan ($250–$750) — that builds trust and feeds these paid offers.").font=Font(italic=True,size=9,color=MUTEC)
ws.cell(row=13,column=1).alignment=WRAP; ws.merge_cells("A13:E14")

# ===== PERMIT & BRIDGE =====
ws=newtab("Permit & Bridge","Permit & Bridge","This may be the most important thing we do.  Best line: Permit & Bridge helps the public understand the path and helps project sponsors move through it.")
r=blocks(ws,4,[
 ("What it is","PublicLogic helps project sponsors move through public systems — and helps the public understand the path before they get there."),
 ("That includes","Permits · boards · funding · grants · stakeholder processes · implementation planning · public-sector coordination."),
 ("Why it exists","Developers, nonprofits, municipalities, and project sponsors usually understand their project. They often don't understand the public systems around it — the permits, the boards, the funding rules, the politics. That gap is where good projects stall."),
 ("What we become","The bridge. We translate between the project and the public system, so the project keeps moving and nothing falls through the cracks between the people involved."),
 ("Why it matters to us","It is the clearest, most sellable version of 'helping projects move through public systems' — and it's exactly the work we've already done in Sutton, Shrewsbury, Phillipston, and Michigan LTC."),
 ("The promise","A project sponsor should leave with a clear path forward, a clear record of how they got there, and a clear understanding of who owns what next."),
],h=42)
# two sides
r+=1
ws.cell(row=r,column=1,value="TWO SIDES OF THE SAME BRIDGE").font=SEC; ws.merge_cells(f"A{r}:B{r}"); r+=1
htable(ws,r,["Layer","Audience","Offer"],
 [
 ["Public Help Layer","Residents · small owners · nonprofits","“Can I do X in my backyard / property / building?”  Cheap, simple, framework-guided."],
 ["Project Sponsor Layer","Developers · towns · nonprofits · businesses","White-glove permit path · stewardship map · grant / funding strategy · implementation support."],
 ],[22,34,62])
r+=4
# logiccommons box (explicit)
bx=ws.cell(row=r,column=1,value="LogicCommons — the public access layer"); bx.font=Font(bold=True,color="FFFFFF",size=11); bx.fill=GHFILL; bx.alignment=MID; bx.border=B
ws.merge_cells(f"A{r}:C{r}"); ws.row_dimensions[r].height=20; r+=1
lc=ws.cell(row=r,column=1,value="Free templates, checklists, and frameworks that help people understand the path before they need paid help.   LogicCommons helps people start. PublicLogic helps them carry it through.")
lc.font=Font(size=11,color=SLATE); lc.alignment=WRAP; lc.fill=GFILL; lc.border=B
ws.merge_cells(f"A{r}:C{r+1}"); ws.row_dimensions[r].height=28; r+=3
# public output layer
ws.cell(row=r,column=1,value="PERMIT & BRIDGE: PUBLIC OUTPUT LAYER").font=SEC; ws.merge_cells(f"A{r}:C{r}"); r+=1
ws.cell(row=r,column=1,value="The free/cheap tier is NOT AI consulting. It is a guided public framework that helps people understand the path. PublicLogic can support a simple public-facing assistance layer for common questions:").font=BODY
ws.cell(row=r,column=1).alignment=WRAP; ws.merge_cells(f"A{r}:C{r+1}"); ws.row_dimensions[r].height=30; r+=2
for q in ["“Can I put this in my backyard?”","“Do I need a permit?”","“Who do I ask first?”",
 "“What board or office handles this?”","“What documents will I likely need?”",
 "“Is this a zoning, building, health, conservation, or licensing issue?”"]:
    c=ws.cell(row=r,column=1,value="•  "+q); c.font=BODY; c.alignment=WRAP; ws.merge_cells(f"A{r}:C{r}"); r+=1
r+=1
ws.cell(row=r,column=1,value="This layer should be low-cost or free — its purpose is access, trust, and triage. It does NOT replace the municipality, inspector, planner, or permitting authority. It helps people ask better questions, gather the right information, and avoid wasting time.").font=Font(italic=True,size=11,color=GREEN)
ws.cell(row=r,column=1).alignment=WRAP; ws.merge_cells(f"A{r}:C{r+1}"); ws.row_dimensions[r].height=30; r+=3
# logiccommons public access layer
ws.cell(row=r,column=1,value="THE FULL LADDER, FROM FREE TO WHITE-GLOVE").font=SEC; ws.merge_cells(f"A{r}:C{r}"); r+=1
ws.cell(row=r,column=1,value="LogicCommons is the public access layer. It offers free templates, checklists, and simple frameworks that help residents, small organizations, and project sponsors understand the path before they need paid help.").font=BODY
ws.cell(row=r,column=1).alignment=WRAP; ws.merge_cells(f"A{r}:C{r+1}"); ws.row_dimensions[r].height=28; r+=2
htable(ws,r,["Layer","PublicLogic piece","Purpose"],
 [
 ["Free public tools","LogicCommons templates","Help people self-serve basic planning, permit, grant, and project questions."],
 ["Cheap triage","Permit Path Scan","“What path am I probably on?”"],
 ["Paid diagnostic","Stewardship Map","“What actually needs to happen, who owns it, and what breaks?”"],
 ["White-glove help","Permit & Bridge / Funding / Implementation","PublicLogic carries the project forward."],
 ],[22,40,56])
r+=6
ws.cell(row=r,column=1,value="LogicCommons helps people start. PublicLogic helps them carry it through.").font=Font(bold=True,size=12,color=GREEN)
ws.cell(row=r,column=1).alignment=WRAP; ws.merge_cells(f"A{r}:C{r}"); r+=2
# offer ladder
ws.cell(row=r,column=1,value="THE OFFER LADDER  —  free tier creates trust; paid tiers handle complexity").font=SEC; ws.merge_cells(f"A{r}:C{r}"); r+=1
htable(ws,r,["Tier","Product","Price"],
 [
 ["0","Public Permit Helper / “Can I Do This?”","Free or very low cost"],
 ["1","Permit Path Scan","$250 – $750"],
 ["2","Stewardship Map","$2,500 – $7,500"],
 ["3","Permit & Bridge Sprint","$7,500 – $15,000"],
 ["4","White-Glove Implementation","$3,500 – $8,500 / month"],
 ["5","Funding / Grant Build","$5,000 – $25,000 (scope-dependent)"],
 ],[8,46,34])
r+=8
ws.cell(row=r,column=1,value="Credit policy: a paid tier credits toward the next engagement, but credits don't stack — up to one tier's fee credits forward. Keeps the entry risk low without eroding margin on a follow-on sprint.").font=Font(italic=True,size=9,color=MUTEC)
ws.cell(row=r,column=1).alignment=WRAP; ws.merge_cells(f"A{r}:C{r+1}"); r+=2
ws.cell(row=r,column=1,value="Permit & Bridge helps the public understand the path and helps project sponsors move through it.  That's the full product.").font=Font(bold=True,size=12,color=GOLD)
ws.cell(row=r,column=1).alignment=WRAP; ws.merge_cells(f"A{r}:C{r+1}")

# ===== PROOF =====
ws=newtab("Proof","Proof — We Can Show It","Stewardship is provable, not just stated. Each tied to work we've actually done.")
htable(ws,4,["Proof","What it shows","How we show it","Example"],
 [
 ["Continuity","A function survives when the person who ran it leaves.","Successor test + structure mapped to authority and stop-rules.","Sutton — the 30-day successor test the town would otherwise fail."],
 ["Record","Every claim, number, and decision traces to a source.","A governed record: source file → workbook row → output.","Michigan LTC corridor record — line-for-line traceability."],
 ["Readiness","The organization can absorb and sustain the change.","Readiness findings in the Map, before any build.","Pre-build assessment that sequences what the org can carry."],
 ["Adoption","Staff actually use the system — not just receive it.","Adoption monitoring through implementation (Allie's work).","Shrewsbury — human-systems support so the build is used."],
 ["Outcome","The work moved forward and got funded.","Funded applications and delivered projects, documented.","Shrewsbury fiber, Sutton — funding secured, projects implemented."],
 ],[16,36,40,44])
ws.cell(row=10,column=1,value="Honesty note: confirm what we're comfortable naming before sharing externally — some of this is client-confidential.").font=SUB
ws.cell(row=10,column=1).alignment=WRAP; ws.merge_cells("A10:D10")

# ===== NATHAN + ALLIE =====
ws=newtab("Nathan + Allie","Why We're Better Together",None)
htable(ws,4,["","","" ],[],[40,40,40])
blocks(ws,4,[
 ("Nathan — institutional systems","How municipalities, money, grants, procurement, and projects actually work — from the inside."),
 ("Allie — human systems","How people, roles, adoption, leadership, and change actually work — and why systems get used or ignored."),
 ("PublicLogic — the intersection","Most firms have one side or the other. Projects fail for whichever side gets ignored. We've each watched work fail for the other reason — which is exactly why we work together this way."),
 ("In practice","Nathan builds the structure; Allie makes sure people will actually use it and that it survives turnover. Neither half is optional."),
],h=46)

# ===== TRANSFER =====
ws=newtab("Transfer","Transfer — Every Engagement Leaves Something Behind","The point isn't to be needed forever. It's to make the next person's job easier.")
htable(ws,4,["Every engagement creates…","Meaning"],
 [
 ["Reusable knowledge","What we learned about how this kind of organization works."],
 ["Reusable templates","Charters, roadmaps, registers, and workbooks we can adapt next time."],
 ["Reusable process","A repeatable way of doing the work — not a one-off."],
 ["Reusable proof","Evidence (continuity / record / readiness / adoption / outcome) we can show the next client."],
 ],[30,78])
ws.cell(row=9,column=1,value="Two tests for every engagement:").font=SEC
ws.cell(row=10,column=1,value="PublicLogic should be stronger after every engagement.").font=Font(bold=True,size=12,color=GOLD)
ws.cell(row=11,column=1,value="The client should be less dependent after every engagement.").font=Font(bold=True,size=12,color=GOLD)
ws.cell(row=12,column=1,value="If neither happens, we are doing consulting instead of stewardship.").font=Font(bold=True,italic=True,size=12,color=GREEN)
ws.cell(row=12,column=1).alignment=WRAP; ws.merge_cells("A12:B12")
ws.cell(row=14,column=1,value="THE CONTINUITY SYSTEM — our delivery infrastructure (PuddleJumper)").font=SEC; ws.merge_cells("A14:B14")
ws.cell(row=15,column=1,value="Every engagement is delivered through a Continuity System — the records, decisions, templates, and processes that keep the work alive after we step back. PuddleJumper is the name of that system. Inside it —  CaseSpaces: a project's governed record.   VAULT: how knowledge is preserved so it survives turnover.   Workbooks · Registers · Templates: the reusable artifacts the work runs on.  ('Continuity System' is what we say to a client; 'PuddleJumper' is what we call it.)").font=Font(size=11,color=SLATE)
ws.cell(row=15,column=1).alignment=WRAP; ws.merge_cells("A15:B16"); ws.row_dimensions[15].height=60
ws.cell(row=17,column=1,value="LogicCommons is separate — the public, free layer (templates for residents and sponsors), not part of a client's Continuity System.").font=Font(size=10,color=MUTEC)
ws.cell(row=17,column=1).alignment=WRAP; ws.merge_cells("A17:B17")
ws.cell(row=18,column=1,value="A client never has to care about CaseSpaces or VAULT. They care about one thing: “If my planner leaves, does this still work?” A Continuity System is how we answer yes.").font=Font(italic=True,size=11,color=GREEN)
ws.cell(row=18,column=1).alignment=WRAP; ws.merge_cells("A18:B19")

# ===== WHO WE PURSUE =====
ws=newtab("Who We Pursue","Who We Pursue — Forced Prioritization","Don't list twenty things. Start where there's already a project, a budget, and trust.")
htable(ws,4,["Tier","Pursue","Why"],
 [
 ["Tier 1 — now","Existing municipal network · Regional planning agencies · Engineering firms · Municipal consulting firms","Already have projects and budgets; warmest relationships; shortest sales cycle."],
 ["Tier 2 — next","Developers · Housing organizations · Community development organizations","Real need (especially Permit & Bridge), longer cycle; pursue once Tier 1 is producing."],
 ["Tier 3 — later","Everything else","Interesting, but not now. Don't chase."],
 ],[16,52,40])
# small ranked pipeline
ws.cell(row=9,column=1,value="TIER-1 PIPELINE  (illustrative Revenue/Probability — replace with real figures; Weighted = Revenue x Probability; sort to rank)").font=SEC
ws.merge_cells("A9:F9")
heads=["Pursue","Target","Relationship","Revenue ($)","Probability","Weighted ($)"]
for i,hh in enumerate(heads):
    c=ws.cell(row=10,column=i+1,value=hh); c.font=HDR; c.fill=HFILL; c.border=B; c.alignment=MID
pipe=[
 ["NOW","CMRPC (regional planning)","Developing",60000,0.45],
 ["NOW","MRPC (regional planning)","Developing",50000,0.40],
 ["NOW","Warm municipal relationships (Sutton/Shrewsbury/Phillipston)","Strong",35000,0.60],
 ["NOW","Existing network","Strongest",25000,0.65],
 ["NOW","Fuss & O'Neill (engineering)","Cold",40000,0.30],
 ["Qualify","MVPC / other RPCs","Cold",50000,0.25],
 ["Qualify","Community Paradigm (municipal consulting)","Cold",30000,0.25],
]
for i,row in enumerate(pipe):
    rr=11+i
    for ci,val in enumerate(row):
        c=ws.cell(row=rr,column=ci+1,value=val); c.border=B; c.alignment=WRAP; c.font=BODY
        if ci in (0,2,3,4): c.fill=IFILL
    ws.cell(row=rr,column=4).number_format='#,##0'; ws.cell(row=rr,column=5).number_format='0%'
    wv=ws.cell(row=rr,column=6,value=f"=IF(AND(ISNUMBER(D{rr}),ISNUMBER(E{rr})),D{rr}*E{rr},0)")
    wv.number_format='#,##0'; wv.border=B; wv.font=BOLD; wv.fill=GFILL
for col,w in zip("ABCDEF",[10,46,18,14,12,14]): ws.column_dimensions[col].width=w
# engagement flow
fr=11+len(pipe)+1
ws.cell(row=fr,column=1,value="ENGAGEMENT FLOW").font=SEC
ws.cell(row=fr,column=2,value="Signal → Fit → Map → Build → Sustain → Prove").font=GLD
flow=[("Signal","A known trigger: turnover, a stalled project, a funding deadline, a capacity gap."),
 ("Fit","Confirm authority, budget, and that it's a known problem we solve."),
 ("Map","A paid Stewardship Map — the entry move."),
 ("Build","Deliver the project, funding, permit, implementation, or capacity work."),
 ("Sustain","Transfer ownership; optional Implementation or Capacity Support."),
 ("Prove","Capture the evidence — which becomes the next Signal.")]
for i,(st,d) in enumerate(flow):
    rr=fr+1+i; a=ws.cell(row=rr,column=1,value=st); a.font=Font(bold=True,color=GOLD); a.fill=GFILL; a.border=B
    c=ws.cell(row=rr,column=2,value=d); c.font=BODY; c.alignment=WRAP; c.border=B; ws.merge_cells(f"B{rr}:F{rr}")

# ===== NEXT 12 MONTHS =====
ws=newtab("Next 12 Months","What's Realistic in the Next 12 Months","Modest and real. We're two people who do the work; the plan should respect that.")
lines(ws,4,[
 "The honest goal for year one is not scale. It's a small number of real, paid engagements with people who already trust us — and a repeatable way to do them.",
 "",
])
htable(ws,6,["Move","What it looks like","Why"],
 [
 ["Sell a few Maps","2–4 paid Stewardship Maps with warm Tier-1 relationships.","Low-risk entry; surfaces the next, bigger engagement; builds reference proof."],
 ["Convert 1–2","Turn a Map into Project / Funding / Permit & Bridge work.","This is where the real revenue is, and where Permit & Bridge proves out."],
 ["Land baseline retainers","1–2 monthly Implementation or Capacity engagements.","Predictable monthly revenue that smooths the project lumpiness."],
 ["Build the templates as we go","Every engagement leaves a reusable asset (Transfer tab).","Each job gets cheaper to deliver; we compound instead of resetting."],
 ],[24,46,40])
ws.cell(row=12,column=1,value="Three questions to keep answering:  (1) Which 5 targets are we actually pursuing this quarter?  (2) What is the first paid engagement we want to sell?  (3) Which transfer asset gets created every time we do it?").font=Font(italic=True,size=11,color=GREEN)
ws.cell(row=12,column=1).alignment=WRAP; ws.merge_cells("A12:C13")

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"final_deck","PublicLogic - Capabilities Workbook (v2.0 2026-06).xlsx")
os.makedirs(os.path.dirname(out),exist_ok=True)
wb.save(out); print("wrote",out,"| tabs:",wb.sheetnames)
