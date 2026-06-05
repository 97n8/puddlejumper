#!/usr/bin/env python3
"""
Builds MichiganLTC_Network_Financial_Model_FINAL.xlsx — the reconciled Michigan LTC
financial model. Implements the headline reconciliation: a DEFENSIBLE BASE CASE
(fuel + carbon black, market-grounded) with RIN + Section 45Z shown SEPARATELY as
contingent incentive upside, plus full per-site -> 3-site rollups and the two raises.

Why this exists: the signed WasteWerx Model 3 headline of ~$35.58M revenue / ~$30.2M
EBITDA per site depends on ~$20.4M of RIN + 45Z credits that are NOT secured. This model
makes the opportunity big WITHOUT making the investor headline depend on unsecured credits.

All scenario revenue/EBITDA and rollups are LIVE formulas off the Inputs sheet. Figures
trace to the canonical financial workbook / signed pro forma; [CONFIRM] items flagged.
Not legal/accounting/securities/tax advice.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F4E5F")
BASEFILL = PatternFill("solid", fgColor="E2EFDA")   # green = base/defensible
UPFILL = PatternFill("solid", fgColor="FCE4D6")     # orange = contingent upside
INFILL = PatternFill("solid", fgColor="FFF2CC")     # yellow = editable input
NOTE = Font(italic=True, size=9, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
M = '#,##0'; M0='$#,##0'

def H(ws, row, cols, start=1):
    for i,c in enumerate(cols):
        x=ws.cell(row=row,column=start+i,value=c); x.font=HDR; x.fill=HFILL; x.border=BORDER; x.alignment=CTR

def T(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(bold=True,size=14,color="1F4E5F")
    if s: ws["A2"]=s; ws["A2"].font=NOTE

wb=openpyxl.Workbook()

# ---------------- README ----------------
ws=wb.active; ws.title="README"
T(ws,"Michigan LTC Network — Reconciled Financial Model (FINAL working)",
  "Base case (fuel + carbon black) anchors the headline; RIN + Section 45Z shown separately as contingent upside. "
  "Yellow = editable inputs. Green = defensible base. Orange = contingent. Date 2026-06-03. Not financial/legal/tax advice.")
for r,ln in enumerate([
 "","THE RECONCILIATION (why this model)",
 "  The signed WasteWerx Model 3 headline (~$35.58M rev / ~$30.2M EBITDA per site) depends on ~$20.4M of",
 "  RIN + Section 45Z credits that are NOT secured (registration, eligibility, monetization, counsel/tax review",
 "  all required). This model anchors a DEFENSIBLE base case and shows credits separately as contingent upside.",
 "","HOW TO READ",
 "  Per_Site  - line-by-line revenue build-up + 4 scenarios (conservative base, upgraded base, signed headline,",
 "              full stack) with EBITDA each, and the per-site capital bridge.",
 "  Network   - 3-site rollup.",
 "  Raises    - $15M/site project raise + use-of-funds reconciliation; separate $2.5M company raise.",
 "  Caveats   - what must be confirmed / reviewed before any investor-facing use.",
 "","THE SAFE INVESTOR HEADLINE",
 "  Lead with BASE EBITDA (no credits). Show RIN/45Z as labeled contingent upside with a footnote. Do NOT make",
 "  the deck depend on credits that are not secured. This keeps the opportunity big and survives diligence.",
 "","GUARDRAILS",
 "  * Carbon black is INCLUDED in this base case as a real saleable product (the signed model excluded it). [CONFIRM price/offtake]",
 "  * SAF/kerosene classification, RFS pathway (D4/D7 RIN), and 45Z eligibility require registration + tax counsel.",
 "  * Securities offering: not an offer; counsel review required; PublicLogic = governance, NOT placement agent; fee fixed/non-contingent.",
 "  * Figures trace to the canonical workbook / signed pro forma; verify before circulation.",
],start=4):
    c=ws.cell(row=r,column=1,value=ln)
    if ln.strip() in ("THE RECONCILIATION (why this model)","HOW TO READ","THE SAFE INVESTOR HEADLINE","GUARDRAILS"):
        c.font=Font(bold=True,size=11,color="1F4E5F")
ws.column_dimensions["A"].width=112

# ---------------- INPUTS ----------------
wi=wb.create_sheet("Inputs")
T(wi,"Inputs — editable assumptions (per site)","Edit yellow cells; everything else recalculates. [CONFIRM] = verify before investor use.")
wi.append([]); wi.append(["Key","Assumption","Value","Unit","Source / note"]); H(wi,wi.max_row,["Key","Assumption","Value","Unit","Source / note"])
inputs=[
 ("tpd","Feedstock throughput",100,"MT/day","WasteWerx model"),
 ("days","Operating days",300,"days/yr","WasteWerx model"),
 ("oil_yield","Oil/fuel yield",0.45,"of feedstock","WasteWerx model"),
 ("cb_yield","Carbon black yield",0.45,"of feedstock","WasteWerx model"),
 ("oil_price","Pyrolysis oil price (conservative/market)",500,"$/MT","market-grounded base [CONFIRM contract]"),
 ("cb_price","Carbon black price",750,"$/MT","[CONFIRM market price + offtake]"),
 ("fuel_base","WasteWerx fuel revenue (distilled RD+SAF, base)",22181680,"$/yr","CONFIRMED: $35,581,680 total minus ~$13.4M credits (13_Financial_Reconciliation)"),
 ("credits","RIN + Section 45Z (CONTINGENT)",13400000,"$/yr","CONFIRMED contingent: '$13.4M+/yr requires EPA RFS registration — NOT guaranteed'"),
 # opex (bottom-up) — labor from the uploaded WasteWerx staffing pro forma; rest editable
 ("labor_fl","Site labor — total loaded, 62 FTE (FL basis)",5522400,"$/yr","WasteWerx Staffing Pro Forma (uploaded); FL market"),
 ("mi_adj","Michigan labor adjustment factor",1.05,"x","FL->MI cost-of-living [Robert to set]"),
 ("xtrain","Cross-training / headcount efficiency factor",1.00,"x","WasteWerx (V.Tizio 5/27): staffing is WORST-CASE; cross-training reduces it [set <1.0]"),
 ("feedstock","Feedstock cost (tires) — co-location may reduce",3720000,"$/yr","deck Operating Savings tab; co-loc w/ ERR/Geocycle may cut this — BIGGEST swing"),
 ("royalty_gal","WasteWerx production royalty rate",0.60,"$/gal","WasteWerx Model 3 (executed)"),
 ("annual_gal","Annual fuel output (85% util)",4467600,"gal/yr","WasteWerx Model 3"),
 ("monitor_mo","WasteWerx monitoring fee",35000,"$/mo","WasteWerx Model 3 (CPI-indexed)"),
 ("utilities","Utilities (electric+water+fees)",96960,"$/yr","deck Operating Savings tab"),
 ("maint","Maintenance + parts",42000,"$/yr","deck Operating Savings tab [low for plant scale — confirm]"),
 ("insurance","Insurance",78000,"$/yr","deck Operating Savings tab"),
 ("property_tax","Property tax",168000,"$/yr","deck Operating Savings tab"),
 ("security","Security",360000,"$/yr","deck Operating Savings tab"),
 ("legal_acct","Legal + accounting (with PublicLogic)",750000,"$/yr","deck Operating Savings tab (revised)"),
 ("pl_gov","PublicLogic governance retainer",120000,"$/yr","deck ($10k/mo)"),
 # capital
 ("raise_site","Project raise per site",15000000,"$","sponsor-directed headline"),
 ("build","Use of funds: build / core plant",8000000,"$","rough verbal"),
 ("addl","Use of funds: added construction",1000000,"$","rough verbal"),
 ("preproc","Use of funds: pre-processing/pre-treatment equip",2000000,"$","rough verbal"),
 ("conting","Use of funds: contingency/fees/payroll/training",5000000,"$","rough verbal"),
 ("ww_license","WasteWerx license fee (Yr1, net of design credit)",5450000,"$","signed Model 3"),
 ("ww_design","WasteWerx design fee (credited)",50000,"$","signed Model 3"),
 ("ww_monitor","WasteWerx Yr1 monitoring fees",175000,"$","signed Model 3"),
 ("ww_royalty","WasteWerx Yr1 production royalty",995328,"$","signed Model 3"),
 ("cust_capex","Customer-side capex (site/utilities/permits/WC)",952500,"$","signed Model 3"),
 ("company_raise","Company-level raise (separate)",2500000,"$","sponsor-directed"),
 ("sites","Number of sites in network",3,"#","Michigan corridor"),
]
k={}; r0=wi.max_row+1
for i,(key,lab,val,unit,note) in enumerate(inputs):
    rr=r0+i
    wi.cell(row=rr,column=1,value=key).font=Font(bold=True,color="888888",size=9)
    wi.cell(row=rr,column=2,value=lab)
    c=wi.cell(row=rr,column=3,value=val); c.fill=INFILL; c.border=BORDER
    c.number_format='0.00' if isinstance(val,float) else (M if isinstance(val,(int,float)) and val>=1000 else '0')
    wi.cell(row=rr,column=4,value=unit); wi.cell(row=rr,column=5,value=note).font=NOTE
    k[key]=f"Inputs!$C${rr}"
for col,w in zip("ABCDE",[12,46,14,12,52]): wi.column_dimensions[col].width=w

def K(key): return k[key]

# ---------------- OPEX BUILD-UP ----------------
wo=wb.create_sheet("Opex_BuildUp")
T(wo,"Annual operating cost — bottom-up (per site, steady state)",
  "KEY FINDING: site labor ALONE exceeds the signed model's entire implied opex. Green = known, orange = estimate to confirm.")
H(wo,4,["Cost line","Amount $/yr","Status"])
opx=[
 ("Site labor — loaded (62 FTE worst-case x MI-adj x cross-training)", f"={K('labor_fl')}*{K('mi_adj')}*{K('xtrain')}","WORST-CASE per WasteWerx (V.Tizio); benefits/MI rates need actuals [Robert]"),
 ("WasteWerx royalty (steady-state, $0.60/gal x output)", f"={K('royalty_gal')}*{K('annual_gal')}","WasteWerx Model 3 — capital-bridge uses Yr1 partial"),
 ("WasteWerx monitoring ($35k/mo x 12)", f"={K('monitor_mo')}*12","WasteWerx Model 3"),
 ("Feedstock cost (tires)", f"={K('feedstock')}","deck; co-loc w/ ERR/Geocycle may reduce — BIGGEST swing"),
 ("Utilities (electric+water+fees)", f"={K('utilities')}","deck Operating Savings tab"),
 ("Maintenance + parts", f"={K('maint')}","deck [low for scale — confirm]"),
 ("Insurance", f"={K('insurance')}","deck Operating Savings tab"),
 ("Property tax", f"={K('property_tax')}","deck Operating Savings tab"),
 ("Security", f"={K('security')}","deck Operating Savings tab"),
 ("Legal + accounting (with PublicLogic)", f"={K('legal_acct')}","deck Operating Savings tab (revised)"),
 ("PublicLogic governance retainer", f"={K('pl_gov')}","deck ($10k/mo)"),
]
r=5
for lab,f,st in opx:
    wo.cell(row=r,column=1,value=lab).alignment=WRAP
    c=wo.cell(row=r,column=2,value=f); c.number_format=M
    c.fill=UPFILL if "swing" in st else BASEFILL
    wo.cell(row=r,column=3,value=st).alignment=WRAP
    for cc in range(1,4): wo.cell(row=r,column=cc).border=BORDER
    r+=1
last=r-1
wo.cell(row=r,column=1,value="TOTAL ANNUAL OPEX (bottom-up)").font=BOLD
tc=wo.cell(row=r,column=2,value=f"=SUM(B5:B{last})"); tc.number_format=M; tc.font=BOLD; tc.fill=BASEFILL
OPEX_ROW=r
implied = 35581680-30200531  # signed-model implied opex
wo.cell(row=r+2,column=1,value="Signed model's IMPLIED opex (rev - EBITDA)").alignment=WRAP
wo.cell(row=r+2,column=2,value=implied).number_format=M
wo.cell(row=r+3,column=1,value="Understatement baked into signed EBITDA (bottom-up - implied)").font=BOLD
us=wo.cell(row=r+3,column=2,value=f"=B{OPEX_ROW}-B{r+2}"); us.number_format=M; us.fill=UPFILL; us.font=BOLD
wo.cell(row=r+5,column=1,value="Labor ALONE (MI-adj x cross-training)").font=BOLD
la=wo.cell(row=r+5,column=2,value=f"={K('labor_fl')}*{K('mi_adj')}*{K('xtrain')}"); la.number_format=M; la.font=BOLD
wo.cell(row=r+6,column=1,value="-> at worst-case staffing, labor alone exceeds the signed model's entire implied opex. Signed EBITDA overstates margin.").font=Font(bold=True,color="7A1F1F")
wo.cell(row=r+8,column=1,value="WasteWerx (V. Tizio, 5/27/26): staffing is intentional worst-case 'overkill'; cross-trained roles will reduce headcount;").font=NOTE
wo.cell(row=r+9,column=1,value="health insurance & benefits need actuals; salaries are FL market — Robert to adjust to Michigan pay rates. Flex via xtrain + mi_adj inputs.").font=NOTE
for col,w in zip("ABC",[54,16,46]): wo.column_dimensions[col].width=w
OPEX_CELL=f"Opex_BuildUp!$B${OPEX_ROW}"

# ---------------- PER SITE ----------------
ws=wb.create_sheet("Per_Site")
T(ws,"Per-site revenue build-up, scenarios & capital bridge",
  "Line-by-line build-up (the answer to 'show me $6.75M fuel -> $35.58M, line by line'). Base = green, contingent = orange.")
# build-up
H(ws,4,["Revenue line","Amount $/yr","Category","Status / caveat"])
oil_q=f"({K('tpd')}*{K('days')}*{K('oil_yield')})"
cb_q=f"({K('tpd')}*{K('days')}*{K('cb_yield')})"
rows=[
 ("WasteWerx fuel (distilled RD+SAF)", f"={K('fuel_base')}","BASE","CONFIRMED: $35.58M total minus ~$13.4M credits (13_Financial_Reconciliation)"),
 ("Carbon black (deck-era; additive?)", f"={cb_q}*{K('cb_price')}","ADD'L (contingent)","$0 in executed WasteWerx pro forma — confirm if additive + offtake before use"),
 ("RIN + Section 45Z (credits)", f"={K('credits')}","CONTINGENT UPSIDE","CONFIRMED contingent: ~$13.4M+; requires EPA RFS registration — NOT guaranteed"),
]
r=5
for lab,f,cat,st in rows:
    ws.cell(row=r,column=1,value=lab).alignment=WRAP
    c=ws.cell(row=r,column=2,value=f); c.number_format=M
    c.fill=UPFILL if "CONTINGENT" in cat else BASEFILL
    ws.cell(row=r,column=3,value=cat).alignment=CTR
    ws.cell(row=r,column=4,value=st).alignment=WRAP
    for cc in range(1,5): ws.cell(row=r,column=cc).border=BORDER
    r+=1
# rows map: 5 oil,6 cb,7 rd,8 saf,9 d4,10 d7,11 z_rd,12 z_saf
ws.cell(row=14,column=1,value="SCENARIO SUMMARY (per site)").font=Font(bold=True,size=11,color="1F4E5F")
H(ws,15,["Scenario","Revenue $/yr","Opex $/yr","EBITDA $/yr","Credits in headline?"])
scen=[
 ("Base — WasteWerx fuel only (no credits, no CB)","=B5","BASE","No — most defensible"),
 ("Base + carbon black (if additive)","=B5+B6","BASE","No (CB itself contingent)"),
 ("WasteWerx confirmed headline (fuel + credits, no CB)","=B5+B7","SIGNED","YES — ~$13.4M credit-dependent"),
 ("Full stack (fuel + carbon black + credits)","=B5+B6+B7","FULL","YES"),
]
r=16
for lab,rev,tag,creds in scen:
    ws.cell(row=r,column=1,value=lab).alignment=WRAP
    rc=ws.cell(row=r,column=2,value=rev); rc.number_format=M
    oc=ws.cell(row=r,column=3,value=f"={OPEX_CELL}"); oc.number_format=M
    ec=ws.cell(row=r,column=4,value=f"=B{r}-C{r}"); ec.number_format=M; ec.font=BOLD
    ec.fill=BASEFILL if tag=="BASE" else UPFILL
    ws.cell(row=r,column=5,value=creds).alignment=CTR
    for cc in range(1,6): ws.cell(row=r,column=cc).border=BORDER
    r+=1
# credit-dependency callout
ws.cell(row=21,column=1,value="Contingent credit value at risk in headline (RIN + 45Z)").font=BOLD
cd=ws.cell(row=21,column=2,value="=B7"); cd.number_format=M; cd.fill=UPFILL; cd.font=BOLD
ws.cell(row=21,column=3,value="<- this much of the signed headline is NOT secured").font=NOTE

# capital bridge
ws.cell(row=23,column=1,value="PER-SITE CAPITAL BRIDGE").font=Font(bold=True,size=11,color="1F4E5F")
cap=[
 ("Project raise per site","=" + K('raise_site').split('!')[1].replace('$','')),  # placeholder fixed below
]
ws.cell(row=24,column=1,value="Project raise per site")
ws.cell(row=24,column=2,value=f"={K('raise_site')}").number_format=M
ws.cell(row=25,column=1,value="WasteWerx Year 1 (license+design+monitoring+royalty)")
ws.cell(row=25,column=2,value=f"={K('ww_license')}+{K('ww_design')}+{K('ww_monitor')}+{K('ww_royalty')}").number_format=M
ws.cell(row=26,column=1,value="Customer-side capex")
ws.cell(row=26,column=2,value=f"={K('cust_capex')}").number_format=M
ws.cell(row=27,column=1,value="Total Year 1 cash out").font=BOLD
ws.cell(row=27,column=2,value="=B25+B26").number_format=M; ws.cell(row=27,column=2).font=BOLD
ws.cell(row=28,column=1,value="Residual project envelope (land, civil, integration, contingency, payroll, reserves)").alignment=WRAP
rcell=ws.cell(row=28,column=2,value="=B24-B27"); rcell.number_format=M; rcell.fill=BASEFILL; rcell.font=BOLD
# use of funds reconciliation
ws.cell(row=30,column=1,value="Use-of-funds (sponsor rough) vs raise — RECONCILE").font=BOLD
ws.cell(row=31,column=1,value="Build + added construction + pre-processing + contingency")
ws.cell(row=31,column=2,value=f"={K('build')}+{K('addl')}+{K('preproc')}+{K('conting')}").number_format=M
ws.cell(row=32,column=1,value="Reconciliation gap vs $15M raise (negative = over)")
g=ws.cell(row=32,column=2,value=f"={K('raise_site')}-B31"); g.number_format=M; g.fill=UPFILL
ws.cell(row=32,column=3,value="[CONFIRM which line absorbs the ~$1M]").font=NOTE
for col,w in zip("ABCDE",[52,16,14,14,8]): ws.column_dimensions[col].width=w

# ---------------- NETWORK ----------------
wn=wb.create_sheet("Network")
T(wn,"Three-site network rollup (live x number of sites)","Base anchors; credits flagged. Company raise kept separate.")
H(wn,4,["Metric","Per site","x Sites","Network total","Note"])
net=[
 ("Base EBITDA — Conservative","='Per_Site'!D16",f"={K('sites')}","BASE — defensible"),
 ("Base EBITDA — Upgraded","='Per_Site'!D17",f"={K('sites')}","BASE — needs distillation+offtake"),
 ("Signed-headline EBITDA (credit-dependent)","='Per_Site'!D18",f"={K('sites')}","CONTINGENT — ~$20.4M/site credits"),
 ("Signed-headline revenue","='Per_Site'!B18",f"={K('sites')}","CONTINGENT"),
 ("Contingent credit value at risk","='Per_Site'!B21",f"={K('sites')}","NOT secured"),
 ("Year 1 cash out","='Per_Site'!B27",f"={K('sites')}","capital"),
 ("Project raises (sites x $15M)","='Per_Site'!B24",f"={K('sites')}","excludes company raise"),
]
r=5
for lab,per,mult,note in net:
    wn.cell(row=r,column=1,value=lab).alignment=WRAP
    wn.cell(row=r,column=2,value=per).number_format=M
    wn.cell(row=r,column=3,value=mult).alignment=CTR
    tc=wn.cell(row=r,column=4,value=f"=B{r}*C{r}"); tc.number_format=M; tc.font=BOLD
    tc.fill=UPFILL if "CONTINGENT" in note or "NOT secured" in note else BASEFILL
    wn.cell(row=r,column=5,value=note).alignment=WRAP
    for cc in range(1,6): wn.cell(row=r,column=cc).border=BORDER
    r+=1
for col,w in zip("ABCDE",[40,16,9,16,34]): wn.column_dimensions[col].width=w

# ---------------- RAISES ----------------
wr=wb.create_sheet("Raises")
T(wr,"The two raises — kept distinct","Site capital != company capital. Never blend.")
H(wr,4,["Item","Amount","Note"])
raises=[
 ("Project raise per site",f"={K('raise_site')}","site deployment"),
 ("Project raises — 3 sites",f"={K('raise_site')}*{K('sites')}","excludes company raise"),
 ("Company-level raise (separate)",f"={K('company_raise')}","working capital, payroll, PublicLogic gov fees, team buildout [CONFIRM split]"),
 ("Combined discussion envelope",f"={K('raise_site')}*{K('sites')}+{K('company_raise')}","present asks separately"),
]
r=5
for lab,f,note in raises:
    wr.cell(row=r,column=1,value=lab)
    c=wr.cell(row=r,column=2,value=f); c.number_format=M; c.font=BOLD
    wr.cell(row=r,column=3,value=note).alignment=WRAP
    for cc in range(1,4): wr.cell(row=r,column=cc).border=BORDER
    r+=1
for col,w in zip("ABC",[34,18,60]): wr.column_dimensions[col].width=w

# ---------------- CAVEATS ----------------
wc=wb.create_sheet("Caveats")
T(wc,"Caveats & confirm-before-send","Every contingent assumption and open item. Investor headline must lead with the base case.")
items=[
 "HEADLINE DISCIPLINE",
 "  * Investor-facing headline = BASE EBITDA (fuel + carbon black). Show RIN + 45Z as separately-labeled contingent upside.",
 "  * Signed-model EBITDA (~$30.2M/site) is ~$20.4M dependent on RIN + 45Z. Do NOT anchor the deck on it unproven.",
 "",
 "CONTINGENT / REGULATORY",
 "  * D4/D7 RIN: requires EPA RFS pathway registration; volumes & prices [CONFIRM].",
 "  * Section 45Z (RD + SAF): requires eligibility determination + tax counsel; rate assumption [CONFIRM].",
 "  * SAF/kerosene product classification [CONFIRM] — drives both fuel revenue and 45Z.",
 "  * Carbon black: included in base as a real product; price + offtake [CONFIRM]. (Signed model excluded it.)",
 "  * OPEX IS NOW SOURCED (bottom-up ~$14.2M/site) from the WasteWerx staffing pro forma + deck Operating Savings tab +",
 "    WasteWerx Model 3 terms — vs the signed model's implied $5.38M. Understatement ~$8.85M.",
 "  * THE $30.2M EBITDA IS WASTEWERX-SCOPE: it sits on WasteWerx's confirmed $5.38M opex (fees + minimal ops) which OMITS the",
 "    customer's full 62-FTE site labor (~$5.5M) and feedstock (~$1-3.7M). On the customer's all-in opex (~$14.2M) the SAME",
 "    $35.58M revenue yields ~$21.35M EBITDA, not $30.2M. (The deck's own reconciled model shows ~$11.2M expenses / ~$5.66M net income.)",
 "  * Royalty corrected to STEADY-STATE: $0.60/gal x 4,467,600 gal = ~$2.68M/yr ($995k in the capital bridge is Yr1 partial). Monitoring $420k/yr.",
 "  * FEEDSTOCK is a CONFIRMED swing: deck $200/ton landed = $3.72M vs Dort Hwy agreement $25/ton tipping = ~$1M. Resolve with Robert. [CONFIRM]",
 "  * REVENUE SPLIT NOT IN CANONICAL WORKBOOK: executed pro forma shows only blended 'Fuel+RIN+§45Z' = $35.58M; the RD/SAF/RIN/45Z",
 "    line items trace to the investor doc, not a source — need the underlying WasteWerx Model 3 to verify the split.",
 "  * CARBON BLACK = $0 in executed WasteWerx model (Disc.#14) but $10.1M in the deck; base case includes it on the DECK assumption — confirm.",
 "    Labor is worst-case (V. Tizio 5/27): set xtrain (cross-training) + mi_adj (Michigan); benefits actuals pending.",
 "",
 "CAPITAL",
 "  * Sponsor use-of-funds (~$16M) exceeds the $15M/site raise by ~$1M — confirm which line absorbs it.",
 "  * $2.5M company raise stays separate from site SPV totals.",
 "  * WasteWerx is a LICENSE, not a sale; title stays with WasteWerx; license fee depreciation differs; transfer/termination terms matter.",
 "",
 "GOVERNANCE / LEGAL",
 "  * Securities: not an offer; securities counsel review required before any circulation.",
 "  * PublicLogic = governance/document partner, NOT placement agent/broker-dealer; fee FIXED & NON-CONTINGENT.",
 "  * Community-benefit %: source files say 10%; sponsor flags a conflict — do not publish a number until reconciled.",
 "  * Final site list (Coldwater Phase 2 vs Lincoln) [CONFIRM]; parent-holdco legal name [CONFIRM].",
 "  * National Salvage: 'largest treated-wood recycler in the U.S.' unless stronger evidence (a verbal 'in the world' is unverified).",
 "",
 "Not financial, accounting, securities, or tax advice. Figures trace to the canonical workbook / signed pro forma; verify before use.",
]
r=4
for ln in items:
    c=wc.cell(row=r,column=1,value=ln)
    c.font=Font(bold=True,size=11,color="1F4E5F") if ln.strip() in ("HEADLINE DISCIPLINE","CONTINGENT / REGULATORY","CAPITAL","GOVERNANCE / LEGAL") else NOTE
    r+=1
wc.column_dimensions["A"].width=116

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"MichiganLTC_Network_Financial_Model_FINAL.xlsx")
wb.save(out); print("wrote",out)
